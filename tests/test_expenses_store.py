"""Unit tests for the Expenses SQLite store. Headline guards: lossless migration from BOTH sources
(xlsx receipts + results/expenses.json invoices/contracts), the dual projection (receipts→xlsx,
all→json), store ids identical to server._expense_row_id (frontend rows don't move), and the
store-backed dedup that replaces _seen.json/_receipt_hashes.json (no re-classification post-migration)."""
import hashlib
import json
import os
import sys
import tempfile
from pathlib import Path

os.environ.setdefault("DATA_DIR", tempfile.mkdtemp(prefix="ceo_exp_test_"))
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from src.modules import expenses_store as es  # noqa: E402


def _server_row_id(row: dict) -> str:
    """Replica of server._expense_row_id (capitalised Excel keys) — the id the frontend uses today."""
    key = (str(row.get("Msg_ID") or "") + "::" + str(row.get("Att_ID") or ""))
    if not key.strip(":"):
        key = json.dumps({k: str(v) for k, v in row.items()}, sort_keys=True)
    return hashlib.sha1(key.encode()).hexdigest()[:16]


def _seed_xlsx(d, rows):
    from openpyxl import Workbook
    xp = Path(d) / "expenses" / "expenses_master.xlsx"
    xp.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook(); ws = wb.active; ws.title = "Expenses"
    ws.append(es._EXCEL_HEADERS)
    for r in rows:
        ws.append([r.get(h, "") for h in es._EXCEL_HEADERS])
    wb.save(xp)


def _seed_results(d, items):
    rp = Path(d) / "results" / "expenses.json"
    rp.parent.mkdir(parents=True, exist_ok=True)
    rp.write_text(json.dumps({"id": "expenses", "items": items, "count": len(items)}))


def _receipt_xlsx_row(**kw):
    base = {"Date": "2026-06-10", "Vendor": "Starbucks", "Amount": 12.5, "Currency": "CAD",
            "GST_HST": 1.5, "Net_Amount": 11.0, "Category": "Meals", "Attachment": "sb.pdf",
            "Email_Subject": "receipt", "From": "me@x.com", "Msg_ID": "m1", "Att_ID": "a1",
            "Processed_Date": "2026-06-11"}
    base.update(kw)
    return base


def _invoice_item(**kw):
    base = {"document_type": "invoice", "vendor": "Acme SaaS", "counterparty": "", "date": "2026-06-01",
            "due_date": "2026-06-30", "amount": 480, "currency": "USD", "gst_hst": None, "net_amount": None,
            "category": "Software", "subject": "June subscription", "confidence": "high",
            "attachment": "inv.pdf", "email_subject": "Invoice", "from": "billing@acme.com",
            "msg_id": "m2", "att_id": "a2", "processed_at": "2026-06-02"}
    base.update(kw)
    return base


def _contract_item(**kw):
    base = {"document_type": "contract", "vendor": "", "counterparty": "Globex", "date": "2026-05-01",
            "due_date": None, "amount": None, "currency": "", "category": "Legal", "subject": "MSA",
            "confidence": "high", "attachment": "msa.pdf", "email_subject": "Contract",
            "from": "legal@globex.com", "msg_id": "m3", "att_id": "a3", "processed_at": "2026-05-02"}
    base.update(kw)
    return base


# ── migration: both sources, lossless ────────────────────────────────────────

def test_migration_pulls_receipts_from_xlsx_and_invoices_from_json(tmp_path):
    _seed_xlsx(tmp_path, [_receipt_xlsx_row()])
    _seed_results(tmp_path, [_invoice_item(), _contract_item()])
    items = es.load_expenses(tmp_path)
    by_type = {it["document_type"]: it for it in items}
    assert set(by_type) == {"receipt", "invoice", "contract"}, f"got {[i['document_type'] for i in items]}"
    assert by_type["receipt"]["vendor"] == "Starbucks"
    assert by_type["invoice"]["due_date"] == "2026-06-30"
    assert by_type["contract"]["counterparty"] == "Globex"
    assert es.get_migration_status(tmp_path)["verdict"] == "lossless"


def test_migration_is_idempotent(tmp_path):
    _seed_xlsx(tmp_path, [_receipt_xlsx_row()])
    _seed_results(tmp_path, [_invoice_item()])
    n1 = len(es.load_expenses(tmp_path))
    n2 = len(es.load_expenses(tmp_path))   # second access must not re-import / duplicate
    assert n1 == n2 == 2


def test_store_id_matches_server_expense_row_id(tmp_path):
    """The store's id for a migrated receipt must equal server._expense_row_id of the Excel row, so
    existing frontend row ids don't change."""
    row = _receipt_xlsx_row()
    _seed_xlsx(tmp_path, [row])
    stored = es.load_expenses(tmp_path)[0]
    assert stored["id"] == _server_row_id(row)


# ── dual projection ──────────────────────────────────────────────────────────

def test_write_projection_receipts_to_xlsx_all_to_json(tmp_path):
    es.upsert_expense(tmp_path, es._row_from_excel(_receipt_xlsx_row()))
    es.upsert_expense(tmp_path, _invoice_item())
    es.upsert_expense(tmp_path, _contract_item())
    # xlsx = receipts only
    from openpyxl import load_workbook
    ws = load_workbook(tmp_path / "expenses" / "expenses_master.xlsx").active
    vendors = [r[1] for r in ws.iter_rows(min_row=2, values_only=True)]
    assert vendors == ["Starbucks"], f"xlsx should hold only the receipt, got {vendors}"
    # results/expenses.json = all three
    res = json.loads((tmp_path / "results" / "expenses.json").read_text())
    assert {it["document_type"] for it in res["items"]} == {"receipt", "invoice", "contract"}


def test_invoice_survives_and_loads_by_type(tmp_path):
    es.upsert_expense(tmp_path, _invoice_item())
    invoices = es.load_expenses(tmp_path, doc_type="invoice")
    assert len(invoices) == 1 and invoices[0]["vendor"] == "Acme SaaS"
    assert es.load_expenses(tmp_path, doc_type="receipt") == []


# ── edit / delete ────────────────────────────────────────────────────────────

def test_update_expense_fields_preserves_siblings(tmp_path):
    eid = es.upsert_expense(tmp_path, _invoice_item())
    es.update_expense_fields(tmp_path, eid, {"amount": 999})
    it = es.load_expenses(tmp_path)[0]
    assert it["amount"] == 999
    assert it["vendor"] == "Acme SaaS" and it["due_date"] == "2026-06-30"   # untouched


def test_delete_expense(tmp_path):
    eid = es.upsert_expense(tmp_path, _invoice_item())
    assert es.delete_expense(tmp_path, eid) is True
    assert es.load_expenses(tmp_path) == []


# ── dedup replaces _seen.json + _receipt_hashes.json ─────────────────────────

def test_seen_import_and_roundtrip(tmp_path):
    d = tmp_path / "expenses"; d.mkdir(parents=True)
    (d / "_seen.json").write_text(json.dumps({"m9::old.pdf": True}))
    (d / "_receipt_hashes.json").write_text(json.dumps({"abc123": "old.pdf"}))
    assert es.is_seen(tmp_path, "m9::old.pdf") is True          # migrated seen key
    assert es.is_seen(tmp_path, "sha256:abc123") is True        # migrated hash key
    assert es.is_seen(tmp_path, "m9::new.pdf") is False
    es.mark_seen(tmp_path, "m9::new.pdf")
    assert es.is_seen(tmp_path, "m9::new.pdf") is True
