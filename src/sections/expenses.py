"""
Expense capture section — scans email attachments for receipts and invoices.

Does NOT use the screener. Screener answers "should CEO read this email?".
This section answers "is this attachment a receipt?" — only Gemini looking at the
file content can answer that; email metadata cannot.

Two-layer dedup:
  Layer 1: _seen.json    — {msg_id::att_name: true}  (same attachment, idempotent)
  Layer 2: _receipt_hashes.json — {sha256hex: att_name}  (same file from different emails)

Data written per run:
  data_dir/results/expenses.json       — standard section result (this run's new items)
  data_dir/expenses/_seen.json         — cumulative seen log
  data_dir/expenses/_receipt_hashes.json — cumulative hash log
  data_dir/expenses/expenses_master.xlsx — cumulative receipt rows
"""
import hashlib
import json
import os
import re
import tempfile
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path

import requests

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
except ImportError:
    openpyxl = None

from src.graph import GraphClient
from src.ai import AIClient

_EXCEL_HEADERS = [
    "Date", "Vendor", "Amount", "Currency", "GST_HST", "Net_Amount",
    "Category", "Attachment", "Email_Subject", "From", "Msg_ID", "Att_ID", "Processed_Date",
]

_ALLOWED_MIME = {
    "application/pdf",
    "image/jpeg", "image/jpg", "image/png", "image/gif",
    "image/tiff", "image/webp", "image/bmp",
}
_ALLOWED_EXT = {".pdf", ".jpg", ".jpeg", ".png", ".gif", ".tiff", ".tif", ".webp", ".bmp"}

_EXTRACT_PROMPT = """Analyze this document and classify it. Return JSON with these fields:

"document_type": one of:
  - "receipt"   — a PAID expense (already paid; small purchase, meal, travel, transportation, retail receipt)
  - "invoice"   — a BILL TO PAY (vendor invoice the user owes, with due date; SaaS billing, services)
  - "contract"  — a legal AGREEMENT (MSA, NDA, SOW, employment, partnership, lease)
  - "other"     — not a business document, skip

If document_type is "other", return only: {"document_type": "other"}

Otherwise also include:
- "vendor": string (company or store name; for contracts use empty string)
- "counterparty": string (for contracts only — the other party in the agreement; for receipts/invoices use empty string)
- "date": string (YYYY-MM-DD; document date, invoice issue date, or contract effective date)
- "due_date": string or null (for invoices only — when payment is due, YYYY-MM-DD)
- "amount": number or null (total amount; null for contracts without explicit value)
- "currency": string (CAD / USD / EUR / etc., or empty if N/A)
- "gst_hst": number or null (receipts only)
- "net_amount": number or null (before tax, receipts only)
- "category": one of ["Travel", "Meals", "Software", "Services", "Equipment", "Utilities", "Legal", "Other"]
- "subject": string (short title — for contracts use the agreement name; for invoices the bill description)
- "confidence": "high" | "medium" | "low"

Respond with valid JSON only."""


def _load_json(path: Path) -> dict:
    if path.exists():
        try:
            return json.loads(path.read_text())
        except Exception:
            pass
    return {}


def _save_json(path: Path, data: dict) -> None:
    tmp = path.with_suffix(".tmp")
    tmp.write_text(json.dumps(data, indent=2, ensure_ascii=False))
    tmp.replace(path)


def _sender_addr(msg: dict) -> str:
    return ((msg.get("from") or {}).get("emailAddress") or {}).get("address", "")


def _get_attachments(graph: GraphClient, msg_id: str) -> list[dict]:
    try:
        data = graph.get(f"/me/messages/{msg_id}/attachments",
                         params={"$select": "id,name,contentType,size,isInline"})
        return data.get("value", [])
    except Exception:
        return []


def _init_excel(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    ws.append(_EXCEL_HEADERS)
    wb.save(path)


def _append_excel(path: Path, row: dict) -> None:
    if not path.exists():
        _init_excel(path)
    wb = load_workbook(path)
    ws = wb.active
    ws.append([
        row.get("date", ""),
        row.get("vendor", ""),
        row.get("amount", ""),
        row.get("currency", ""),
        row.get("gst_hst", ""),
        row.get("net_amount", ""),
        row.get("category", ""),
        row.get("attachment", ""),
        row.get("email_subject", ""),
        row.get("from", ""),
        row.get("msg_id", ""),
        row.get("att_id", ""),
        row.get("processed_at", ""),
    ])
    wb.save(path)


def _extract_document(ai: AIClient, file_bytes: bytes, filename: str) -> dict | None:
    """Upload file to Gemini Files API and classify + extract fields. Returns None on failure."""
    ext = Path(filename).suffix.lower() or ".pdf"
    mime_map = {
        ".pdf": "application/pdf",
        ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
        ".png": "image/png", ".gif": "image/gif",
        ".tiff": "image/tiff", ".tif": "image/tiff",
        ".webp": "image/webp", ".bmp": "image/bmp",
    }
    mime = mime_map.get(ext, "application/pdf")

    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as f:
        f.write(file_bytes)
        tmp_path = f.name

    try:
        from google.genai import types as genai_types

        uploaded = ai.client.files.upload(file=tmp_path)
        while uploaded.state.name == "PROCESSING":
            time.sleep(3)
            uploaded = ai.client.files.get(name=uploaded.name)

        for attempt in range(3):
            try:
                response = ai.client.models.generate_content(
                    model=ai.model,
                    contents=[
                        genai_types.Part.from_uri(file_uri=uploaded.uri, mime_type=mime),
                        _EXTRACT_PROMPT,
                    ],
                )
                raw = (response.text or "").strip()
                if raw.startswith("```"):
                    raw = raw.split("\n", 1)[-1].rsplit("```", 1)[0]
                match = re.search(r"\{.*\}", raw, re.DOTALL)
                if match:
                    return json.loads(match.group())
                if attempt < 2:
                    time.sleep(3)
            except Exception as e:
                if "429" in str(e) and attempt < 2:
                    time.sleep(10 * (2 ** attempt))
                elif attempt < 2:
                    time.sleep(5)
                else:
                    break
    except Exception:
        pass
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass

    return None


def run(
    graph: GraphClient,
    ai: AIClient,
    data_dir: Path,
    days: int = 7,
    progress=None,
) -> dict:
    """
    Scan email attachments for receipts and invoices.

    Returns standard section result dict.
    """
    if openpyxl is None:
        return {
            "id": "expenses",
            "status": "not_run",
            "last_run": None,
            "items": [],
            "count": 0,
            "empty": True,
            "error": "openpyxl not installed",
        }

    def log(msg: str):
        if progress:
            progress(msg)

    data_dir = Path(data_dir)
    expenses_dir = data_dir / "expenses"
    expenses_dir.mkdir(parents=True, exist_ok=True)
    results_dir = data_dir / "results"
    results_dir.mkdir(parents=True, exist_ok=True)

    seen_path = expenses_dir / "_seen.json"
    hashes_path = expenses_dir / "_receipt_hashes.json"
    excel_path = expenses_dir / "expenses_master.xlsx"

    seen: dict = _load_json(seen_path)
    receipt_hashes: dict = _load_json(hashes_path)

    since = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%dT%H:%M:%SZ")
    log(f"Fetching emails with attachments (last {days} days)...")

    try:
        messages = graph.get_messages(
            top=200,
            filter=f"receivedDateTime ge {since}",
            orderby=None,
        )
    except Exception as e:
        log(f"Fetch failed: {e}")
        messages = []

    log(f"Scanning {len(messages)} recent emails for receipts")

    new_items: list[dict] = []

    for msg in messages:
        msg_id = msg.get("id", "")
        subject = msg.get("subject") or "(no subject)"
        sender = _sender_addr(msg)

        has_att    = msg.get("hasAttachments", False)
        body_html  = msg.get("body", {}).get("content", "")
        has_inline = "cid:" in body_html or ("<img" in body_html and "src=" in body_html)
        if not has_att and not has_inline:
            continue

        attachments = _get_attachments(graph, msg_id)
        for att in attachments:
            att_id = att.get("id", "")
            att_name = att.get("name") or "attachment"
            content_type = (att.get("contentType") or "").lower().split(";")[0].strip()

            # Layer 1: seen check
            seen_key = f"{msg_id}::{att_name}"
            if seen_key in seen:
                continue

            # Extension/MIME filter
            ext = Path(att_name).suffix.lower()
            if content_type not in _ALLOWED_MIME and ext not in _ALLOWED_EXT:
                seen[seen_key] = True
                continue

            # Get file bytes — either inline contentBytes or download
            file_bytes_b64 = att.get("contentBytes")
            if file_bytes_b64:
                import base64
                try:
                    file_bytes = base64.b64decode(file_bytes_b64)
                except Exception:
                    seen[seen_key] = True
                    continue
            else:
                try:
                    file_bytes = graph.download(f"/me/messages/{msg_id}/attachments/{att_id}/$value")
                except Exception:
                    seen[seen_key] = True
                    continue

            if not file_bytes:
                seen[seen_key] = True
                continue

            # Layer 2: hash check
            file_hash = hashlib.sha256(file_bytes).hexdigest()
            if file_hash in receipt_hashes:
                seen[seen_key] = True
                continue

            log(f"  Analyzing: {att_name} (from: {sender})")

            # Gemini extraction + classification
            extracted = _extract_document(ai, file_bytes, att_name)
            seen[seen_key] = True

            if not extracted:
                continue
            doc_type = extracted.get("document_type", "other")
            if doc_type not in ("receipt", "invoice", "contract"):
                continue

            receipt_hashes[file_hash] = att_name

            item = {
                "document_type": doc_type,
                "vendor":        extracted.get("vendor", ""),
                "counterparty":  extracted.get("counterparty", ""),
                "date":          extracted.get("date", ""),
                "due_date":      extracted.get("due_date"),
                "amount":        extracted.get("amount", ""),
                "currency":      extracted.get("currency", ""),
                "gst_hst":       extracted.get("gst_hst"),
                "net_amount":    extracted.get("net_amount"),
                "category":      extracted.get("category", "Other"),
                "subject":       extracted.get("subject", ""),
                "confidence":    extracted.get("confidence", ""),
                "attachment":    att_name,
                "email_subject": subject,
                "from":          sender,
                "msg_id":        msg_id,
                "att_id":        att_id,
                "processed_at":  datetime.now().strftime("%Y-%m-%d"),
            }
            new_items.append(item)

            # Only receipts go to the reimbursement Excel
            if doc_type == "receipt":
                # Upload receipt to OneDrive so the user can view it later.
                # Failure is non-fatal — log and continue (we still want the Excel row).
                ext = Path(att_name).suffix.lower() or ".bin"
                mime_for_upload = {
                    ".pdf":  "application/pdf",
                    ".jpg":  "image/jpeg", ".jpeg": "image/jpeg",
                    ".png":  "image/png",  ".gif":  "image/gif",
                    ".webp": "image/webp", ".bmp":  "image/bmp",
                    ".tiff": "image/tiff", ".tif":  "image/tiff",
                }.get(ext, "application/octet-stream")
                try:
                    graph.upload_to_onedrive(
                        f"CEO Platform/Receipts/{att_name}",
                        file_bytes,
                        mime_for_upload,
                    )
                except Exception as up_err:
                    log(f"    OneDrive upload failed (kept Excel row): {up_err}")

                _append_excel(excel_path, item)
                log(f"    Receipt: {item['vendor']} {item['amount']} {item['currency']} [{item['category']}]")
            elif doc_type == "invoice":
                log(f"    Invoice: {item['vendor']} {item['amount']} {item['currency']} due {item.get('due_date') or '?'}")
            elif doc_type == "contract":
                log(f"    Contract: {item['counterparty']} — {item['subject']}")

    _save_json(seen_path, seen)
    _save_json(hashes_path, receipt_hashes)

    result = {
        "id":       "expenses",
        "status":   "fresh",
        "last_run": datetime.now(timezone.utc).isoformat(),
        "items":    new_items,
        "count":    len(new_items),
        "empty":    len(new_items) == 0,
    }

    result_path = results_dir / "expenses.json"
    _save_json(result_path, result)

    counts = {"receipt": 0, "invoice": 0, "contract": 0}
    for it in new_items:
        counts[it.get("document_type", "")] = counts.get(it.get("document_type", ""), 0) + 1
    log(f"Expenses done — {counts['receipt']} receipt(s), {counts['invoice']} invoice(s), {counts['contract']} contract(s)")
    return result
