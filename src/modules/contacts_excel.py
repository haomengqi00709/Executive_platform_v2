"""
contacts_excel — maintain a master Excel of Teams-ingested contacts in OneDrive.

Location: OneDrive 'Conferences/contacts_master.xlsx' (per-user).
Each row = one captured contact. The file is a read-only mirror of CRM data —
CRM is source of truth; this excel is for convenience (open in Excel app,
share with team, etc.). If the user edits this file directly, changes won't
flow back to CRM.

Columns:
  Added | Name | Email | Company | Role | Phone | LinkedIn | Source | Tags | Notes | Draft Link
"""
import os
import tempfile
from datetime import datetime, timezone

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
except ImportError:
    openpyxl = None


_HEADERS = ["Added", "Name", "Email", "Company", "Role", "Phone", "LinkedIn",
            "Source", "Tags", "Notes", "Draft Link"]
_ONEDRIVE_PATH = "Conferences/contacts_master.xlsx"
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _normalize_added(added_at: str) -> str:
    """Convert ISO datetime to 'YYYY-MM-DD HH:MM' for display."""
    if not added_at:
        return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")
    try:
        dt = datetime.fromisoformat(added_at.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return added_at[:16]


def append_contacts(graph, contacts: list) -> dict:
    """Append a batch of contacts to the OneDrive master excel.

    Creates the file with headers if it doesn't exist yet. Each contact dict
    should have the fields the CRM uses (email, name, company, role, phone,
    linkedin, source, tags, notes, draft_link, added_at).

    Returns {"appended": N, "error": str | None}.
    """
    if openpyxl is None:
        return {"appended": 0, "error": "openpyxl not installed"}
    if not contacts:
        return {"appended": 0, "error": None}

    # Try to download existing master file
    existing_bytes = None
    try:
        existing_bytes = graph.download(f"/me/drive/root:/{_ONEDRIVE_PATH}:/content")
    except Exception:
        pass  # File doesn't exist yet — will create new

    tmp_in = None
    if existing_bytes:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(existing_bytes)
            tmp_in = f.name
        try:
            wb = load_workbook(tmp_in)
            ws = wb.active
        except Exception:
            # Corrupted or empty — start fresh
            wb = Workbook()
            ws = wb.active
            ws.title = "Contacts"
            ws.append(_HEADERS)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Contacts"
        ws.append(_HEADERS)

    for c in contacts:
        tags_val = c.get("tags") or []
        if isinstance(tags_val, list):
            tags_str = ", ".join(tags_val)
        else:
            tags_str = str(tags_val)
        ws.append([
            _normalize_added(c.get("added_at", "")),
            c.get("name", ""),
            c.get("email", ""),
            c.get("company", ""),
            c.get("role", ""),
            c.get("phone", ""),
            c.get("linkedin", ""),
            c.get("source", ""),
            tags_str,
            c.get("notes", ""),
            c.get("draft_link", ""),
        ])

    # Save to a temp file and upload
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        tmp_out = f.name
    wb.save(tmp_out)

    try:
        with open(tmp_out, "rb") as f:
            out_bytes = f.read()
        graph.upload_to_onedrive(_ONEDRIVE_PATH, out_bytes, _XLSX_MIME)
        return {"appended": len(contacts), "error": None}
    except Exception as e:
        return {"appended": 0, "error": f"upload failed: {e}"}
    finally:
        for path in (tmp_in, tmp_out):
            if path:
                try:
                    os.unlink(path)
                except Exception:
                    pass
