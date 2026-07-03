"""Expenses store — brings the per-user expense ledger into the shared store.db, mirroring
projects_store.py. Expenses was the LAST file-only domain: receipts lived in
expenses/expenses_master.xlsx, and invoices/contracts had no durable home (dropped entirely on the
Teams path; kept only for the last email run in results/expenses.json) and no frontend view.

Each expense is one row keyed by a stable id = sha1("{msg_id}::{att_id}") — IDENTICAL to server's
_expense_row_id, so existing frontend row ids don't change — stored as a JSON blob so every field is
preserved verbatim. `document_type` ∈ {receipt, invoice, contract} distinguishes them; all three now
live in the same table (fixing "invoice sent to Audrey, acknowledged, but never in the dashboard").

Projections (store = truth; files = synced mirrors so legacy readers + the reimbursement export keep
working with no reader change):
  - expenses/expenses_master.xlsx  ← receipts ONLY (reimbursement export)
  - results/expenses.json          ← ALL types (section route / any legacy reader now sees invoices)

Dedup (replaces the two loose JSON caches _seen.json + _receipt_hashes.json) lives in expenses_seen:
examined-attachment keys ("{msg_id}::{att_name}") AND file hashes ("sha256:{hex}") in one table, so a
rejected/non-receipt attachment is never re-classified (the token-burn guard from commit 9ce035d).

One-time lazy import (idempotent via expenses_meta.migrated_from_json, reversible by deleting
store.db): receipts from the xlsx + invoices/contracts from results/expenses.json + the two dedup
caches, all carried over verbatim.
"""
import hashlib
import json
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

from src.modules.db_helpers import open_sqlite

_EXCEL_HEADERS = [
    "Date", "Vendor", "Amount", "Currency", "GST_HST", "Net_Amount",
    "Category", "Attachment", "Email_Subject", "From", "Msg_ID", "Att_ID", "Processed_Date",
]

# canonical unified item keys (the sections/expenses.py item shape + store-only source fields)
_ITEM_KEYS = [
    "document_type", "vendor", "counterparty", "date", "due_date", "amount", "currency",
    "gst_hst", "net_amount", "category", "subject", "confidence", "attachment",
    "email_subject", "from", "msg_id", "att_id", "processed_at",
    "source_type", "sha256", "onedrive_path",
]


def _db_path(data_dir) -> Path:
    return Path(data_dir) / "store.db"


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _s(v) -> str:
    return "" if v is None else str(v)


def expense_id(item: dict) -> str:
    """Stable id — identical formula to server._expense_row_id, computed from the item's lowercase
    msg_id/att_id so store ids == the ids the frontend already renders. Falls back to a content hash
    when both are empty (manual/rare)."""
    key = (_s(item.get("msg_id")) + "::" + _s(item.get("att_id")))
    if not key.strip(":"):
        key = json.dumps({k: _s(item.get(k)) for k in _ITEM_KEYS}, sort_keys=True)
    return hashlib.sha1(key.encode()).hexdigest()[:16]


def _conn(data_dir):
    path = _db_path(data_dir)
    path.parent.mkdir(parents=True, exist_ok=True)
    con = open_sqlite(path)
    con.row_factory = sqlite3.Row
    con.execute("CREATE TABLE IF NOT EXISTS expenses (id TEXT PRIMARY KEY, data TEXT NOT NULL, updated_at TEXT)")
    con.execute("CREATE TABLE IF NOT EXISTS expenses_seen (k TEXT PRIMARY KEY, at TEXT)")
    con.execute("CREATE TABLE IF NOT EXISTS expenses_meta (k TEXT PRIMARY KEY, v TEXT)")
    con.commit()
    _maybe_import(con, data_dir)
    return con


# ── one-time lazy import ─────────────────────────────────────────────────────

def _read_json(path: Path) -> dict:
    try:
        return json.loads(path.read_text()) if path.exists() else {}
    except Exception:
        return {}


def _row_from_excel(r: dict) -> dict:
    """Excel row (capitalised headers) → canonical lowercase item (always a receipt)."""
    frm = r.get("From") or ""
    src = "teams" if frm == "teams" else ("manual" if _s(r.get("Msg_ID")).startswith("manual_") else "email")
    return {
        "document_type": "receipt",
        "vendor":        r.get("Vendor", "") or "",
        "counterparty":  "",
        "date":          _s(r.get("Date")),
        "due_date":      None,
        "amount":        r.get("Amount", "") if r.get("Amount") is not None else "",
        "currency":      r.get("Currency", "") or "",
        "gst_hst":       r.get("GST_HST"),
        "net_amount":    r.get("Net_Amount"),
        "category":      r.get("Category", "Other") or "Other",
        "subject":       "",
        "confidence":    "",
        "attachment":    r.get("Attachment", "") or "",
        "email_subject": r.get("Email_Subject", "") or "",
        "from":          frm,
        "msg_id":        _s(r.get("Msg_ID")),
        "att_id":        _s(r.get("Att_ID")),
        "processed_at":  _s(r.get("Processed_Date")),
        "source_type":   src,
    }


def _excel_rows(xlsx_path: Path) -> list[dict]:
    if not xlsx_path.exists():
        return []
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    try:
        ws = load_workbook(xlsx_path).active
        headers = [c.value for c in ws[1]]
        out = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            out.append(_row_from_excel(dict(zip(headers, row))))
        return out
    except Exception as e:
        print(f"[expenses_store] excel read failed: {e}")
        return []


def _maybe_import(con, data_dir) -> None:
    row = con.execute("SELECT v FROM expenses_meta WHERE k='migrated_from_json'").fetchone()
    if row:
        if not con.execute("SELECT v FROM expenses_meta WHERE k='migration_check'").fetchone():
            _verify_import(con, data_dir)   # backfill verdict for an already-migrated user
        return
    data_dir = Path(data_dir)
    now = _now_iso()
    items = _excel_rows(data_dir / "expenses" / "expenses_master.xlsx")   # receipts (cumulative)
    for it in (_read_json(data_dir / "results" / "expenses.json").get("items") or []):
        if it.get("document_type") in ("receipt", "invoice", "contract"):
            items.append({**it, "source_type": it.get("source_type") or "email"})
    try:
        for it in items:
            con.execute("INSERT OR REPLACE INTO expenses (id, data, updated_at) VALUES (?,?,?)",
                        (expense_id(it), json.dumps(it, ensure_ascii=False), now))
        # examined-attachment + hash caches → one table (so nothing gets re-classified post-migration)
        for k in (_read_json(data_dir / "expenses" / "_seen.json") or {}).keys():
            con.execute("INSERT OR IGNORE INTO expenses_seen (k, at) VALUES (?,?)", (k, now))
        for h in (_read_json(data_dir / "expenses" / "_receipt_hashes.json") or {}).keys():
            con.execute("INSERT OR IGNORE INTO expenses_seen (k, at) VALUES (?,?)", (f"sha256:{h}", now))
    except Exception as e:
        print(f"[expenses_store] import skipped/failed: {e}")
    con.execute("INSERT OR REPLACE INTO expenses_meta (k, v) VALUES ('migrated_from_json', ?)", (now,))
    con.commit()
    _verify_import(con, data_dir)


def _verify_import(con, data_dir: Path) -> bool:
    """Best-effort self-check: store expense count must equal (xlsx receipts + json invoices/contracts,
    de-duplicated by id). Writes a durable verdict to expenses_meta + logs. Never raises."""
    try:
        data_dir = Path(data_dir)
        src = _excel_rows(data_dir / "expenses" / "expenses_master.xlsx")
        for it in (_read_json(data_dir / "results" / "expenses.json").get("items") or []):
            if it.get("document_type") in ("receipt", "invoice", "contract"):
                src.append(it)
        src_ids = {expense_id(it) for it in src}
        store_ids = {r["id"] for r in con.execute("SELECT id FROM expenses")}
        missing = sorted(src_ids - store_ids)
        ok = not missing
        name = Path(data_dir).name
        payload = {"verdict": "lossless" if ok else "mismatch", "source": len(src_ids),
                   "store": len(store_ids), "lost": missing[:20], "at": _now_iso()}
        con.execute("INSERT OR REPLACE INTO expenses_meta (k, v) VALUES ('migration_check', ?)",
                    (json.dumps(payload),))
        con.commit()
        if ok:
            print(f"[expenses_store] migration verified dir={name} expenses={len(store_ids)} (lossless)")
        else:
            print(f"[expenses_store] ⚠️ EXPENSES MIGRATION MISMATCH dir={name} source={len(src_ids)} "
                  f"store={len(store_ids)} lost={missing[:20]} — files preserved; rollback possible")
        return ok
    except Exception as e:
        print(f"[expenses_store] migration self-check skipped: {e}")
        return True


# ── dedup (replaces _seen.json + _receipt_hashes.json) ───────────────────────

def is_seen(data_dir, key: str) -> bool:
    """True if this attachment key ("{msg_id}::{att_name}") or hash key ("sha256:{hex}") was already
    examined. Callers pass the raw seen-key or a "sha256:"-prefixed hash."""
    con = _conn(data_dir)
    try:
        return con.execute("SELECT 1 FROM expenses_seen WHERE k=?", (key,)).fetchone() is not None
    finally:
        con.close()


def mark_seen(data_dir, key: str) -> None:
    con = _conn(data_dir)
    try:
        con.execute("INSERT OR IGNORE INTO expenses_seen (k, at) VALUES (?,?)", (key, _now_iso()))
        con.commit()
    finally:
        con.close()


# ── write paths ──────────────────────────────────────────────────────────────

def upsert_expense(data_dir, item: dict, project: bool = True) -> str:
    """Insert/replace one expense of any document_type (receipt/invoice/contract). Returns the stable
    id. Regenerates the xlsx (receipts) + results/expenses.json (all) projections unless project=False
    (batch scans upsert many with project=False, then call write_projection once)."""
    eid = expense_id(item)
    con = _conn(data_dir)
    try:
        con.execute("INSERT OR REPLACE INTO expenses (id, data, updated_at) VALUES (?,?,?)",
                    (eid, json.dumps(item, ensure_ascii=False), _now_iso()))
        con.commit()
    finally:
        con.close()
    if project:
        write_projection(data_dir)
    return eid


def update_expense_fields(data_dir, eid: str, updates: dict) -> dict | None:
    """Field-level edit (the server PATCH chokepoint). `updates` keys are the LOWERCASE item keys
    (the route maps its capitalised editable fields down first). Returns the updated item (+id) or
    None if not found."""
    con = _conn(data_dir)
    try:
        r = con.execute("SELECT data FROM expenses WHERE id=?", (eid,)).fetchone()
        if not r:
            return None
        item = json.loads(r["data"])
        item.update(updates or {})
        con.execute("INSERT OR REPLACE INTO expenses (id, data, updated_at) VALUES (?,?,?)",
                    (eid, json.dumps(item, ensure_ascii=False), _now_iso()))
        con.commit()
    finally:
        con.close()
    write_projection(data_dir)
    return {**item, "id": eid}


def delete_expense(data_dir, eid: str) -> bool:
    con = _conn(data_dir)
    try:
        cur = con.execute("DELETE FROM expenses WHERE id=?", (eid,))
        con.commit()
        removed = cur.rowcount > 0
    finally:
        con.close()
    write_projection(data_dir)
    return removed


def set_last_run(data_dir, iso: str = None) -> None:
    con = _conn(data_dir)
    try:
        con.execute("INSERT OR REPLACE INTO expenses_meta (k, v) VALUES ('last_run', ?)",
                    (json.dumps(iso or _now_iso()),))
        con.commit()
    finally:
        con.close()


def clear(data_dir) -> None:
    """Drop all expense rows + caches + meta (onboarding reset). Next access re-imports from the
    (now empty/absent) files."""
    path = _db_path(data_dir)
    if not path.exists():
        return
    con = open_sqlite(path)
    try:
        con.execute("DROP TABLE IF EXISTS expenses")
        con.execute("DROP TABLE IF EXISTS expenses_seen")
        con.execute("DROP TABLE IF EXISTS expenses_meta")
        con.commit()
    finally:
        con.close()


# ── read path ────────────────────────────────────────────────────────────────

def _meta(con, key, default=None):
    r = con.execute("SELECT v FROM expenses_meta WHERE k=?", (key,)).fetchone()
    if not r:
        return default
    try:
        return json.loads(r["v"])
    except Exception:
        return default


def load_expenses(data_dir, doc_type=None) -> list[dict]:
    """All expenses (each item dict + injected `id`), newest first. Optional document_type filter."""
    con = _conn(data_dir)
    try:
        out = []
        for r in con.execute("SELECT id, data FROM expenses"):
            it = json.loads(r["data"])
            it["id"] = r["id"]
            if doc_type and it.get("document_type") != doc_type:
                continue
            out.append(it)
    finally:
        con.close()
    out.sort(key=lambda x: (_s(x.get("processed_at")), _s(x.get("date"))), reverse=True)
    return out


def write_projection(data_dir) -> None:
    """Regenerate expenses_master.xlsx (receipts only — reimbursement export) + results/expenses.json
    (all types — section route / legacy readers) from the store. Best-effort, atomic."""
    data_dir = Path(data_dir)
    items = load_expenses(data_dir)
    con = _conn(data_dir)
    try:
        last_run = _meta(con, "last_run")
    finally:
        con.close()
    # results/expenses.json — ALL types
    try:
        res = {
            "id":       "expenses",
            "status":   "fresh" if items else "not_run",
            "last_run": last_run or _now_iso(),
            "items":    items,
            "count":    len(items),
            "empty":    len(items) == 0,
        }
        p = data_dir / "results" / "expenses.json"
        p.parent.mkdir(parents=True, exist_ok=True)
        tmp = p.with_suffix(".json.tmp")
        tmp.write_text(json.dumps(res, indent=2, ensure_ascii=False))
        tmp.replace(p)
    except Exception as e:
        print(f"[expenses_store] json projection failed: {e}")
    # expenses_master.xlsx — receipts only
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses"
        ws.append(_EXCEL_HEADERS)
        for it in items:
            if it.get("document_type") != "receipt":
                continue
            ws.append([
                it.get("date", ""), it.get("vendor", ""), it.get("amount", ""),
                it.get("currency", ""), it.get("gst_hst", ""), it.get("net_amount", ""),
                it.get("category", ""), it.get("attachment", ""), it.get("email_subject", ""),
                it.get("from", ""), it.get("msg_id", ""), it.get("att_id", ""), it.get("processed_at", ""),
            ])
        xp = data_dir / "expenses" / "expenses_master.xlsx"
        xp.parent.mkdir(parents=True, exist_ok=True)
        tmp = xp.with_suffix(".xlsx.tmp")
        wb.save(tmp)
        tmp.replace(xp)
    except ImportError:
        pass
    except Exception as e:
        print(f"[expenses_store] xlsx projection failed: {e}")


def get_migration_status(data_dir) -> dict:
    """Read-only verdict for the admin endpoint — does NOT trigger migration."""
    path = _db_path(data_dir)
    if not path.exists():
        return {"state": "not_accessed"}
    con = sqlite3.connect(str(path))
    con.row_factory = sqlite3.Row
    try:
        try:
            mig = con.execute("SELECT v FROM expenses_meta WHERE k='migrated_from_json'").fetchone()
            chk = con.execute("SELECT v FROM expenses_meta WHERE k='migration_check'").fetchone()
        except Exception as e:
            return {"state": "error", "error": str(e)}
    finally:
        con.close()
    if not mig:
        return {"state": "no_flag"}
    if not chk:
        return {"state": "migrated_unchecked"}
    try:
        return {"state": "checked", **json.loads(chk["v"])}
    except Exception:
        return {"state": "checked", "raw": chk["v"]}
