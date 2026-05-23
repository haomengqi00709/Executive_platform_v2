"""
Compatibility bridge for Teams bot receipt handling.
bot.py and teams_bot.py import from here.
Core email-scan implementation lives in src/sections/expenses.py.
"""
import hashlib
import json
from pathlib import Path

from src.sections.expenses import _extract_document, _load_json, _save_json

# Sentinel paths — callers that pass data_dir will construct their own paths and
# override these. These only apply when teams_bot falls back to EXPENSES_DIR.
EXPENSES_DIR = Path(".data/_default/expenses")
MASTER_FILE  = EXPENSES_DIR / "expenses_master.xlsx"

_EXCEL_HEADERS = [
    "Date", "Vendor", "Amount", "Currency", "GST_HST", "Net_Amount",
    "Category", "Attachment", "Email_Subject", "From", "Msg_ID", "Att_ID", "Processed_Date",
]


def _extract_from_attachment(img_bytes: bytes, mime: str, filename: str, ai) -> dict:
    """Extract document fields from raw bytes using Gemini vision.
    Returns dict with `document_type` (receipt/invoice/contract/other) and extracted fields.
    For backward-compat with teams_bot.py callers, also sets `is_receipt` boolean."""
    result = _extract_document(ai, img_bytes, filename)
    if not result:
        return {}
    # Back-compat for teams_bot.py which checks is_receipt
    result["is_receipt"] = (result.get("document_type") == "receipt")
    return result


def _init_workbook():
    """Create a new Workbook with expense column headers."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    ws.append(_EXCEL_HEADERS)
    return wb


def _append_row(ws, row: dict) -> None:
    """
    Append a row to an open worksheet.
    row keys must match _EXCEL_HEADERS (capitalised), e.g. {"Date": ..., "Vendor": ...}.
    """
    ws.append([
        row.get("Date", ""),
        row.get("Vendor", ""),
        row.get("Amount", ""),
        row.get("Currency", ""),
        row.get("GST_HST", ""),
        row.get("Net_Amount", ""),
        row.get("Category", ""),
        row.get("Attachment", ""),
        row.get("Email_Subject", ""),
        row.get("From", ""),
        row.get("Msg_ID", ""),
        row.get("Att_ID", ""),
        row.get("Processed_Date", ""),
    ])


def _load_seen(path: Path | None) -> set:
    """Load the seen-key set from a JSON file (keys of the dict)."""
    if not path or not path.exists():
        return set()
    try:
        data = json.loads(path.read_text())
        if isinstance(data, dict):
            return set(data.keys())
        return set(data)
    except Exception:
        return set()


def _save_seen(data: set, path: Path | None) -> None:
    """Persist the seen-key set as {key: true} JSON."""
    if not path:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    _save_json(path, {k: True for k in data})


def _load_hashes(path: Path | None) -> dict:
    """Load sha256 → filename hash map."""
    if not path:
        return {}
    return _load_json(path)


def _save_hashes(data: dict, path: Path | None) -> None:
    """Persist sha256 → filename hash map."""
    if not path:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    _save_json(path, data)


def _compute_hash(data: bytes) -> str:
    """SHA-256 hex digest of raw bytes."""
    return hashlib.sha256(data).hexdigest()


def _find_field_match(master_file: Path, vendor: str, amount, date: str) -> dict | None:
    """
    Scan expenses_master.xlsx for an existing row with the same vendor + date + amount (±$0.01).
    Returns the matching row dict (column names as keys) or None.
    Used for field-level dedup before confirming a Teams receipt.
    """
    if not master_file or not master_file.exists():
        return None
    try:
        from openpyxl import load_workbook
        wb      = load_workbook(master_file)
        ws      = wb.active
        headers = [c.value for c in ws[1]]
        for row_vals in ws.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, row_vals))
            same_vendor = (row.get("Vendor") or "").lower() == (vendor or "").lower()
            same_date   = row.get("Date", "") == date
            try:
                same_amount = abs(float(row.get("Amount") or 0) - float(amount or 0)) < 0.01
            except Exception:
                same_amount = False
            if same_vendor and same_date and same_amount:
                return row
    except Exception:
        pass
    return None
