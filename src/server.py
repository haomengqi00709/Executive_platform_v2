"""
CEO AI Platform v2 — FastAPI server
Auth + Settings + Teams bot registration
"""
import json
import re
import secrets
import threading
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path

from fastapi import BackgroundTasks, Cookie, Depends, FastAPI, File, Header, HTTPException, Request, Response, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from dotenv import load_dotenv

from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.events import EVENT_JOB_EXECUTED, EVENT_JOB_ERROR

from src import auth
from src.graph import GraphClient
from src.ai import AIClient
from src.modules.tz import get_user_tz, now_local, today_local_str
from src.sections import ai_summary, expenses, reply_needed, followup_needed, commitments_extract, upcoming_commitments
from src.sections import recent_meetings, meeting_action_items
from src.sections import market_intelligence, company_intelligence
from src.sections import relationship_health, business_insights
from src.sections import meetings_today, due_today, yesterday_recap
from src.sections import projects_needing_attention, project_status
from src.sections import meeting_prep

load_dotenv(override=True)

import os
FRONTEND_URL = os.getenv("FRONTEND_URL", "http://localhost:3000")

app = FastAPI(title="CEO AI Platform v2")

# CORS restricted to deployed origin(s). FRONTEND_URL may be a single URL
# or a comma-separated list (handy for dev with multiple local ports).
_allowed_origins = [o.strip() for o in FRONTEND_URL.split(",") if o.strip()]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_allowed_origins,
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Content-Type", "X-Timezone"],
    allow_credentials=True,
)

# Session cookies use Secure flag when serving over HTTPS.
# Auto-detect from REDIRECT_URI so local dev (http://localhost) still works.
_SECURE_COOKIE = os.getenv("REDIRECT_URI", "").startswith("https://")

(auth.DATA_DIR / "_sessions").mkdir(parents=True, exist_ok=True)

_device_flow: dict = {}
_device_flow_lock = threading.Lock()
_bot_device_flow: dict = {}
_bot_device_flow_lock = threading.Lock()

_scheduler = BackgroundScheduler(timezone="UTC")


# ── File helpers ──────────────────────────────────────────

def _read_json(path: Path) -> dict:
    try:
        return json.loads(path.read_text()) if path.exists() else {}
    except Exception:
        return {}

def _write_json(path: Path, data: dict):
    """Atomic, concurrency-safe JSON write — delegates to src/storage.py.
    (The old tmp+os.replace used a SHARED tmp name, which still raced under
    concurrent writers on Azure Files SMB → the 2026-06 'Extra data' corruption.)"""
    from src.storage import atomic_write_json
    atomic_write_json(path, data)

def _within(iso_ts: str | None, secs: int) -> bool:
    if not iso_ts:
        return False
    try:
        ts = datetime.fromisoformat(iso_ts.replace("Z", "+00:00"))
        if ts.tzinfo is None:
            ts = ts.replace(tzinfo=timezone.utc)
        return (datetime.now(timezone.utc) - ts).total_seconds() <= secs
    except Exception:
        return False

def _udir(user_id: str) -> Path:
    return auth.DATA_DIR / user_id

def _user_settings(user_id: str) -> Path:
    return _udir(user_id) / "settings.json"

def _user_bot_link_path(user_id: str) -> Path:
    return _udir(user_id) / "bot_link.json"


# ── Ops: scheduler job health (stream 2) ──────────────────
# An APScheduler listener records last_run / last_success / last_error per job id
# to a fleet-global file, so the ops dashboard can tell whether scheduled work is
# actually running — auth can read "healthy" while a dead job means no briefings
# go out. Recording is best-effort and never raises into the scheduler.

_OPS_DIR = auth.DATA_DIR / "_ops"
_JOB_HEALTH_PATH = _OPS_DIR / "job_health.json"

# Expected cadence per job (seconds), so the ops side can flag staleness
# sensibly — a 10s bot poll and a weekly cleanup need very different thresholds.
_JOB_EXPECTED_INTERVAL_SECS = {
    "teams_bot_poll":          10,
    "email_monitor_poll":      60,
    "expense_scan_poll":       60,
    "meeting_prep_poll":       300,
    "meeting_recordings_poll": 1200,
    "init_stuck_detector":     60,
    "crm_daily_refresh":       86400,
    "projects_daily_refresh":  86400,
    "companies_daily_refresh": 86400,
    "db_cleanup_weekly":       604800,
    "canary_probe":            900,
    "push_qa_poll":            3600,
}


def _record_job_event(event) -> None:
    """APScheduler listener — records success/failure timing per job id.
    Best-effort: any error here must never disturb the scheduler."""
    try:
        job_id = getattr(event, "job_id", None)
        if not job_id:
            return
        now = datetime.now(timezone.utc).isoformat()
        all_health = _read_json(_JOB_HEALTH_PATH)
        rec = all_health.get(job_id) or {}
        rec["last_run"] = now
        if getattr(event, "exception", None):
            rec["last_error"] = str(event.exception)[:500]
            rec["last_error_at"] = now
            rec["consecutive_failures"] = rec.get("consecutive_failures", 0) + 1
        else:
            rec["last_success"] = now
            rec["last_error"] = None
            rec["consecutive_failures"] = 0
        exp = _JOB_EXPECTED_INTERVAL_SECS.get(job_id)
        if exp:
            rec["expected_interval_secs"] = exp
        all_health[job_id] = rec
        _write_json(_JOB_HEALTH_PATH, all_health)
    except Exception:
        pass


def _run_canary_probe() -> None:
    """Synthetic end-to-end probe: exercises the critical path (token refresh →
    Graph reach) for a designated canary account, so breakage surfaces before a
    real customer hits it. No-op unless CANARY_UID names an existing account.
    Raises on failure so the job-health listener records it uniformly — a failing
    canary then shows up as a stalled `canary_probe` job and pages ops."""
    uid = (os.getenv("CANARY_UID") or "").strip()
    if not uid:
        return  # canary not configured — opt-in via env
    token = auth.get_valid_access_token(uid)   # raises if auth is broken
    GraphClient(token).get_me()                # raises if Graph is unreachable


def _poll_push_qa_all_users() -> None:
    """Sample recent section pushes per user and score quality (stream 4).
    Gated OFF by default — set PUSH_QA_ENABLED=true to run. Writes a fleet
    roll-up to _ops/push_qa.json that the ops dashboard surfaces. Best-effort:
    never raises into the scheduler."""
    if (os.getenv("PUSH_QA_ENABLED") or "").strip().lower() not in ("1", "true", "yes", "on"):
        return  # not enabled — makes Gemini calls, so opt-in only
    from src.modules import push_qa

    sessions_dir = auth.DATA_DIR / "_sessions"
    if not sessions_dir.exists():
        return
    ai = AIClient()
    roll_up = {"generated_at": datetime.now(timezone.utc).isoformat(), "users": []}
    for token_file in sessions_dir.glob("*.json"):
        uid = token_file.stem
        bot_path = _bot_state_path(uid)
        if bot_path.exists() and _is_bound_bot(_read_json(bot_path)):
            continue  # QA the human owner's pushes, not the bot account
        try:
            graph = GraphClient(auth.get_valid_access_token(uid))
        except Exception:
            graph = None  # auth broken — heuristics/judge still run, grounding skips
        try:
            settings = _read_json(_user_settings(uid))
            display_name = settings.get("display_name") or "the executive"
            summary = push_qa.run_for_user(_udir(uid), ai, graph, display_name)
            roll_up["users"].append({
                "uid":      uid,
                "username": (auth.load_user_tokens(uid) or {}).get("username", ""),
                "fail":     summary.get("fail", 0),
                "warn":     summary.get("warn", 0),
                "ok":       summary.get("ok", 0),
                "sections": [
                    {"section_id": s.get("section_id"), "verdict": s.get("verdict"),
                     "score": (s.get("judge") or {}).get("overall"),
                     "ai_checked": s.get("ai_checked", False)}
                    for s in summary.get("sections", [])
                ],
            })
        except Exception as e:
            print(f"[PushQA] Error for {uid}: {e}")
    try:
        _write_json(_OPS_DIR / "push_qa.json", roll_up)
    except Exception:
        pass


def _get_bot_display_name(user_id: str) -> str:
    """Per-user AI assistant display name. Resolved in priority order:
      1. settings.bot_display_name — explicit override (rarely set)
      2. Bot's OAuth email local-part — auto-derived from Step 3 setup
         (e.g. "Audrey@imodel3d.com" → "Audrey", "Edileen@x.com" → "Edileen")
      3. Fallback to "Audrey"
    The auto path means the user never has to configure this — the name is
    whatever they (or their IT) called the assistant's M365 account."""
    try:
        s = _read_json(_user_settings(user_id))
        explicit = (s.get("bot_display_name") or "").strip()
        if explicit:
            return explicit
    except Exception:
        pass
    # Auto from bot account email
    try:
        link    = _read_json(_user_bot_link_path(user_id))
        bot_uid = link.get("bot_uid")
        if bot_uid:
            bot_email = (auth.load_user_tokens(bot_uid) or {}).get("username") or ""
            if "@" in bot_email:
                local = bot_email.split("@")[0]
                if local:
                    # Capitalize first char if all lower ("audrey" → "Audrey")
                    # but leave names already cased ("iModelBot") alone.
                    return local[0].upper() + local[1:] if local[0].islower() else local
    except Exception:
        pass
    return "Audrey"


# ── Session helpers ───────────────────────────────────────

def get_current_session(session_token: str = Cookie(None)) -> dict | None:
    if not session_token:
        return None
    return auth.decode_jwt(session_token)

def require_session(session: dict | None = Depends(get_current_session)) -> dict:
    if not session:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return session


# ── Redirect URI detection ────────────────────────────────

def _detect_redirect_uri(request: Request) -> str:
    proto = request.headers.get("x-forwarded-proto") or ("https" if request.url.scheme == "https" else "http")
    host  = request.headers.get("x-forwarded-host") or request.headers.get("host") or "localhost:8000"
    return f"{proto}://{host}/auth/callback"


# ── Web OAuth endpoints ───────────────────────────────────

@app.get("/auth/login")
def login(request: Request):
    state        = secrets.token_urlsafe(16)
    redirect_uri = _detect_redirect_uri(request)
    url          = auth.get_auth_url(state, redirect_uri=redirect_uri)
    resp         = RedirectResponse(url)
    resp.set_cookie("oauth_state",    state,        httponly=True, max_age=600)
    resp.set_cookie("oauth_redirect", redirect_uri, httponly=True, max_age=600)
    return resp


@app.get("/auth/callback")
def auth_callback(code: str = None, state: str = None, error: str = None,
                  oauth_state: str = Cookie(None), oauth_redirect: str = Cookie(None)):
    if error:
        return RedirectResponse(f"{FRONTEND_URL}/?error={error}")
    if not code:
        return RedirectResponse(f"{FRONTEND_URL}/?error=no_code")
    try:
        result = auth.exchange_code(code, redirect_uri=oauth_redirect)
    except Exception as e:
        return RedirectResponse(f"{FRONTEND_URL}/?error={str(e)[:80]}")

    claims   = result.get("id_token_claims", {})
    user_id  = claims.get("oid") or claims.get("sub", "unknown")
    username = claims.get("preferred_username") or claims.get("email", "")
    expiry   = (datetime.now(timezone.utc) + timedelta(seconds=result.get("expires_in", 3600))).isoformat()

    auth.save_user_tokens(user_id, {
        "access_token":  result["access_token"],
        "refresh_token": result.get("refresh_token", ""),
        "expiry":        expiry,
        "username":      username,
    })

    jwt_token = auth.create_jwt(user_id, username)
    resp = RedirectResponse(f"{FRONTEND_URL}/")
    resp.set_cookie("session_token", jwt_token, httponly=True, secure=_SECURE_COOKIE, max_age=60*60*24*7, samesite="lax")
    resp.delete_cookie("oauth_state")
    resp.delete_cookie("oauth_redirect")
    _ensure_onedrive_folders(user_id)
    from src.modules.profile import is_profile_confirmed, init_has_started, save_init_status
    if not is_profile_confirmed(_udir(user_id)) and not init_has_started(_udir(user_id)):
        # Don't start init yet — wait for the user to confirm the right account is logged in.
        # Frontend reads init_status, shows a confirmation screen, then POSTs /api/init/start.
        save_init_status(_udir(user_id), "awaiting_confirmation",
                         current_message=f"Logged in as {username}. Confirm to start setup.")
        print(f"[init] Awaiting user confirmation for {user_id} ({username})")
    return resp


@app.get("/auth/logout")
def logout(session_token: str = Cookie(None)):
    session = auth.decode_jwt(session_token) if session_token else None
    if session:
        auth.delete_user_tokens(session["user_id"])
    resp = RedirectResponse(f"{FRONTEND_URL}/")
    resp.delete_cookie("session_token")
    return resp


# ── Auth info endpoints ───────────────────────────────────

@app.get("/api/auth/me")
def auth_me(request: Request, session: dict | None = Depends(get_current_session)):
    if not session:
        return {"authenticated": False}
    tz = request.headers.get("X-Timezone", "").strip()
    if tz and session.get("user_id"):
        path     = _user_settings(session["user_id"])
        settings = _read_json(path)
        if settings.get("timezone") != tz:
            settings["timezone"] = tz
            _write_json(path, settings)
    return {"authenticated": True, "username": session.get("username"), "user_id": session.get("user_id")}


@app.get("/api/auth/status")
def auth_status():
    return {"authenticated": False}


# ── Device Code Flow (main user login) ────────────────────

@app.post("/api/auth/start")
def auth_start():
    with _device_flow_lock:
        global _device_flow
        flow = auth.start_device_flow()
        _device_flow = flow
        return {
            "user_code":        flow["user_code"],
            "verification_uri": flow["verification_uri"],
            "message":          flow["message"],
            "expires_in":       flow.get("expires_in", 900),
        }


@app.post("/api/auth/poll")
def auth_poll(response: Response):
    with _device_flow_lock:
        if not _device_flow:
            return {"status": "no_flow"}
        try:
            token = auth.complete_device_flow(_device_flow)
            _device_flow.clear()
            cache   = auth._load_cache()
            app_    = auth._build_legacy_app(cache)
            account = next(iter(app_.get_accounts()), None)
            if account:
                import requests as _req
                me = _req.get(
                    "https://graph.microsoft.com/v1.0/me",
                    headers={"Authorization": f"Bearer {token}"},
                    timeout=10,
                ).json()
                user_id  = me.get("id", account.get("local_account_id", "local"))
                username = me.get("mail") or me.get("userPrincipalName") or account.get("username", "")
                expiry   = (datetime.now(timezone.utc) + timedelta(hours=1)).isoformat()
                auth.save_user_tokens(user_id, {
                    "access_token":  token,
                    "refresh_token": "",
                    "expiry":        expiry,
                    "username":      username,
                })
                jwt_token = auth.create_jwt(user_id, username)
                response.set_cookie("session_token", jwt_token, httponly=True,
                                    secure=_SECURE_COOKIE, max_age=60*60*24*7, samesite="lax")
                _ensure_onedrive_folders(user_id)
                from src.modules.profile import is_profile_confirmed, init_has_started, save_init_status
                if not is_profile_confirmed(_udir(user_id)) and not init_has_started(_udir(user_id)):
                    save_init_status(_udir(user_id), "awaiting_confirmation",
                                     current_message=f"Logged in as {username}. Confirm to start setup.")
                    print(f"[init] Awaiting user confirmation for {user_id} ({username}) — device code")
            return {"status": "success"}
        except Exception as e:
            msg = str(e)
            if "authorization_pending" in msg or "slow_down" in msg:
                return {"status": "pending"}
            return {"status": "error", "message": msg}


# ── Settings ──────────────────────────────────────────────

@app.get("/api/settings")
def get_settings(request: Request, session: dict = Depends(require_session)):
    uid      = session["user_id"]
    path     = _user_settings(uid)
    settings = _read_json(path)
    ms_email = (auth.load_user_tokens(uid) or {}).get("username", "")
    if not settings.get("report_email") and ms_email:
        settings["report_email"] = ms_email
    if not settings.get("display_name") and ms_email:
        local = ms_email.split("@")[0]
        settings["display_name"] = " ".join(p.capitalize() for p in local.replace(".", " ").split())
    tz = request.headers.get("X-Timezone", "").strip()
    if tz and settings.get("timezone") != tz:
        settings["timezone"] = tz
    _write_json(path, settings)
    # bot_display_name is computed (not stored) — derived from the bot's M365
    # account email so the UI can render it without the user configuring
    # anything. Computed AFTER the disk write so it doesn't leak into the
    # persisted settings file (avoids stale data if user reassigns the bot).
    if not settings.get("bot_display_name"):
        settings["bot_display_name"] = _get_bot_display_name(uid)
    return settings


@app.patch("/api/settings")
def update_settings(body: dict, session: dict = Depends(require_session)):
    uid  = session["user_id"]
    path = _user_settings(uid)
    settings = _read_json(path)
    settings.update(body)
    _write_json(path, settings)
    return settings


# ── Bot helpers ───────────────────────────────────────────

def _bot_state_path(user_id: str) -> Path:
    return _udir(user_id) / "teams_bot.json"


def _is_bound_bot(state: dict) -> bool:
    """True if this account is a bot currently bound to an owner. Keyed on
    owner_uid (which IS cleared on unbind) — not on is_registered_bot alone,
    which is write-once and was historically never cleared — so a stale
    registration flag can no longer permanently mis-route an account as a bot
    (the 'zombie bot' bug). Per-user OWNER polls skip these; bot-side polls
    additionally require `enabled`."""
    return bool(state.get("is_registered_bot") and state.get("owner_uid"))


def _heal_zombie_bot_flags_at_startup() -> None:
    """One-time self-heal at boot: reset 'zombie' bot flags left by an aborted
    bot setup. A zombie is is_registered_bot=True with no owner_uid — created
    when someone signed in as their OWN account during the bot device-flow
    (bot_auth_poll set the flag before activate_bot rejected the self-bind, and
    nothing ever cleared it). Such an account is really an owner, but every
    per-user event poll skipped it as a bot. Reset it to a clean owner state so
    its meetings/expenses process again. Only touches zombies — real bound bots
    (owner_uid set) are left untouched."""
    sessions_dir = auth.DATA_DIR / "_sessions"
    if not sessions_dir.exists():
        return
    for tf in sessions_dir.glob("*.json"):
        bp = _bot_state_path(tf.stem)
        if not bp.exists():
            continue
        bs = _read_json(bp)
        if bs.get("is_registered_bot") and not bs.get("owner_uid"):
            bs["is_registered_bot"] = False
            bs["enabled"] = False
            _write_json(bp, bs)
            print(f"[startup] reset zombie bot flag for {tf.stem}")


def _find_bot_for_user(user_id: str):
    link_path = _user_bot_link_path(user_id)
    if link_path.exists():
        bot_uid = _read_json(link_path).get("bot_uid")
        if bot_uid:
            bp = _bot_state_path(bot_uid)
            if bp.exists():
                bs = json.loads(bp.read_text())
                if bs.get("enabled") and bs.get("owner_uid") == user_id:
                    return bot_uid, bs.get("chat_id")
            link_path.unlink(missing_ok=True)

    sessions_dir = auth.DATA_DIR / "_sessions"
    for tf in sorted(sessions_dir.glob("*.json")):
        bid = tf.stem
        if bid == user_id:
            continue
        bp = _bot_state_path(bid)
        if not bp.exists():
            continue
        bs = json.loads(bp.read_text())
        if bs.get("enabled") and bs.get("owner_uid") == user_id and bs.get("chat_id"):
            _write_json(link_path, {"bot_uid": bid})
            return bid, bs["chat_id"]

    return None, None


def _bind_bot_to_user(bot_uid: str, user_id: str, username: str):
    existing = _read_json(_user_bot_link_path(user_id)).get("bot_uid")
    if existing and existing != bot_uid:
        ep = _bot_state_path(existing)
        if ep.exists():
            es = json.loads(ep.read_text())
            es.update({"is_registered_bot": False, "owner_uid": None, "peer_email": None, "chat_id": None, "last_seen_ts": None})
            _write_json(ep, es)

    bp = _bot_state_path(bot_uid)
    bs = json.loads(bp.read_text()) if bp.exists() else {}
    prev_owner = bs.get("owner_uid")
    if prev_owner and prev_owner != user_id:
        prev_link = _user_bot_link_path(prev_owner)
        if prev_link.exists() and _read_json(prev_link).get("bot_uid") == bot_uid:
            prev_link.unlink(missing_ok=True)

    bs.update({
        "enabled":           True,
        "is_registered_bot": True,
        "owner_uid":         user_id,
        "peer_email":        username,
        "chat_id":           None,
        "last_seen_ts":      None,
    })
    _write_json(bp, bs)
    _write_json(_user_bot_link_path(user_id), {"bot_uid": bot_uid})


def _unbind_bot_from_user(user_id: str) -> str | None:
    link_path = _user_bot_link_path(user_id)
    bot_uid = _read_json(link_path).get("bot_uid") if link_path.exists() else None
    if not bot_uid:
        sessions_dir = auth.DATA_DIR / "_sessions"
        for tf in sessions_dir.glob("*.json"):
            bid = tf.stem
            if bid == user_id:
                continue
            bp = _bot_state_path(bid)
            if not bp.exists():
                continue
            bs = json.loads(bp.read_text())
            if bs.get("enabled") and bs.get("owner_uid") == user_id:
                bot_uid = bid
                break
    if bot_uid:
        bp = _bot_state_path(bot_uid)
        if bp.exists():
            bs = json.loads(bp.read_text())
            bs.update({"is_registered_bot": False, "owner_uid": None, "peer_email": None, "chat_id": None, "last_seen_ts": None})
            _write_json(bp, bs)
        link_path.unlink(missing_ok=True)
    return bot_uid


def _send_activation_greeting(bot_uid: str, peer_email: str, display_name: str):
    import time
    time.sleep(3)
    try:
        token   = auth.get_valid_access_token(bot_uid)
        graph   = GraphClient(token)
        chat_id = graph.find_chat_with_user(peer_email)
        if not chat_id:
            return
        bp    = _bot_state_path(bot_uid)
        state = json.loads(bp.read_text())
        state["chat_id"]      = chat_id
        state["last_seen_ts"] = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")
        _write_json(bp, state)

        bot_me      = graph.get_me()
        bot_display = bot_me.get("displayName") or peer_email.split("@")[0]
        greeting = (
            f"Hi {display_name}! 👋 I'm {bot_display}, your AI executive assistant.\n\n"
            f"I'm now connected. Here's what I can do:\n"
            f"• 📧 Alert you to important emails and draft replies\n"
            f"• 📅 Send pre-meeting briefs and reminders\n"
            f"• 🌅 Deliver your morning briefing each day\n"
            f"• 💬 Answer questions — just ask me anything here!"
        )
        graph.send_chat_message(chat_id, greeting)
    except Exception as e:
        print(f"[TeamsBot] Greeting failed for {peer_email}: {e}")


# ── OneDrive folder bootstrap ─────────────────────────────

def _ensure_onedrive_folders(user_id: str) -> None:
    """Create standard platform folders in the user's OneDrive (idempotent)."""
    def _bg():
        try:
            token = auth.get_valid_access_token(user_id)
            g = GraphClient(token)
            for folder in [
                "CEO Platform/Receipts",
                "CEO Platform/Meetings",
                "CEO Platform/CRM",
            ]:
                g.ensure_folder_path(folder)
            print(f"[OneDrive] Platform folders ready for {user_id}")
        except Exception as e:
            print(f"[OneDrive] Folder setup failed for {user_id}: {e}")
    threading.Thread(target=_bg, daemon=True).start()


# ── Teams bot polling (scheduler) ─────────────────────────

def _poll_teams_bot_all_users():
    from src.modules.teams_bot import poll_and_reply
    sessions_dir = auth.DATA_DIR / "_sessions"
    if not sessions_dir.exists():
        return
    for token_file in sessions_dir.glob("*.json"):
        uid  = token_file.stem
        path = _bot_state_path(uid)
        if not path.exists():
            continue
        try:
            state = _read_json(path)
            if not state.get("enabled") or not state.get("is_registered_bot"):
                continue
            owner_uid = state.get("owner_uid")
            if not owner_uid:
                # Orphan bot — registered but unbound. Don't fall back to
                # treating the bot as its own owner (that silently uses wrong
                # context). Log once per bot, then skip every poll.
                if not state.get("_orphan_logged"):
                    print(f"[TeamsBot] Bot {uid} has no owner_uid — skipping poll. "
                          f"Use /api/admin/bot/unbind/{uid} to clean up.")
                    state["_orphan_logged"] = True
                    _write_json(path, state)
                continue
            token       = auth.get_valid_access_token(uid)
            graph       = GraphClient(token)
            owner_settings = _read_json(_user_settings(owner_uid))
            owner_graph = None
            try:
                owner_graph = GraphClient(auth.get_valid_access_token(owner_uid))
            except Exception:
                pass
            new_state = poll_and_reply(
                state, graph, AIClient(),
                owner_graph=owner_graph,
                owner_wiki_dir=_udir(owner_uid) / "wiki",
                owner_settings=owner_settings,
                owner_settings_path=_user_settings(owner_uid),
                owner_context_path=_udir(owner_uid) / "context.json",
                owner_data_dir=_udir(owner_uid),
                bot_state_path=path,
            )
            path.write_text(json.dumps(new_state, indent=2, ensure_ascii=False))
        except Exception as e:
            msg = str(e)
            if not any(c in msg for c in ("502", "503", "504", "ConnectionError", "Timeout")):
                print(f"[TeamsBot] Error for {uid}: {e}")


# ── Email monitor polling (scheduler) ─────────────────────

def _poll_email_monitor_all_users():
    from src.modules.email_monitor import poll_and_notify
    sessions_dir = auth.DATA_DIR / "_sessions"
    if not sessions_dir.exists():
        return
    for token_file in sessions_dir.glob("*.json"):
        uid  = token_file.stem
        path = _bot_state_path(uid)
        if not path.exists():
            continue
        try:
            state = _read_json(path)
            if not state.get("enabled") or not state.get("is_registered_bot"):
                continue
            owner_uid = state.get("owner_uid")
            if not owner_uid:
                # Orphan bot — skip silently. The TeamsBot polling loop already
                # logged the warning once; no need to spam it here.
                continue
            chat_id   = state.get("chat_id")
            if not chat_id:
                continue
            owner_settings = _read_json(_user_settings(owner_uid))
            owner_graph = None
            try:
                owner_graph = GraphClient(auth.get_valid_access_token(owner_uid))
            except Exception:
                continue
            bot_token = auth.get_valid_access_token(uid)
            graph = GraphClient(bot_token)
            poll_and_notify(
                graph=graph,
                owner_graph=owner_graph,
                data_dir=_udir(owner_uid),
                settings=owner_settings,
                chat_id=chat_id,
            )
        except Exception as e:
            msg = str(e)
            if not any(c in msg for c in ("502", "503", "504", "ConnectionError", "Timeout")):
                print(f"[EmailMonitor] Error for {uid}: {e}")


# ── Expense auto-scan polling (scheduler) ─────────────────

def _poll_expense_scan_all_users():
    """Auto-scan inbox every 10 min for receipt attachments. Notifies via Audrey if new receipts found."""
    sessions_dir = auth.DATA_DIR / "_sessions"
    if not sessions_dir.exists():
        return
    for token_file in sessions_dir.glob("*.json"):
        uid = token_file.stem
        bot_path = _bot_state_path(uid)
        if bot_path.exists() and _is_bound_bot(_read_json(bot_path)):
            continue
        try:
            owner_graph = GraphClient(auth.get_valid_access_token(uid))
            ai          = AIClient()
            result      = expenses.run(owner_graph, ai, _udir(uid), days=1)
            new_items   = result.get("items", [])
            if not new_items:
                continue
            print(f"[ExpensePoll] {len(new_items)} new receipt(s) for {uid}")
            try:
                bot_uid, chat_id = _find_bot_for_user(uid)
                if bot_uid and chat_id:
                    bot_token = auth.get_valid_access_token(bot_uid)
                    bot_graph = GraphClient(bot_token)
                    bot_graph.send_chat_message(chat_id, _format_section_for_teams(result, get_user_tz(_udir(uid))))
            except Exception as ne:
                print(f"[ExpensePoll] Teams notify failed for {uid}: {ne}")
        except Exception as e:
            msg = str(e)
            if not any(c in msg for c in ("502", "503", "504", "ConnectionError", "Timeout")):
                print(f"[ExpensePoll] Error for {uid}: {e}")


def _poll_meeting_prep_all_users():
    """Every 5 min, for each user with meeting.prep_enabled=true, run meeting_prep.
    For each new prep item, push its teams_message directly to the user's Audrey chat."""
    sessions_dir = auth.DATA_DIR / "_sessions"
    if not sessions_dir.exists():
        return
    for token_file in sessions_dir.glob("*.json"):
        uid = token_file.stem
        bot_path = _bot_state_path(uid)
        if bot_path.exists() and _is_bound_bot(_read_json(bot_path)):
            continue
        try:
            from src.modules.schedules import load_schedules
            sched = load_schedules(_udir(uid))
            if not sched.get("meeting", {}).get("prep_enabled"):
                continue
            owner_graph = GraphClient(auth.get_valid_access_token(uid))
            ai          = AIClient()
            settings    = _read_json(_user_settings(uid))
            result      = meeting_prep.run(owner_graph, ai, _udir(uid), settings)
            preps       = result.get("items", [])
            if not preps:
                continue
            print(f"[MeetingPrep] {len(preps)} new prep(s) for {uid}")
            try:
                bot_uid, chat_id = _find_bot_for_user(uid)
                if bot_uid and chat_id:
                    bot_token = auth.get_valid_access_token(bot_uid)
                    bot_graph = GraphClient(bot_token)
                    for p in preps:
                        msg = p.get("teams_message", "")
                        if msg:
                            bot_graph.send_chat_message(chat_id, msg)
            except Exception as ne:
                print(f"[MeetingPrep] Teams notify failed for {uid}: {ne}")
        except Exception as e:
            msg = str(e)
            if not any(c in msg for c in ("502", "503", "504", "ConnectionError", "Timeout")):
                print(f"[MeetingPrep] Error for {uid}: {e}")


def _poll_meeting_recordings_all_users():
    """Every 20 min, scan each user's OneDrive Recordings/ for newly added mp4s.
    Processes at most 3 new recordings per user per cycle (Gemini transcription is slow + expensive).
    Already-processed mp4s are skipped via wiki/_index.json.
    For each newly processed meeting, also pushes a summary card to the user's
    Audrey 1:1 chat (gated by sched.meeting.summary_enabled, default true)."""
    from src.modules.m03_meeting import run as m03_run, format_meeting_summary_html
    from src.modules.schedules import load_schedules
    sessions_dir = auth.DATA_DIR / "_sessions"
    if not sessions_dir.exists():
        return
    for token_file in sessions_dir.glob("*.json"):
        uid = token_file.stem
        bot_path = _bot_state_path(uid)
        # Skip bot accounts (they don't have their own OneDrive recordings to process)
        if bot_path.exists() and _is_bound_bot(_read_json(bot_path)):
            continue
        try:
            token    = auth.get_valid_access_token(uid)
            graph    = GraphClient(token)
            ai       = AIClient()
            settings = _read_json(_user_settings(uid))
            result   = m03_run(graph, ai, _udir(uid), settings=settings,
                               months=6, max_to_process=3)
            n = result.get("processed", 0)
            if not n:
                continue
            print(f"[m03 poll] Processed {n} new recording(s) for {uid}")

            # Push summary to Teams unless the user disabled it
            try:
                sched = load_schedules(_udir(uid))
                if not sched.get("meeting", {}).get("summary_enabled", True):
                    continue
                bot_uid, chat_id = _find_bot_for_user(uid)
                if not (bot_uid and chat_id):
                    print(f"[m03 poll] No bot bound for {uid} — skipping Teams summary")
                    continue
                bot_token = auth.get_valid_access_token(bot_uid)
                bot_graph = GraphClient(bot_token)
                for rec in result.get("results", []):
                    if rec.get("status") != "processed":
                        continue
                    html = format_meeting_summary_html(rec)
                    bot_graph.send_html_message(chat_id, html)
                    print(f"[m03 poll] Teams summary sent: {rec.get('meeting_id','')} — {rec.get('title','')[:50]}")
            except Exception as ne:
                print(f"[m03 poll] Teams summary failed for {uid}: {ne}")
        except Exception as e:
            msg = str(e)
            if not any(c in msg for c in ("502", "503", "504", "ConnectionError", "Timeout")):
                print(f"[m03 poll] Error for {uid}: {e}")


# ── Teams bot API ─────────────────────────────────────────

@app.get("/api/teams/bot")
def get_bot_status(session: dict = Depends(require_session)):
    uid   = session["user_id"]
    path  = _bot_state_path(uid)
    state = json.loads(path.read_text()) if path.exists() else {}

    # Also check if a *separate* bot account is bound to this user
    bot_uid, chat_id = _find_bot_for_user(uid)
    if bot_uid:
        bp  = _bot_state_path(bot_uid)
        bs  = json.loads(bp.read_text()) if bp.exists() else {}
        bot_email = (auth.load_user_tokens(bot_uid) or {}).get("username", "")
        bot_health = auth.get_auth_health(bot_uid)
        return {
            "enabled":      True,
            "connected":    bot_health.get("status") != "broken",
            "peer_email":   bs.get("peer_email", ""),
            "bot_email":    bot_email,
            "chat_id":      chat_id,
            "last_seen_ts": bs.get("last_seen_ts"),
        }

    return {
        "enabled":      state.get("enabled", False),
        "connected":    True,
        "peer_email":   state.get("peer_email", ""),
        "bot_email":    "",
        "chat_id":      state.get("chat_id"),
        "last_seen_ts": state.get("last_seen_ts"),
    }


@app.post("/api/teams/bot/auth-start")
def bot_auth_start(session: dict = Depends(require_session)):
    # See src/ai.py::_call_with_timeout for why the
    # `with ThreadPoolExecutor` + future.result(timeout=...) pattern is broken:
    # the with-block calls shutdown(wait=True) on exit, which blocks on the
    # hung worker even after the future times out — so the timeout is fake.
    # Use the same daemon-thread fix as ai.py.
    global _bot_device_flow
    with _bot_device_flow_lock:
        result: dict = {}
        def runner():
            try:
                result["value"] = auth.start_device_flow()
            except BaseException as e:
                result["error"] = e
        t = threading.Thread(target=runner, daemon=True)
        t.start()
        t.join(timeout=20)
        if t.is_alive():
            raise HTTPException(504, "Timeout connecting to Microsoft — please retry")
        if "error" in result:
            raise result["error"]
        flow = result["value"]
        _bot_device_flow = flow
        return {
            "user_code":        flow["user_code"],
            "verification_url": flow["verification_uri"],
            "expires_in":       flow.get("expires_in", 900),
        }


@app.post("/api/teams/bot/auth-poll")
def bot_auth_poll(session: dict = Depends(require_session)):
    """Single-shot OAuth poll — returns immediately with pending/success/error.

    We can't use MSAL's auth.complete_device_flow() here because it's a
    blocking call that waits until the user completes auth or the code
    expires (5–15 min). On Railway/Cloudflare that connection gets
    idle-killed mid-flight (ERR_HTTP2_PROTOCOL_ERROR), and the frontend
    polling loop dies. Hitting the raw token endpoint returns in <1s with
    'authorization_pending' until the user completes, then with the token."""
    global _bot_device_flow
    with _bot_device_flow_lock:
        if not _bot_device_flow:
            return {"status": "no_flow"}

        import requests as _req
        try:
            r = _req.post(
                "https://login.microsoftonline.com/common/oauth2/v2.0/token",
                data={
                    "client_id":   auth.CLIENT_ID,
                    "grant_type":  "urn:ietf:params:oauth:grant-type:device_code",
                    "device_code": _bot_device_flow.get("device_code", ""),
                },
                timeout=8,
            )
            result = r.json()
        except Exception as e:
            return {"status": "error", "message": f"Token poll failed: {e}"}

        if "error" in result:
            err = result["error"]
            if err in ("authorization_pending", "slow_down"):
                return {"status": "pending"}
            if err == "expired_token":
                _bot_device_flow.clear()
                return {"status": "expired", "message": "Device code expired"}
            return {"status": "error", "message": result.get("error_description", err)}

        token = result.get("access_token")
        if not token:
            return {"status": "error", "message": "No access_token in response"}

        _bot_device_flow.clear()

        try:
            me = _req.get(
                "https://graph.microsoft.com/v1.0/me",
                headers={"Authorization": f"Bearer {token}"},
                timeout=10,
            ).json()
        except Exception as e:
            return {"status": "error", "message": f"Graph /me failed: {e}"}

        bot_uid   = me.get("id", "")
        bot_email = me.get("mail") or me.get("userPrincipalName", "")
        if bot_uid and bot_uid == session["user_id"]:
            # Self-registration guard. The bot role must never land on the
            # user's OWN account. activate_bot rejects a self-bind too, but only
            # AFTER the flag below is written — leaving a permanent zombie that
            # every per-user event poll then skips. Reject here, before any write.
            _bot = _get_bot_display_name(session["user_id"])
            return {"status": "error", "message":
                    "You signed in as your own account. The AI assistant must be "
                    f"a separate Microsoft account (e.g. {_bot.lower()}@yourcompany.com). "
                    "Restart setup and sign in as the assistant account."}
        if bot_uid:
            expiry = (datetime.now(timezone.utc) + timedelta(hours=1)).isoformat()

            # Push the raw token response into the MSAL cache so the
            # device-flow refresh path (Path 2 in get_valid_access_token) can
            # find it when the access_token expires ~1 hour from now. Without
            # this, the session would keep a CLIENT_ID-issued refresh_token
            # that Path 1 (PROD_CLIENT_ID + PROD_SECRET) can't refresh —
            # Microsoft responds with AADSTS7000215 because the secret
            # doesn't match the issuing app — and Path 2 has no cache entry
            # to fall back to. Net result: bot dies after 1h. See
            # docs/auth-health-review-findings.md and conversation around
            # commit 2c74f2b for the full root-cause analysis.
            try:
                with auth.file_lock(auth._bot_cache_file(bot_uid)):
                    _cache = auth._load_cache(bot_uid)
                    _cache.add({
                        "client_id":      auth.CLIENT_ID,
                        "scope":          (result.get("scope") or "").split() or list(auth.SCOPES_LOCAL),
                        "token_endpoint": "https://login.microsoftonline.com/common/oauth2/v2.0/token",
                        "response":       result,
                        "data":           {},
                        "grant_type":     "urn:ietf:params:oauth:grant-type:device_code",
                    })
                    auth._save_cache(_cache, bot_uid)
            except Exception as e:
                # If injection fails, log loudly. First hour still works on
                # the access_token we already have; the bot will simply
                # appear broken after that, same as before this fix.
                print(f"[BotAuth] MSAL cache injection failed for {bot_uid} ({bot_email}): {e}")

            # Save the session with an EMPTY refresh_token so the next
            # get_valid_access_token call goes straight to the MSAL silent
            # path (Path 2). If we left the raw refresh_token in here, Path 1
            # would try first and fail with AADSTS7000215 on every refresh
            # — wasted Graph quota and noise in the diag log.
            auth.save_user_tokens(bot_uid, {
                "access_token":  token,
                "refresh_token": "",
                "expiry":        expiry,
                "username":      bot_email,
            })
            bp = _bot_state_path(bot_uid)
            bs = json.loads(bp.read_text()) if bp.exists() else {}
            bs["enabled"]           = True
            bs["is_registered_bot"] = True
            _write_json(bp, bs)

        return {"status": "success", "bot_email": bot_email, "bot_uid": bot_uid}


@app.post("/api/teams/bot/activate")
def activate_bot(background_tasks: BackgroundTasks, session: dict = Depends(require_session),
                 bot_uid: str = None):
    user_id  = session["user_id"]
    username = session["username"]
    if not bot_uid:
        raise HTTPException(400, "bot_uid is required")
    if bot_uid == user_id:
        # The AI assistant must be a separate Microsoft account. Self-binding
        # creates a confusing "you chatting with yourself" state and loses the
        # two-account failure isolation the architecture relies on.
        _bot = _get_bot_display_name(user_id)
        raise HTTPException(
            400,
            "The AI assistant must be a different Microsoft account than your "
            "own. When prompted at microsoft.com/devicelogin, sign in as your "
            f"AI assistant account (e.g. {_bot.lower()}@yourcompany.com), not your "
            "personal account.",
        )
    bp = _bot_state_path(bot_uid)
    if not bp.exists():
        raise HTTPException(404, "Bot account not found")
    bs = json.loads(bp.read_text())
    if not bs.get("enabled"):
        raise HTTPException(400, "Bot account is not enabled")
    existing_owner = bs.get("owner_uid")
    if existing_owner and existing_owner != user_id:
        raise HTTPException(409, "This bot is already claimed by another user")

    _bind_bot_to_user(bot_uid, user_id, username)
    bot_email    = (auth.load_user_tokens(bot_uid) or {}).get("username", "bot")
    display_name = username.split("@")[0].replace(".", " ").title()
    background_tasks.add_task(_send_activation_greeting, bot_uid, username, display_name)

    return {"ok": True, "message": f"AI assistant ({bot_email}) is now monitoring {username}"}


@app.post("/api/teams/bot/disable")
def disable_bot(session: dict = Depends(require_session)):
    uid = session["user_id"]
    bot_uid = _unbind_bot_from_user(uid)
    # Disable the BOT (its teams_bot.json), not the owner's. Without this the
    # unbound bot still passes the `enabled && is_registered_bot` check in the
    # poll loop and would keep running with owner_uid=None.
    if bot_uid:
        bp = _bot_state_path(bot_uid)
        if bp.exists():
            bs = json.loads(bp.read_text())
            bs["enabled"] = False
            bs["is_registered_bot"] = False
            _write_json(bp, bs)
    return {"ok": True}


# ── Auth health (for the AI assistant status indicator) ──

def _account_health_dot(user_id: str, label: str) -> dict:
    from src.auth_errors import classify_aadsts
    h = auth.get_auth_health(user_id)
    status_raw = h.get("status", "healthy")
    consec = h.get("consecutive_failures", 0)
    if status_raw == "broken":
        dot = "red"
    elif consec > 0:
        dot = "yellow"
    else:
        dot = "green"

    classification = None
    if status_raw == "broken" and h.get("last_error"):
        err = h["last_error"]
        c = classify_aadsts(err.get("code"), err.get("description") or "")
        classification = {
            "message": c["message"],
            "action": c["action"],
            "code": c.get("code"),
        }

    return {
        "label":               label,
        "user_id":             user_id,
        "dot":                 dot,
        "consecutive_failures": consec,
        "last_success_at":     h.get("last_success_at"),
        "last_failure_at":     h.get("last_failure_at"),
        "broken_since":        h.get("broken_since"),
        "classification":      classification,
    }


@app.get("/api/auth/health")
def get_auth_health_endpoint(session: dict = Depends(require_session)):
    uid = session["user_id"]
    accounts = [_account_health_dot(uid, "Your account")]

    bot_uid, _ = _find_bot_for_user(uid)
    if bot_uid:
        bot_email = (auth.load_user_tokens(bot_uid) or {}).get("username", "")
        accounts.append(_account_health_dot(bot_uid, f"AI assistant ({bot_email})" if bot_email else "AI assistant"))

    rank = {"green": 0, "yellow": 1, "red": 2}
    overall = max((a["dot"] for a in accounts), key=lambda d: rank.get(d, 0))

    return {"overall": overall, "accounts": accounts}


def _tail_file(path: Path, lines: int) -> list[str]:
    if not path.exists():
        return []
    try:
        with open(path) as f:
            all_lines = f.readlines()
        return [l.rstrip("\n") for l in all_lines[-lines:]]
    except Exception as e:
        return [f"(read failed: {e})"]


@app.get("/api/admin/diag/{user_id}")
def admin_user_diag(user_id: str, lines: int = 200, session: dict = Depends(require_session)):
    """Returns everything we know about a single user's auth state — health
    record, token expiry summary, last N lines of diag log, last N dead-letter
    entries. Use when a user reports their AI assistant is offline."""
    lines = max(1, min(lines, 2000))

    health = auth.get_auth_health(user_id)

    tokens = auth.load_user_tokens(user_id) or {}
    token_summary = {
        "username":          tokens.get("username", ""),
        "has_refresh_token": bool(tokens.get("refresh_token")),
        "expiry":            tokens.get("expiry"),
    }

    bot_state_path = auth.DATA_DIR / user_id / "teams_bot.json"
    bot_state = None
    if bot_state_path.exists():
        try:
            bot_state = json.loads(bot_state_path.read_text())
        except Exception:
            pass

    diag_path = auth.DATA_DIR / user_id / ".auth_diag.log"
    deadletter_path = auth.DATA_DIR / user_id / ".auth_notifier_deadletter.log"

    return {
        "user_id":          user_id,
        "health":           health,
        "token":            token_summary,
        "bot_state":        bot_state,
        "diag_log_tail":    _tail_file(diag_path, lines),
        "deadletter_tail":  _tail_file(deadletter_path, min(lines, 100)),
    }


# ── Graph test endpoint ───────────────────────────────────

@app.get("/api/admin/bot-bindings")
def admin_bot_bindings(session: dict = Depends(require_session)):
    sessions_dir = auth.DATA_DIR / "_sessions"
    results = []
    for tf in sessions_dir.glob("*.json"):
        bid = tf.stem
        bp  = _bot_state_path(bid)
        if not bp.exists():
            continue
        bs        = json.loads(bp.read_text())
        bot_email = (auth.load_user_tokens(bid) or {}).get("username", "")
        owner_uid = bs.get("owner_uid")
        owner_email = ""
        if owner_uid:
            ot = auth.load_user_tokens(owner_uid) or {}
            owner_email = ot.get("username", owner_uid)
        results.append({
            "bot_uid":      bid,
            "bot_email":    bot_email,
            "enabled":      bs.get("enabled", False),
            "owner_uid":    owner_uid,
            "owner_email":  owner_email,
            "peer_email":   bs.get("peer_email", ""),
            "last_seen_ts": bs.get("last_seen_ts"),
        })
    return results


@app.post("/api/admin/bot/unbind/{bot_uid}")
def admin_unbind_bot(bot_uid: str, session: dict = Depends(require_session)):
    bp = _bot_state_path(bot_uid)
    if not bp.exists():
        raise HTTPException(404, "Bot not found")
    bs = json.loads(bp.read_text())
    prev_owner = bs.get("owner_uid")
    bs.update({
        "enabled":           False,
        "is_registered_bot": False,
        "owner_uid":         None,
        "peer_email":        None,
        "chat_id":           None,
        "last_seen_ts":      None,
    })
    _write_json(bp, bs)
    if prev_owner:
        _user_bot_link_path(prev_owner).unlink(missing_ok=True)
    return {"ok": True, "bot_uid": bot_uid, "message": "Bot unbound and disabled"}


def _build_user_health_snapshot(uid: str) -> dict:
    """Aggregate per-user health for the ops dashboard. Read-only — no
    side effects. Returns flat dict that the dashboard can render directly.
    Every disk read is guarded; a single corrupt file should not break the
    whole snapshot."""
    tokens = auth.load_user_tokens(uid) or {}
    health = auth.get_auth_health(uid)

    bot_state = None
    bot_state_path = auth.DATA_DIR / uid / "teams_bot.json"
    if bot_state_path.exists():
        try:
            bot_state = json.loads(bot_state_path.read_text())
        except Exception:
            bot_state = None

    is_bot = bool(bot_state and bot_state.get("is_registered_bot"))
    role = "bot" if is_bot else "user"

    linked_bot_uid = None
    linked_bot_email = None
    if not is_bot:
        link_path = _user_bot_link_path(uid)
        if link_path.exists():
            try:
                link = json.loads(link_path.read_text())
                linked_bot_uid = link.get("bot_uid")
                if linked_bot_uid:
                    bt = auth.load_user_tokens(linked_bot_uid) or {}
                    linked_bot_email = bt.get("username", "")
            except Exception:
                pass

    owner_email = None
    if is_bot:
        owner_uid = bot_state.get("owner_uid") if bot_state else None
        if owner_uid:
            ot = auth.load_user_tokens(owner_uid) or {}
            owner_email = ot.get("username", "")

    briefings_enabled = 0
    briefings_total = 0
    try:
        sched_path = auth.DATA_DIR / uid / "schedules.json"
        if sched_path.exists():
            sched = json.loads(sched_path.read_text())
            briefings = sched.get("briefings") or []
            briefings_total = len(briefings)
            briefings_enabled = sum(1 for b in briefings if b.get("enabled"))
    except Exception:
        pass

    last_briefing_run = None
    try:
        ai_summary_path = auth.DATA_DIR / uid / "results" / "ai_summary.json"
        if ai_summary_path.exists():
            r = json.loads(ai_summary_path.read_text())
            last_briefing_run = r.get("last_run")
    except Exception:
        pass

    diag_tail = _tail_file(auth.DATA_DIR / uid / ".auth_diag.log", 5)

    return {
        "uid":               uid,
        "username":          tokens.get("username", ""),
        "role":              role,
        "owner_email":       owner_email,
        "linked_bot_uid":    linked_bot_uid,
        "linked_bot_email":  linked_bot_email,
        "health": {
            "status":               health.get("status"),
            "consecutive_failures": health.get("consecutive_failures", 0),
            "last_success_at":      health.get("last_success_at"),
            "last_failure_at":      health.get("last_failure_at"),
            "last_error":           health.get("last_error"),
            "broken_since":         health.get("broken_since"),
        },
        "bot": {
            "enabled":      bool(bot_state and bot_state.get("enabled")),
            "is_bot":       is_bot,
            "chat_id":      (bot_state or {}).get("chat_id"),
            "last_seen_ts": (bot_state or {}).get("last_seen_ts"),
        } if bot_state else None,
        "briefings": {
            "total":            briefings_total,
            "enabled":          briefings_enabled,
            "last_summary_run": last_briefing_run,
        },
        "diag_log_tail": diag_tail,
    }


@app.get("/api/admin/fleet-health")
def admin_fleet_health(x_admin_token: str | None = Header(None, alias="X-Admin-Token")):
    """Aggregated health snapshot for the ops dashboard. Single call returns
    every user/bot so the ops poller doesn't hit N endpoints. Auth: shared
    secret in env var OPS_ADMIN_TOKEN sent as X-Admin-Token header — this
    endpoint is bot-to-bot, not browser-facing, so no session cookie."""
    expected = os.getenv("OPS_ADMIN_TOKEN")
    if not expected:
        raise HTTPException(503, "OPS_ADMIN_TOKEN not configured on this server")
    if not x_admin_token or not secrets.compare_digest(x_admin_token, expected):
        raise HTTPException(401, "invalid admin token")

    sessions_dir = auth.DATA_DIR / "_sessions"
    users = []
    if sessions_dir.exists():
        for sf in sorted(sessions_dir.glob("*.json")):
            try:
                users.append(_build_user_health_snapshot(sf.stem))
            except Exception as e:
                users.append({"uid": sf.stem, "error": str(e)})

    healthy = sum(1 for u in users if (u.get("health") or {}).get("status") == "healthy")
    broken  = sum(1 for u in users if (u.get("health") or {}).get("status") == "broken")

    # Scheduler job health (stream 2) — flat list the ops side checks for staleness.
    jobs = []
    try:
        jh = _read_json(_JOB_HEALTH_PATH)
        for jid, rec in sorted(jh.items()):
            jobs.append({"job_id": jid, **(rec if isinstance(rec, dict) else {})})
    except Exception:
        pass

    # Push-quality roll-up (stream 4) — present only once push_qa has run.
    push_qa = None
    try:
        pq = _read_json(_OPS_DIR / "push_qa.json")
        if pq:
            push_qa = pq
    except Exception:
        pass

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "user_count":   len(users),
        "healthy":      healthy,
        "broken":       broken,
        "users":        users,
        "jobs":         jobs,
        "push_qa":      push_qa,
    }


@app.get("/api/test/graph")
def test_graph(session: dict = Depends(require_session)):
    uid = session["user_id"]
    try:
        token = auth.get_valid_access_token(uid)
        graph = GraphClient(token)
        me    = graph.get_me()
        msgs  = graph.get_messages(top=3)
        return {
            "ok":   True,
            "user": {
                "displayName":       me.get("displayName"),
                "mail":              me.get("mail"),
                "userPrincipalName": me.get("userPrincipalName"),
            },
            "latest_emails": [
                {
                    "subject":  m.get("subject"),
                    "from":     m.get("from", {}).get("emailAddress", {}).get("address"),
                    "received": m.get("receivedDateTime", "")[:16],
                }
                for m in msgs
            ],
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})


# ── First-login init chain ────────────────────────────────

def _make_progress_cb(uid: str):
    """Returns a callable that pipes progress messages into init_status.json."""
    from src.modules.profile import update_init_message
    def cb(msg: str):
        try:
            update_init_message(_udir(uid), msg)
        except Exception:
            pass
    return cb


def _generate_profile(uid: str, progress_cb=None, on_doc_ready=None) -> None:
    """Call run_profile_init for a user — no status side-effects.

    on_doc_ready: optional callback (doc_key, content) — onboarding uses this
        to reveal each of the 3 profile docs to the user one at a time as
        they're written.
    """
    from src.modules.profile_init import run_profile_init
    token    = auth.get_valid_access_token(uid)
    graph    = GraphClient(token)
    ai       = AIClient()
    settings = _read_json(_user_settings(uid))
    run_profile_init(_udir(uid), graph, ai, settings,
                     progress=progress_cb, on_doc_ready=on_doc_ready)


def _regenerate_profile_for_user(uid: str) -> None:
    """
    /api/profile/regenerate worker — runs in background.
    Preserves user_confirmed status so the user isn't kicked back to Onboarding.
    """
    from src.modules.profile import (
        save_init_status, load_init_status, reset_init_steps,
        update_init_step,
    )
    was_confirmed = load_init_status(_udir(uid)).get("stage") == "user_confirmed"
    reset_init_steps(_udir(uid))
    update_init_step(_udir(uid), "crm",      "done")  # already done, skip
    update_init_step(_udir(uid), "projects", "done")
    update_init_step(_udir(uid), "profile",  "in_progress", "Re-drafting profile from current data...")
    try:
        _generate_profile(uid, progress_cb=_make_progress_cb(uid))
        update_init_step(_udir(uid), "profile", "done", "Profile updated")
    except Exception as e:
        print(f"[profile_init] Regenerate failed for {uid}: {e}")
        update_init_step(_udir(uid), "profile", "failed", f"Failed: {e}")
    finally:
        # Restore prior status if user had already confirmed; otherwise expose the draft.
        save_init_status(_udir(uid), "user_confirmed" if was_confirmed else "draft_ready")


def _build_crm_with_progress(uid: str, progress_cb, months: int = 6) -> None:
    from src.modules.crm import build_crm, save_crm
    try:
        token  = auth.get_valid_access_token(uid)
        graph  = GraphClient(token)
        ai     = AIClient()
        seg, biz = _get_crm_ai_context(uid)
        result = build_crm(graph, ai, _udir(uid),
                           months=months,
                           market_segments_content=seg,
                           business_context=biz,
                           progress_cb=progress_cb)
        save_crm(_udir(uid), result)
        print(f"[CRM] Initial build done for {uid} — {len(result.get('contacts', {}))} contacts ({months}mo scan)")
    except Exception as e:
        print(f"[CRM] Initial build failed for {uid}: {e}")
        raise


def _build_projects_with_progress(uid: str, progress_cb, months: int = 6) -> None:
    from src.modules.projects import build_projects, save_projects
    try:
        token    = auth.get_valid_access_token(uid)
        graph    = GraphClient(token)
        ai       = AIClient()
        settings = _read_json(_user_settings(uid))
        result   = build_projects(graph, ai, _udir(uid), settings=settings,
                                  months=months,
                                  progress_cb=progress_cb)
        save_projects(_udir(uid), result)
        print(f"[Projects] Initial build done for {uid} — {len(result.get('projects', {}))} projects ({months}mo scan)")
    except Exception as e:
        print(f"[Projects] Initial build failed for {uid}: {e}")
        raise


def _sample_contact_names(uid: str, n: int = 3) -> tuple[int, list[str]]:
    """Return (total_count, [n sample display names]) from crm.json."""
    crm_path = _udir(uid) / "crm.json"
    if not crm_path.exists():
        return 0, []
    try:
        crm = json.loads(crm_path.read_text())
        contacts = list(crm.get("contacts", {}).values())
        contacts = [c for c in contacts if not (c.get("ignore") or c.get("archived"))]
        contacts.sort(key=lambda c: c.get("thread_count", 0), reverse=True)
        samples = [(c.get("name") or c.get("email") or "").strip() for c in contacts[:n]]
        samples = [s for s in samples if s]
        return len(contacts), samples
    except Exception:
        return 0, []


def _sample_project_names(uid: str, n: int = 3) -> tuple[int, list[str]]:
    proj_path = _udir(uid) / "projects.json"
    if not proj_path.exists():
        return 0, []
    try:
        data = json.loads(proj_path.read_text())
        projs = [p for p in data.get("projects", {}).values() if not (p.get("ignore") or p.get("archived"))]
        projs.sort(key=lambda p: p.get("thread_count", 0), reverse=True)
        samples = [(p.get("name") or "").strip() for p in projs[:n]]
        samples = [s for s in samples if s]
        return len(projs), samples
    except Exception:
        return 0, []


def _sample_company_names(uid: str, n: int = 3) -> tuple[int, list[str]]:
    comp_path = _udir(uid) / "companies.json"
    if not comp_path.exists():
        return 0, []
    try:
        data = json.loads(comp_path.read_text())
        comps = [c for c in data.get("companies", {}).values() if not c.get("ignore")]
        comps.sort(key=lambda c: c.get("thread_count_total", 0), reverse=True)
        samples = [(c.get("name") or (c.get("identity") or {}).get("original_name") or "").strip() for c in comps[:n]]
        samples = [s for s in samples if s]
        return len(comps), samples
    except Exception:
        return 0, []


def _preview_doc(text: str, limit: int = 240) -> str:
    """Trim to ~limit chars on a clean boundary for the reveal card."""
    text = (text or "").strip()
    if len(text) <= limit:
        return text
    cut = text[:limit].rsplit(" ", 1)[0]
    return cut.rstrip(".,;:—-") + "…"


def _run_first_login_init(uid: str, history_months: int | None = None) -> None:
    """
    Serial init chain on first OAuth login. Drives the 6 reveal phases
    shown in onboarding Step 5:

      1. crm                — Reading your inbox
      2. projects           — Detecting active projects
      3. companies          — Mapping companies
      4. personal_profile   — Drafting personal profile  ┐
      5. business_profile   — Drafting business profile  │  All from
      6. market_segments    — Identifying market segments┘  ONE AI call,
                                                            staggered for
                                                            sequential reveal.

    Each step writes a `reveal_data` payload (counts + samples for the 3 DBs,
    text preview for the 3 docs) that the frontend uses to animate the card.
    """
    from src.modules.profile import (
        save_init_status, reset_init_steps, update_init_step,
    )
    udir = _udir(uid)
    if history_months is None:
        settings = _read_json(_user_settings(uid))
        history_months = int(settings.get("onboarding_history_months") or 12)
    history_months = max(1, min(int(history_months), 24))

    progress_cb = _make_progress_cb(uid)
    try:
        save_init_status(udir, "generating")
        reset_init_steps(udir)

        # ── Phase 1: CRM ──────────────────────────────────
        if (udir / "crm.json").exists():
            n, samples = _sample_contact_names(uid)
            update_init_step(udir, "crm", "done", "CRM already built",
                             reveal_data={"count": n, "samples": samples})
        else:
            update_init_step(udir, "crm", "in_progress", f"Scanning inbox ({history_months}mo)...")
            try:
                _build_crm_with_progress(uid, progress_cb, months=history_months)
                n, samples = _sample_contact_names(uid)
                update_init_step(udir, "crm", "done", f"{n} contacts indexed",
                                 reveal_data={"count": n, "samples": samples})
            except Exception as e:
                update_init_step(udir, "crm", "failed", f"CRM scan failed: {e}")

        # ── Phase 2: Projects ─────────────────────────────
        if (udir / "projects.json").exists():
            n, samples = _sample_project_names(uid)
            update_init_step(udir, "projects", "done", "Projects already built",
                             reveal_data={"count": n, "samples": samples})
        else:
            update_init_step(udir, "projects", "in_progress",
                             f"Identifying active projects ({history_months}mo)...")
            try:
                _build_projects_with_progress(uid, progress_cb, months=history_months)
                n, samples = _sample_project_names(uid)
                update_init_step(udir, "projects", "done", f"{n} active projects detected",
                                 reveal_data={"count": n, "samples": samples})
            except Exception as e:
                update_init_step(udir, "projects", "failed", f"Project scan failed: {e}")

        # ── Phase 3: Companies (no AI, derived from CRM + Projects) ──
        update_init_step(udir, "companies", "in_progress", "Aggregating company list...")
        try:
            _build_companies_for_user(uid)
            n, samples = _sample_company_names(uid)
            update_init_step(udir, "companies", "done", f"{n} companies mapped",
                             reveal_data={"count": n, "samples": samples})
        except Exception as e:
            print(f"[Companies] Init build failed for {uid}: {e}")
            update_init_step(udir, "companies", "failed", f"Company build failed: {e}")

        # ── Phases 4-6: Profile docs (run_profile_init internal phases) ──
        # business_profile uses a two-pass generation (website-only v1 then
        # CRM-enriched v2) — only the v2 fires on_doc_ready('business'), so
        # the user sees a single ✓ after the final pass. personal + segments
        # come from one combined AI call between v1 and v2.
        update_init_step(udir, "personal_profile", "in_progress", "Drafting your personal profile...")
        update_init_step(udir, "business_profile", "in_progress", "")
        update_init_step(udir, "market_segments",  "in_progress", "")

        def _on_doc_ready(doc_key: str, content: str):
            mapping = {
                "personal": ("personal_profile", "Personal profile drafted"),
                "business": ("business_profile", "Business profile drafted"),
                "segments": ("market_segments",  "Market segments mapped"),
            }
            step_key, msg = mapping.get(doc_key, (doc_key, "Done"))
            update_init_step(udir, step_key, "done", msg,
                             reveal_data={"preview": _preview_doc(content)})
            time.sleep(0.6)  # stagger so cards animate one at a time

        try:
            _generate_profile(uid, progress_cb=progress_cb, on_doc_ready=_on_doc_ready)
            # Any docs the AI failed to produce are still in_progress — mark failed.
            from src.modules.profile import load_init_status as _load
            for s in _load(udir).get("steps", []):
                if s["key"] in ("personal_profile", "business_profile", "market_segments") \
                   and s.get("status") == "in_progress":
                    update_init_step(udir, s["key"], "failed",
                                     "AI didn't return this section")
        except Exception as e:
            print(f"[profile_init] Generate failed for {uid}: {e}")
            for key in ("personal_profile", "business_profile", "market_segments"):
                update_init_step(udir, key, "failed", f"Draft generation failed: {e}")

        # Background: meeting recordings backfill — doesn't block onboarding
        threading.Thread(
            target=_run_meeting_backfill_for_user,
            args=(uid,),
            daemon=True,
        ).start()
        print(f"[init] Meeting backfill thread started for {uid} (runs in parallel; not in onboarding gate)")
    finally:
        save_init_status(udir, "draft_ready")


def _run_meeting_backfill_for_user(uid: str) -> None:
    """One-time scan of the user's OneDrive Recordings/ for mp4s from the last 6 months.
    Runs in a background thread on first login. Idempotent — already-processed
    recordings are skipped via wiki/_index.json."""
    from src.modules.m03_meeting import run as m03_run
    try:
        token    = auth.get_valid_access_token(uid)
        graph    = GraphClient(token)
        ai       = AIClient()
        settings = _read_json(_user_settings(uid))
        result = m03_run(graph, ai, _udir(uid), settings=settings, months=6)
        print(f"[m03 backfill] {uid} done — {result.get('processed', 0)} processed, "
              f"{result.get('skipped', 0)} already in index")
    except Exception as e:
        print(f"[m03 backfill] {uid} failed: {e}")


# ── CRM lifecycle helpers ─────────────────────────────────

def _get_market_segments(uid: str) -> str:
    from src.modules.profile import load_market_segments
    return load_market_segments(_udir(uid))


def _get_crm_ai_context(uid: str) -> tuple[str, str]:
    """Returns (market_segments_content, business_context) for CRM enrichment.
    business_context = personal + business profile + market segments combined,
    which lets the AI reason about same-domain=internal and buyer/seller direction."""
    from src.modules.profile import load_market_segments, load_profile_context
    return load_market_segments(_udir(uid)), load_profile_context(_udir(uid))


def _build_crm_for_user(uid: str) -> None:
    from src.modules.crm import build_crm, save_crm
    try:
        token  = auth.get_valid_access_token(uid)
        graph  = GraphClient(token)
        ai     = AIClient()
        seg, biz = _get_crm_ai_context(uid)
        result = build_crm(graph, ai, _udir(uid),
                           market_segments_content=seg,
                           business_context=biz)
        save_crm(_udir(uid), result)
        print(f"[CRM] Initial build done for {uid} — {len(result.get('contacts', {}))} contacts")
    except Exception as e:
        print(f"[CRM] Initial build failed for {uid}: {e}")


def _refresh_crm_for_user(uid: str) -> None:
    from src.modules.crm import refresh_crm, save_crm
    if not (_udir(uid) / "crm.json").exists():
        return
    try:
        token  = auth.get_valid_access_token(uid)
        graph  = GraphClient(token)
        ai     = AIClient()
        seg, biz = _get_crm_ai_context(uid)
        result = refresh_crm(graph, ai, _udir(uid),
                             market_segments_content=seg,
                             business_context=biz)
        save_crm(_udir(uid), result)
        print(f"[CRM] Daily refresh done for {uid}")
    except Exception as e:
        print(f"[CRM] Daily refresh failed for {uid}: {e}")


def _refresh_crm_all_users() -> None:
    sessions_dir = auth.DATA_DIR / "_sessions"
    if not sessions_dir.exists():
        return
    for f in sessions_dir.glob("*.json"):
        threading.Thread(target=_refresh_crm_for_user, args=(f.stem,), daemon=True).start()


# ── Projects lifecycle helpers ────────────────────────────

def _build_projects_for_user(uid: str) -> None:
    from src.modules.projects import build_projects, save_projects
    try:
        token    = auth.get_valid_access_token(uid)
        graph    = GraphClient(token)
        ai       = AIClient()
        settings = _read_json(_user_settings(uid))
        result   = build_projects(graph, ai, _udir(uid), settings=settings)
        save_projects(_udir(uid), result)
        print(f"[Projects] Initial build done for {uid} — {len(result.get('projects', {}))} projects")
    except Exception as e:
        print(f"[Projects] Initial build failed for {uid}: {e}")


def _refresh_projects_for_user(uid: str) -> None:
    from src.modules.projects import refresh_projects, save_projects
    if not (_udir(uid) / "projects.json").exists():
        return
    try:
        token    = auth.get_valid_access_token(uid)
        graph    = GraphClient(token)
        ai       = AIClient()
        settings = _read_json(_user_settings(uid))
        result   = refresh_projects(graph, ai, _udir(uid), settings=settings)
        save_projects(_udir(uid), result)
        print(f"[Projects] Daily refresh done for {uid}")
    except Exception as e:
        print(f"[Projects] Daily refresh failed for {uid}: {e}")


def _refresh_projects_all_users() -> None:
    sessions_dir = auth.DATA_DIR / "_sessions"
    if not sessions_dir.exists():
        return
    for f in sessions_dir.glob("*.json"):
        threading.Thread(target=_refresh_projects_for_user, args=(f.stem,), daemon=True).start()


# ── Companies lifecycle helpers ───────────────────────────
# Companies is a derived aggregation of CRM contacts + Projects participants.
# No AI calls, no Graph API calls — cheap enough that we just rebuild it on
# every refresh instead of bothering with incremental logic. User-editable
# fields (monitor_intelligence, ignore, notes, priority, name) are preserved
# across rebuilds by companies.build_companies itself.

def _build_companies_for_user(uid: str) -> None:
    from src.modules.companies import build_companies, save_companies
    try:
        result = build_companies(_udir(uid))
        save_companies(_udir(uid), result)
        print(f"[Companies] Build done for {uid} — {len(result.get('companies', {}))} companies")
    except Exception as e:
        print(f"[Companies] Build failed for {uid}: {e}")


def _refresh_companies_all_users() -> None:
    sessions_dir = auth.DATA_DIR / "_sessions"
    if not sessions_dir.exists():
        return
    for f in sessions_dir.glob("*.json"):
        threading.Thread(target=_build_companies_for_user, args=(f.stem,), daemon=True).start()


# ── Startup cleanup: stuck "running" sections ────────────
# A BackgroundTasks job that's mid-execution when the container is killed
# (Railway redeploy, OOM, etc.) dies without writing a final status, leaving
# the result file pinned at status="running" forever — visible to the user as
# a hung run with no clean recovery path. Daniel hit this on 2026-06-01 when
# a redeploy interrupted his company_intelligence run.
#
# Fix: at startup, scan results/*.json for any file still in "running" and
# flip it to "error". The new process has no in-flight tasks yet, so any
# "running" marker is by definition from a dead previous process.

def _reset_stuck_running_results() -> int:
    """Reset result files left in status='running' by a previous process.
    Returns count of files reset.

    Scope: only `.data/{uid}/results/*.json`. Does not touch init_status,
    email_monitor state, bot state, or anything outside results/."""
    reset_count = 0
    try:
        for results_dir in auth.DATA_DIR.glob("*/results"):
            if not results_dir.is_dir():
                continue
            for f in results_dir.glob("*.json"):
                try:
                    data = json.loads(f.read_text())
                except Exception:
                    continue
                if not isinstance(data, dict) or data.get("status") != "running":
                    continue
                section_id = data.get("id") or f.stem
                started_at = data.get("started_at") or ""
                logs       = data.get("logs") or []
                _write_json(f, {
                    "id":         section_id,
                    "status":     "error",
                    "error":      ("Previous run was interrupted by service "
                                   "restart or crash. Trigger again to retry."),
                    "started_at": started_at,
                    "last_run":   datetime.now(timezone.utc).isoformat(),
                    "logs":       logs,
                    "items":      [],
                    "count":      0,
                    "empty":      True,
                })
                reset_count += 1
                try:
                    print(f"[Startup] Reset stuck running: {f.relative_to(auth.DATA_DIR)}", flush=True)
                except ValueError:
                    print(f"[Startup] Reset stuck running: {f}", flush=True)
    except Exception as e:
        print(f"[Startup] Reset stuck running failed: {e}", flush=True)
    if reset_count:
        print(f"[Startup] Reset {reset_count} stuck 'running' result file(s)", flush=True)
    return reset_count


# ── Section registry ─────────────────────────────────────
# Add entries here as sections are implemented.

_SECTION_RUNNERS: dict[str, object] = {
    "ai_summary": lambda graph, ai, data_dir, settings, progress:
        ai_summary.run(graph, ai, data_dir, settings, progress,
                       runner_lookup=lambda sid: _SECTION_RUNNERS.get(sid)),
    "reply_needed": lambda graph, ai, data_dir, settings, progress:
        reply_needed.run(graph, ai, data_dir, settings, progress),
    "followup_needed": lambda graph, ai, data_dir, settings, progress:
        followup_needed.run(graph, ai, data_dir, settings, progress),
    "commitments_extract": lambda graph, ai, data_dir, settings, progress:
        commitments_extract.run(graph, ai, data_dir, settings, progress),
    "upcoming_commitments": lambda graph, ai, data_dir, settings, progress:
        upcoming_commitments.run(graph, ai, data_dir, settings, progress),
    "expenses": lambda graph, ai, data_dir, settings, progress:
        expenses.run(graph, ai, data_dir, progress=progress),
    "recent_meetings": lambda graph, ai, data_dir, settings, progress:
        recent_meetings.run(data_dir),
    "meeting_action_items": lambda graph, ai, data_dir, settings, progress:
        meeting_action_items.run(data_dir),
    "market_intelligence": lambda graph, ai, data_dir, settings, progress:
        market_intelligence.run(graph, ai, data_dir, settings, progress),
    "company_intelligence": lambda graph, ai, data_dir, settings, progress:
        company_intelligence.run(graph, ai, data_dir, settings, progress),
    "relationship_health": lambda graph, ai, data_dir, settings, progress:
        relationship_health.run(graph, ai, data_dir, settings, progress),
    "business_insights": lambda graph, ai, data_dir, settings, progress:
        business_insights.run(graph, ai, data_dir, settings, progress),
    "meetings_today": lambda graph, ai, data_dir, settings, progress:
        meetings_today.run(graph, ai, data_dir, settings, progress),
    "due_today": lambda graph, ai, data_dir, settings, progress:
        due_today.run(graph, ai, data_dir, settings, progress),
    "yesterday_recap": lambda graph, ai, data_dir, settings, progress:
        yesterday_recap.run(graph, ai, data_dir, settings, progress),
    "projects_needing_attention": lambda graph, ai, data_dir, settings, progress:
        projects_needing_attention.run(graph, ai, data_dir, settings, progress),
    "project_status": lambda graph, ai, data_dir, settings, progress:
        project_status.run(graph, ai, data_dir, settings, progress),
    "meeting_prep": lambda graph, ai, data_dir, settings, progress:
        meeting_prep.run(graph, ai, data_dir, settings, progress),
}


# Friendly "nothing to report" messages, sent when a section runs but produces
# no items. Keeps the user informed that the section was checked.
_EMPTY_SECTION_MESSAGES: dict[str, str] = {
    "reply_needed":                "📧 **Reply Needed** — inbox is clear, no emails awaiting your reply.",
    "followup_needed":             "📤 **Sent — No Response** — every email you sent has a reply. No follow-ups pending.",
    "commitments_extract":         "📋 **Commitments** — no new commitments extracted from recent emails.",
    "upcoming_commitments":        "📅 **Upcoming Commitments** — no commitments due in the next 7 days.",
    "due_today":                   "✅ **Due Today** — nothing due today, you're free.",
    "yesterday_recap":             "📰 **Yesterday's Recap** — quiet day, no activity to report.",
    "meetings_today":              "📅 **Today's Meetings** — no meetings on your calendar today.",
    "recent_meetings":             "🎬 **Recent Meetings** — no new meeting recordings captured.",
    "meeting_action_items":        "📝 **Meeting Action Items** — no open action items.",
    "expenses":                    "💳 **Expenses** — no new receipts captured.",
    "market_intelligence":         "📡 **Market Intelligence** — no new market signals worth surfacing.",
    "company_intelligence":        "🏢 **Company Intelligence** — no new company signals worth surfacing.",
    "relationship_health":         "💚 **Relationship Health** — all key relationships are healthy. Nothing cooling.",
    "projects_needing_attention":  "✅ **Projects Needing Attention** — all projects on track.",
    "project_status":              "📁 **Project Status** — no active projects in your portfolio yet.",
    "meeting_prep":                "🕐 **Meeting Prep** — no upcoming meetings in the prep window.",
    "ai_summary":                  "☀️ **Morning Briefing** — no briefing produced (data may still be loading).",
}


def _format_section_for_teams(result: dict, tz: "ZoneInfo | None" = None) -> str:
    section_id = result.get("id", "")
    # Section formatters that need to compute "today"/"tomorrow" labels use
    # the user's tz when provided; UTC fallback keeps callers that haven't
    # been migrated working without crashing.
    from zoneinfo import ZoneInfo as _ZI
    tz = tz or _ZI("UTC")
    _now_local = datetime.now(tz)
    if section_id == "ai_summary":
        briefing = result.get("briefing", "")
        if briefing:
            return briefing
        return _EMPTY_SECTION_MESSAGES["ai_summary"]

    if section_id == "business_insights":
        narrative = result.get("narrative") or ""
        items = result.get("items", [])
        period_label = result.get("period_label") or result.get("week_of", "")
        if not narrative and not items:
            return "📊 **Weekly Brief** — quiet week, nothing notable to report." + (f" ({period_label})" if period_label else "")
        cat_tag = {
            "pipeline":   "🟢 Pipeline",
            "engagement": "🟡 Engagement",
            "execution":  "🔵 Execution",
            "intel":      "📡 Intel",
        }
        pri_tag = {"high": "🔴", "medium": "🟡", "low": "🟢"}
        header = f"📊 **Weekly Brief** · {period_label}" if period_label else "📊 **Weekly Brief**"
        lines = [header]
        if narrative:
            lines.append("")
            lines.append(narrative)

        # Group headlines by timeframe: last week (retrospective) / now (current
        # backlog) / coming up (next 7 days). Old results without a timeframe
        # field fall back to "now".
        def _fmt_headline(it):
            cat = cat_tag.get(it.get("category", ""), it.get("category", "").title())
            pri = pri_tag.get(it.get("priority", ""), "")
            out = [f"{pri} {cat} — **{it.get('title', '')}**"]
            detail = (it.get("detail") or "")[:240]
            if detail:
                out.append(detail)
            return out

        for tf, label in (("past", "📋 Last week"),
                          ("now", "📌 Now"),
                          ("ahead", "📅 Coming up (next 7 days)")):
            group = [it for it in items if (it.get("timeframe") or "now") == tf]
            if not group:
                continue
            lines.append("")
            lines.append(f"**{label}**")
            for it in group[:8]:
                lines.append("")
                lines.extend(_fmt_headline(it))
        # Deltas snippet
        deltas = (result.get("stats") or {}).get("deltas") or {}
        notable_deltas = [
            (k, v) for k, v in deltas.items()
            if v.get("pct") is not None and abs(v.get("pct", 0)) >= 25
        ]
        if notable_deltas:
            lines.append("")
            lines.append("**Notable week-over-week changes**")
            for k, v in notable_deltas[:6]:
                label = k.replace("_", " ")
                pct = v.get("pct", 0)
                arrow = "↑" if pct > 0 else "↓"
                lines.append(f"  {arrow} {label}: {v['previous']} → {v['current']} ({pct:+d}%)")
        return "\n".join(lines)

    items = result.get("items", [])
    if not items:
        return _EMPTY_SECTION_MESSAGES.get(
            section_id,
            f"**{section_id.replace('_', ' ').title()}** — nothing to report.",
        )

    if section_id == "reply_needed":
        lines = [f"**Reply Needed** — {len(items)} email(s)"]
        priority_tag = {"high": "🔴", "medium": "🟡", "low": "🟢"}
        for i, item in enumerate(items[:10], start=1):
            tag = priority_tag.get(item.get("priority", "medium"), "🟡")
            subject = (item.get("subject") or "(no subject)")[:80]
            from_name = (item.get("from_name") or "").strip()
            from_email = (item.get("from_email") or "").strip()
            sender = from_name or from_email or "Unknown sender"
            summary = (item.get("reason") or item.get("preview") or "").strip()
            summary = " ".join(summary.split())[:200]
            lines.append("")
            lines.append(f"{i}. {tag} **{subject}** — {sender}")
            if summary:
                lines.append(f"   {summary}")
        if len(items) > 10:
            lines.append("")
            lines.append(f"... and {len(items) - 10} more")
        return "\n".join(lines)

    if section_id == "followup_needed":
        lines = [f"**Follow-up Needed** — {len(items)} email(s)"]
        urgency_tag = {"high": "🔴", "medium": "🟡", "low": "🟢"}
        for i, item in enumerate(items[:10], start=1):
            tag = urgency_tag.get(item.get("urgency", "medium"), "🟡")
            subject = (item.get("subject") or "(no subject)")[:55]
            to_name = item.get("to_name") or item.get("to_email") or "—"
            days = item.get("days_waiting", 0)
            lines.append(f"{i}. {tag} \"{subject}\" → {to_name} ({days}d, no reply)")
        if len(items) > 10:
            lines.append(f"... and {len(items) - 10} more")
        return "\n".join(lines)

    if section_id == "commitments_extract":
        priority_tag = {"high": "🔴", "medium": "🟡", "low": "🟢"}
        today_str = _now_local.strftime("%Y-%m-%d")
        tomorrow_str = (_now_local + timedelta(days=1)).strftime("%Y-%m-%d")

        dated = [it for it in items if it.get("due_date")]
        undated = [it for it in items if not it.get("due_date")]
        dated.sort(key=lambda it: it.get("due_date") or "")
        undated.sort(key=lambda it: it.get("received") or "", reverse=True)

        def _fmt_item(item: dict, kind: str) -> str:
            tag = priority_tag.get(item.get("priority", "medium"), "🟡")
            them = " ↩" if item.get("type") == "their_commitment" else ""
            desc = (item.get("description") or "")[:120]
            contact = item.get("contact_name") or item.get("contact_email") or ""
            contact_str = f" → {contact}" if contact else ""
            if kind == "dated":
                due = item.get("due_date") or ""
                if due == today_str:
                    meta = "due today"
                elif due == tomorrow_str:
                    meta = "due tomorrow"
                else:
                    meta = f"due {due[5:]}" if due else ""
            else:
                received = (item.get("received") or "")[:16].replace("T", " ")
                meta = f"from email {received}" if received else "no deadline"
            return f"{tag}{them} {desc}{contact_str} ({meta})"

        lines = [f"**Commitments** — {len(items)} item(s)"]

        if dated:
            lines.append("")
            lines.append(f"**With deadline ({len(dated)})** — earliest first")
            for i, item in enumerate(dated[:10], start=1):
                lines.append(f"{i}. {_fmt_item(item, 'dated')}")
            if len(dated) > 10:
                lines.append(f"... and {len(dated) - 10} more dated item(s)")

        if undated:
            lines.append("")
            lines.append(f"**No deadline ({len(undated)})** — latest email first")
            for i, item in enumerate(undated[:5], start=1):
                lines.append(f"{i}. {_fmt_item(item, 'undated')}")
            if len(undated) > 5:
                lines.append(f"... and {len(undated) - 5} more undated item(s)")

        return "\n".join(lines)

    if section_id == "upcoming_commitments":
        window = result.get("window_days", 7)
        priority_tag = {"high": "🔴", "medium": "🟡", "low": "🟢"}
        today_d = _now_local.date()
        tomorrow_d = today_d + timedelta(days=1)

        # Bucket items
        overdue, today_b, tomorrow_b, later = [], [], [], []
        for it in items:
            due_raw = (it.get("due_date") or "")[:10]
            try:
                due_d = datetime.strptime(due_raw, "%Y-%m-%d").date() if due_raw else None
            except Exception:
                due_d = None
            if it.get("overdue") or (due_d and due_d < today_d):
                overdue.append((due_d, it))
            elif due_d == today_d:
                today_b.append((due_d, it))
            elif due_d == tomorrow_d:
                tomorrow_b.append((due_d, it))
            else:
                later.append((due_d or datetime.max.date(), it))

        # Sort each bucket by due_date ascending
        for bucket in (overdue, today_b, tomorrow_b, later):
            bucket.sort(key=lambda p: p[0] or datetime.max.date())

        def _fmt(idx: int, it: dict, with_due: bool = True, days_ago: int = 0) -> str:
            tag = priority_tag.get(it.get("priority", "medium"), "🟡")
            them = " ↩" if it.get("type") == "their_commitment" else ""
            desc = (it.get("description") or "")[:140]
            contact = it.get("contact_name") or it.get("contact_email") or ""
            contact_str = f" → {contact}" if contact else ""
            extra = ""
            if with_due:
                due = (it.get("due_date") or "")[:10]
                if days_ago > 0:
                    extra = f" · was due {due[5:]} ({days_ago}d ago)"
                elif due:
                    extra = f" · due {due[5:]}"
            return f"{idx}. {tag}{them} {desc}{contact_str}{extra}"

        lines = [f"**Upcoming Commitments** — {len(items)} item(s) · next {window} days"]
        shown = 0
        cap = 15  # max items in the Teams message

        def _emit_bucket(label: str, header: str, bucket: list, days_ago_for: bool = False):
            nonlocal shown
            if not bucket or shown >= cap:
                return
            lines.append("")
            lines.append(header.format(n=len(bucket)))
            for due_d, it in bucket:
                if shown >= cap:
                    break
                days_ago = (today_d - due_d).days if (days_ago_for and due_d) else 0
                lines.append("  " + _fmt(shown + 1, it, with_due=True, days_ago=days_ago))
                shown += 1

        _emit_bucket("overdue",  "⚠️ **Overdue ({n})**",      overdue,   days_ago_for=True)
        _emit_bucket("today",    "📌 **Due today ({n})**",    today_b)
        _emit_bucket("tomorrow", "📅 **Due tomorrow ({n})**", tomorrow_b)
        _emit_bucket("later",    "📆 **Later this week ({n})**", later)

        if len(items) > shown:
            lines.append(f"\n+{len(items) - shown} more not shown")
        return "\n".join(lines)

    if section_id == "meeting_action_items":
        # Group by meeting, sort meetings by date desc.
        meetings: dict[str, list[dict]] = {}
        meeting_meta: dict[str, dict] = {}
        for it in items:
            mid = it.get("meeting_id") or "_unknown"
            meetings.setdefault(mid, []).append(it)
            if mid not in meeting_meta:
                meeting_meta[mid] = {
                    "title": it.get("meeting_title") or "(unknown meeting)",
                    "date":  it.get("meeting_date") or "",
                }
        ordered_mids = sorted(meetings.keys(),
                              key=lambda m: meeting_meta[m]["date"], reverse=True)

        lines = [f"**Meeting Action Items** — {len(items)} item(s) across {len(meetings)} meeting(s)"]
        for mid in ordered_mids:
            meta = meeting_meta[mid]
            title = meta["title"]
            date = meta["date"]
            grp = meetings[mid]
            lines.append("")
            header = f"**{title}**"
            if date:
                header += f" — {date}"
            header += f" ({len(grp)} item(s))"
            lines.append(header)
            for it in grp[:8]:
                owner = (it.get("owner") or "—").strip()
                action = " ".join((it.get("action") or "").split())[:220]
                due = it.get("due_date")
                due_str = f" (due {due[5:] if len(due) >= 10 else due})" if due else ""
                lines.append(f"  • **{owner}**: {action}{due_str}")
            if len(grp) > 8:
                lines.append(f"  ... +{len(grp) - 8} more in this meeting")
        return "\n".join(lines)

    if section_id == "market_intelligence":
        lines = [f"**Market Intelligence** — {len(items)} signal(s)"]
        for i, item in enumerate(items[:12], start=1):
            headline   = (item.get("headline") or "")[:120]
            summary    = (item.get("summary") or "")[:250]
            source     = (item.get("source") or "").strip()
            source_url = (item.get("source_url") or "").strip()
            date_part  = (item.get("published_date") or "")[5:]
            signal     = item.get("signal_type", "other")
            priority   = item.get("priority", "medium")

            source_link = f"[{source}]({source_url})" if (source and source_url) else source
            meta_parts  = [p for p in [signal, priority if priority != "medium" else "", date_part, source_link] if p]
            meta        = " · ".join(meta_parts)

            lines.append("")
            lines.append(f"**{i}. {headline}**")
            if summary:
                lines.append(summary)
            if meta:
                lines.append(meta)
        if len(items) > 12:
            lines.append(f"\n+{len(items) - 12} more signals")
        return "\n".join(lines)

    if section_id == "company_intelligence":
        lines = [f"**Company Intelligence** — {len(items)} signal(s)"]
        # Group by company
        by_company: dict[str, list[dict]] = {}
        for item in items[:15]:
            co = item.get("company") or "Unknown"
            by_company.setdefault(co, []).append(item)
        for company, co_items in by_company.items():
            lines.append("")
            lines.append(f"**{company}**")
            for item in co_items:
                headline   = (item.get("headline") or "")[:120]
                summary    = (item.get("summary") or "")[:250]
                person     = item.get("person") or ""
                source     = (item.get("source") or "").strip()
                source_url = (item.get("source_url") or "").strip()
                date_part  = (item.get("published_date") or "")[5:]
                priority   = item.get("priority", "medium")

                source_link = f"[{source}]({source_url})" if (source and source_url) else source
                meta_parts  = [p for p in [person, priority if priority != "medium" else "", date_part, source_link] if p]
                meta        = " · ".join(meta_parts)

                lines.append(f"**{headline}**")
                if summary:
                    lines.append(summary)
                if meta:
                    lines.append(meta)
        remaining = len(items) - 15
        if remaining > 0:
            lines.append(f"\n+{remaining} more signals")
        return "\n".join(lines)

    if section_id == "relationship_health":
        health_tag = {
            "at_risk": "🔴 At Risk",
            "stalled": "⏸ Stalled",
            "cooling": "🟡 Cooling",
            "new":     "🆕 New",
        }
        lines = [f"**Relationship Health** — {len(items)} contact(s) need attention"]
        for item in items[:12]:
            name      = item.get("contact_name") or item.get("contact_email") or "—"
            company   = item.get("company") or ""
            health    = item.get("health", "")
            tag       = health_tag.get(health, health.upper())
            reminder  = (item.get("reminder") or "")[:300]
            action    = (item.get("suggested_action") or "")[:200]
            who       = f"**{name}**" + (f" · {company}" if company else "")
            lines.append("")
            lines.append(f"{tag} — {who}")
            if reminder:
                lines.append(reminder)
            if action:
                lines.append(f"→ {action}")
        if len(items) > 12:
            lines.append(f"\n+{len(items) - 12} more")
        return "\n".join(lines)

    if section_id == "expenses":
        lines = [f"💳 Expense Capture — {len(items)} new receipt(s) captured"]
        totals: dict[str, float] = {}
        for item in items:
            cat = item.get("category", "Other")
            try:
                totals[cat] = totals.get(cat, 0) + float(item.get("amount") or 0)
            except Exception:
                pass
        for i, item in enumerate(items[:8], start=1):
            vendor   = (item.get("vendor") or "?")[:30]
            amount   = item.get("amount", "")
            currency = item.get("currency", "")
            cat      = item.get("category", "")
            date     = (item.get("date") or "")
            lines.append(f"{i}. {vendor} — {amount} {currency} ({cat}) on {date}")
        if totals:
            lines.append("By category: " + " · ".join(
                f"{k} ${v:.0f}" for k, v in sorted(totals.items())
            ))
        return "\n".join(lines)

    if section_id == "meetings_today":
        the_date = result.get("date", "")
        lines = [f"**Today's Meetings** — {len(items)} meeting(s)" + (f" ({the_date})" if the_date else "")]
        for it in items[:12]:
            subject = (it.get("subject") or "(no subject)")[:100]
            start = it.get("start_time") or ""
            end = it.get("end_time") or ""
            time_str = (f"{start}–{end}" if start and end else start) if not it.get("is_all_day") else "all day"
            location = (it.get("location") or "").strip()
            attendees = it.get("attendees", [])
            who = f" · {len(attendees)} attendee(s)" if attendees else ""
            loc_str = f" · {location[:60]}" if location else ""
            lines.append("")
            lines.append(f"**{time_str}** — {subject}{who}{loc_str}")
            if attendees:
                lines.append(f"  with {', '.join(attendees[:5])}" + ("..." if len(attendees) > 5 else ""))
        if len(items) > 12:
            lines.append(f"\n+{len(items) - 12} more")
        return "\n".join(lines)

    if section_id == "due_today":
        the_date = result.get("date", "")
        lines = [f"**Due Today** — {len(items)} commitment(s)" + (f" ({the_date})" if the_date else "")]
        priority_tag = {"high": "🔴", "medium": "🟡", "low": "🟢"}
        for i, it in enumerate(items[:15], start=1):
            tag = priority_tag.get(it.get("priority", "medium"), "🟡")
            desc = (it.get("description") or "")[:160]
            contact = it.get("contact_name") or it.get("contact_email") or ""
            contact_str = f" → {contact}" if contact else ""
            lines.append(f"{i}. {tag} {desc}{contact_str}")
        if len(items) > 15:
            lines.append(f"... +{len(items) - 15} more")
        return "\n".join(lines)

    if section_id == "yesterday_recap":
        stats = result.get("stats", {})
        the_date = result.get("date", "")
        lines = [f"**Yesterday's Recap** — {the_date}"]
        lines.append(
            f"📥 {stats.get('inbound_count', 0)} in  ·  "
            f"📤 {stats.get('outbound_count', 0)} out  ·  "
            f"📅 {stats.get('meetings_count', 0)} meeting(s)  ·  "
            f"✅ {stats.get('commitments_count', 0)} commitment(s)"
        )

        meetings = [i for i in items if i.get("type") == "meeting"]
        inbound = [i for i in items if i.get("type") == "inbound_email"]
        outbound = [i for i in items if i.get("type") == "outbound_email"]
        commits = [i for i in items if i.get("type") == "commitment"]

        if meetings:
            lines.append("")
            lines.append(f"**Meetings ({len(meetings)})**")
            for m in meetings:
                rng = f"{m.get('start','')}–{m.get('end','')}" if m.get("start") else ""
                subj = (m.get("subject") or "")[:100]
                lines.append(f"  • {rng} {subj}")

        if inbound:
            lines.append("")
            lines.append(f"**Top inbound ({len(inbound)})**")
            for m in inbound:
                lines.append(f"  • {m.get('time','')} \"{(m.get('subject') or '')[:90]}\" — {m.get('from','')}")

        if outbound:
            lines.append("")
            lines.append(f"**Top outbound ({len(outbound)})**")
            for m in outbound:
                lines.append(f"  • {m.get('time','')} \"{(m.get('subject') or '')[:90]}\" → {m.get('to','')}")

        if commits:
            lines.append("")
            lines.append(f"**New commitments ({len(commits)})**")
            for c in commits:
                kind = " ↩" if c.get("commit_kind") == "their_commitment" else ""
                due = f" (due {c.get('due_date','')[5:]})" if c.get("due_date") else ""
                lines.append(f"  •{kind} {(c.get('description') or '')[:140]}{due}")
        return "\n".join(lines)

    if section_id == "projects_needing_attention":
        total = result.get("total_projects", 0)
        lines = [f"**Projects Needing Attention** — {len(items)} of {total} project(s)"]
        status_tag = {"needs_attention": "🔴 Needs Attention", "early_stage": "🆕 Early Stage"}
        for it in items[:10]:
            name = it.get("name", "")
            tag = status_tag.get(it.get("status", ""), it.get("status", "").upper())
            momentum = it.get("momentum", "")
            mom_str = f" · {momentum} momentum" if momentum else ""
            last = it.get("last_activity", "")
            last_str = f" · last activity {last}" if last else ""
            summary = (it.get("summary") or "")[:240]
            next_action = (it.get("next_action") or "")[:200]
            lines.append("")
            lines.append(f"{tag} — **{name}**{mom_str}{last_str}")
            if summary:
                lines.append(summary)
            if next_action:
                lines.append(f"→ {next_action}")
        if len(items) > 10:
            lines.append(f"\n+{len(items) - 10} more")
        return "\n".join(lines)

    if section_id == "project_status":
        counts = result.get("status_counts", {})
        lines = [f"**Project Status** — {len(items)} active project(s)"]
        if counts:
            parts = []
            for k in ("ongoing", "needs_attention", "paused", "early_stage"):
                if counts.get(k):
                    parts.append(f"{k.replace('_', ' ')} {counts[k]}")
            if parts:
                lines.append("  ·  ".join(parts))

        status_tag = {
            "ongoing":         "🟢",
            "needs_attention": "🔴",
            "paused":          "⏸",
            "early_stage":     "🆕",
        }
        # Group by status
        by_status: dict[str, list[dict]] = {}
        for it in items:
            by_status.setdefault(it.get("status", ""), []).append(it)
        for status in ("needs_attention", "ongoing", "paused", "early_stage"):
            grp = by_status.get(status, [])
            if not grp:
                continue
            lines.append("")
            lines.append(f"{status_tag.get(status, '')} **{status.upper()} ({len(grp)})**")
            for it in grp[:8]:
                name = it.get("name", "")
                momentum = it.get("momentum", "")
                last = it.get("last_activity", "")
                meta_parts = [p for p in [momentum, f"last {last}" if last else ""] if p]
                meta = " · ".join(meta_parts)
                lines.append(f"  • {name}" + (f" — {meta}" if meta else ""))
            if len(grp) > 8:
                lines.append(f"  ... +{len(grp) - 8} more")
        return "\n".join(lines)

    title = section_id.replace("_", " ").title()
    lines = [f"**{title}** — {len(items)} item(s)"]
    for i, item in enumerate(items[:10], start=1):
        label = item.get("subject") or item.get("vendor") or item.get("title") or str(item)
        lines.append(f"{i}. {str(label)[:80]}")
    if len(items) > 10:
        lines.append(f"... and {len(items) - 10} more")
    return "\n".join(lines)


def _log_outbound_to_history(uid: str, content: str) -> None:
    """Append an outbound system message (briefing push, prep, expense notify, …)
    to .data/{uid}/bot_history.db so the web chat panel can see it too."""
    if not content:
        return
    import time
    from src.modules.db_helpers import open_sqlite
    db_path = _udir(uid) / "bot_history.db"
    db_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        con = open_sqlite(db_path)
        con.execute(
            "CREATE TABLE IF NOT EXISTS history "
            "(id INTEGER PRIMARY KEY AUTOINCREMENT, role TEXT, content TEXT, ts REAL)"
        )
        con.execute("INSERT INTO history (role, content, ts) VALUES (?, ?, ?)",
                    ("model", content, time.time()))
        con.commit()
        con.close()
    except Exception as e:
        print(f"[Bot] history log failed for {uid}: {e}")


def _send_to_bot(uid: str, result: dict) -> None:
    bot_uid, chat_id = _find_bot_for_user(uid)
    if not bot_uid or not chat_id:
        return
    message = _format_section_for_teams(result, get_user_tz(_udir(uid)))
    if not message:
        return
    try:
        bot_token = auth.get_valid_access_token(bot_uid)
        bot_graph = GraphClient(bot_token)
        bot_graph.send_chat_message(chat_id, message)
        _log_outbound_to_history(uid, message)
    except Exception as e:
        print(f"[Bot] Failed to send {result.get('id')} to {uid}: {e}")


@app.post("/api/sections/{section_id}/run")
def run_section(section_id: str, background_tasks: BackgroundTasks,
                session: dict = Depends(require_session)):
    if section_id not in _SECTION_RUNNERS:
        raise HTTPException(404, detail=f"Section '{section_id}' not implemented yet")
    uid = session["user_id"]

    results_dir = _udir(uid) / "results"
    results_dir.mkdir(parents=True, exist_ok=True)
    _write_json(results_dir / f"{section_id}.json", {
        "id": section_id, "status": "running",
        "started_at": datetime.now(timezone.utc).isoformat(),
    })

    def _run():
        logs: list[str] = []

        def _progress(msg: str):
            logs.append(msg)
            print(f"[Section:{section_id}] {msg}")
            _write_json(results_dir / f"{section_id}.json", {
                "id": section_id, "status": "running",
                "started_at": datetime.now(timezone.utc).isoformat(),
                "logs": logs[-20:],
            })

        try:
            token    = auth.get_valid_access_token(uid)
            graph    = GraphClient(token)
            ai       = AIClient()
            settings = _read_json(_user_settings(uid))

            result = _SECTION_RUNNERS[section_id](
                graph, ai, _udir(uid), settings, _progress,
            )
            result["logs"] = logs
            _write_json(results_dir / f"{section_id}.json", result)
            _send_to_bot(uid, result)
            print(f"[Section:{section_id}] Done for {uid}")
        except Exception as e:
            _write_json(results_dir / f"{section_id}.json", {
                "id": section_id, "status": "error",
                "error": str(e),
                "logs": logs,
                "last_run": datetime.now(timezone.utc).isoformat(),
            })
            print(f"[Section:{section_id}] Failed for {uid}: {e}")

    background_tasks.add_task(_run)
    return {"ok": True}


@app.get("/api/sections/due_today")
def get_due_today(session: dict = Depends(require_session)):
    """Derived section — real-time filter of commitments_extract for due_date == today."""
    uid = session["user_id"]
    today_str = today_local_str(_udir(uid))
    path = _udir(uid) / "results" / "commitments_extract.json"
    if not path.exists():
        return {"id": "due_today", "status": "not_run", "items": [], "count": 0, "empty": True}
    data = _read_json(path)
    items = [
        item for item in data.get("items", [])
        if item.get("due_date") == today_str and item.get("type") == "my_commitment"
    ]
    return {
        "id": "due_today", "status": "fresh",
        "date": today_str, "items": items,
        "count": len(items), "empty": len(items) == 0,
    }


@app.get("/api/sections/{section_id}")
def get_section(section_id: str, session: dict = Depends(require_session)):
    uid  = session["user_id"]
    path = _udir(uid) / "results" / f"{section_id}.json"
    if not path.exists():
        return {"id": section_id, "status": "not_run", "items": [], "count": 0, "empty": True}
    return _read_json(path)


@app.get("/api/sections/{section_id}/instructions")
def get_section_instructions(section_id: str, session: dict = Depends(require_session)):
    uid  = session["user_id"]
    path = _udir(uid) / "instructions" / f"{section_id}.md"
    return {"content": path.read_text() if path.exists() else ""}


@app.put("/api/sections/{section_id}/instructions")
def update_section_instructions(section_id: str, body: dict, session: dict = Depends(require_session)):
    uid  = session["user_id"]
    path = _udir(uid) / "instructions" / f"{section_id}.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(body.get("content", ""))
    return {"ok": True}


def _run_section_for_user(uid: str, section_id: str) -> None:
    """Run a section directly (no HTTP). Used by bot.py to trigger sections from Teams chat."""
    if section_id not in _SECTION_RUNNERS:
        print(f"[Section:{section_id}] Not implemented — skipping")
        return
    results_dir = _udir(uid) / "results"
    results_dir.mkdir(parents=True, exist_ok=True)
    _write_json(results_dir / f"{section_id}.json", {
        "id": section_id, "status": "running",
        "started_at": datetime.now(timezone.utc).isoformat(),
    })
    try:
        token    = auth.get_valid_access_token(uid)
        graph    = GraphClient(token)
        ai       = AIClient()
        settings = _read_json(_user_settings(uid))
        result   = _SECTION_RUNNERS[section_id](graph, ai, _udir(uid), settings, None)
        _write_json(results_dir / f"{section_id}.json", result)
        _send_to_bot(uid, result)
        print(f"[Section:{section_id}] Done (bot-triggered) for {uid}")
    except Exception as e:
        _write_json(results_dir / f"{section_id}.json", {
            "id": section_id, "status": "error", "error": str(e),
            "last_run": datetime.now(timezone.utc).isoformat(),
        })
        print(f"[Section:{section_id}] Failed (bot-triggered) for {uid}: {e}")


# ── CRM ──────────────────────────────────────────────────

@app.get("/api/crm")
def get_crm(session: dict = Depends(require_session)):
    uid      = session["user_id"]
    crm_path = _udir(uid) / "crm.json"
    if not crm_path.exists():
        return {"last_scan": None, "months_scanned": 0, "total": 0, "contacts": []}
    try:
        data     = json.loads(crm_path.read_text())
        contacts = list(data.get("contacts", {}).values())
        contacts.sort(key=lambda x: x.get("last_contact", ""), reverse=True)
        return {
            "last_scan":      data.get("last_scan"),
            "months_scanned": data.get("months_scanned", 0),
            "total":          len(contacts),
            "contacts":       contacts,
        }
    except Exception as e:
        raise HTTPException(500, str(e))


@app.patch("/api/crm/{email}")
def patch_crm_contact(email: str, body: dict, session: dict = Depends(require_session)):
    uid  = session["user_id"]
    from src.modules.crm import load_crm, save_crm
    data = load_crm(_udir(uid))
    addr = email.lower()
    if addr not in data.get("contacts", {}):
        raise HTTPException(404, "Contact not found")
    data["contacts"][addr].update(body)
    save_crm(_udir(uid), data)
    return data["contacts"][addr]


@app.post("/api/crm")
def create_crm_contact(body: dict, session: dict = Depends(require_session)):
    """Manually add a new contact. Required: email. Optional: name, company,
    role, phone, linkedin, status, priority, summary, writing_style, notes."""
    uid  = session["user_id"]
    from src.modules.crm import load_crm, save_crm
    email = (body.get("email") or "").strip().lower()
    if not email or "@" not in email:
        raise HTTPException(400, "Valid email is required")
    data = load_crm(_udir(uid))
    contacts = data.setdefault("contacts", {})
    if email in contacts:
        raise HTTPException(409, "Contact already exists — use PATCH to update")
    today = today_local_str(_udir(uid))
    contact = {
        "email":         email,
        "name":          body.get("name", ""),
        "company":       body.get("company", ""),
        "role":          body.get("role", ""),
        "phone":         body.get("phone", ""),
        "linkedin":      body.get("linkedin", ""),
        "website":       body.get("website", ""),
        "status":        body.get("status", "other"),
        "priority":      body.get("priority", "medium"),
        "summary":       body.get("summary", ""),
        "writing_style": body.get("writing_style", ""),
        "notes":         body.get("notes", ""),
        "thread_count":  body.get("thread_count", 0),
        "last_contact":  body.get("last_contact", today),
        "ignore":        bool(body.get("ignore", False)),
        "updated_at":    today,
        "manual":        True,
    }
    contacts[email] = contact
    save_crm(_udir(uid), data)
    return contact


@app.post("/api/crm/scan")
def trigger_crm_scan(background_tasks: BackgroundTasks,
                     session: dict = Depends(require_session),
                     months: int = 6):
    uid = session["user_id"]

    def _run():
        try:
            from src.modules.crm import build_crm, save_crm
            from src.ai import AIClient
            token = auth.get_valid_access_token(uid)
            graph = GraphClient(token)
            ai    = AIClient()
            seg, biz = _get_crm_ai_context(uid)
            result = build_crm(graph, ai, _udir(uid),
                               months=months,
                               market_segments_content=seg,
                               business_context=biz)
            save_crm(_udir(uid), result)
            print(f"[CRM] Scan complete for {uid}: {len(result.get('contacts', {}))} contacts")
        except Exception as e:
            print(f"[CRM] Scan failed for {uid}: {e}")

    background_tasks.add_task(_run)
    return {"ok": True, "message": "CRM scan started"}


@app.post("/api/projects/scan")
def trigger_projects_scan(background_tasks: BackgroundTasks,
                          session: dict = Depends(require_session)):
    """Force a full Projects DB rebuild from inbox history."""
    uid = session["user_id"]

    def _run():
        try:
            from src.modules.projects import build_projects, save_projects
            from src.ai import AIClient
            token    = auth.get_valid_access_token(uid)
            graph    = GraphClient(token)
            ai       = AIClient()
            settings = _read_json(_user_settings(uid))
            result   = build_projects(graph, ai, _udir(uid), settings=settings)
            save_projects(_udir(uid), result)
            print(f"[Projects] Scan complete for {uid}: {len(result.get('projects', {}))} projects")
        except Exception as e:
            print(f"[Projects] Scan failed for {uid}: {e}")

    background_tasks.add_task(_run)
    return {"ok": True, "message": "Projects scan started"}


# ── Bulk upload (CRM + Projects) ──────────────────────────

_BULK_ALLOWED_EXTS = {".csv", ".xlsx", ".xls", ".docx", ".pdf", ".txt", ".md"}
_BULK_MAX_BYTES = 10 * 1024 * 1024  # 10 MB per file


async def _read_uploads(files: list[UploadFile]) -> list[tuple[bytes, str]]:
    out: list[tuple[bytes, str]] = []
    for f in files:
        ext = Path(f.filename or "").suffix.lower()
        if ext not in _BULK_ALLOWED_EXTS:
            raise HTTPException(400, f"Unsupported file type: {ext or f.filename}")
        data = await f.read()
        if len(data) > _BULK_MAX_BYTES:
            raise HTTPException(400, f"{f.filename} is larger than 10MB")
        if not data:
            continue
        out.append((data, f.filename or f"upload{ext}"))
    return out


@app.post("/api/crm/bulk-upload")
async def crm_bulk_upload(files: list[UploadFile] = File(...),
                          session: dict = Depends(require_session)):
    """Parse uploaded files into a preview list of CRM contacts (NOT yet saved)."""
    from src.modules.bulk_loader import parse_crm_file
    uploads = await _read_uploads(files)
    if not uploads:
        return {"contacts": [], "files": []}

    ai = AIClient()
    all_contacts: list[dict] = []
    file_summaries: list[dict] = []
    for data, name in uploads:
        try:
            parsed = parse_crm_file(data, name, ai)
        except Exception as e:
            file_summaries.append({"name": name, "count": 0, "error": str(e)[:200]})
            continue
        file_summaries.append({"name": name, "count": len(parsed), "error": None})
        all_contacts.extend(parsed)

    # Deduplicate by email across files
    seen: set = set()
    dedup: list[dict] = []
    for c in all_contacts:
        e = c.get("email", "").lower()
        if not e or e in seen:
            continue
        seen.add(e)
        dedup.append(c)

    # Mark which ones already exist in CRM
    uid = session["user_id"]
    existing: set = set()
    crm_path = _udir(uid) / "crm.json"
    if crm_path.exists():
        try:
            existing = {e.lower() for e in json.loads(crm_path.read_text()).get("contacts", {}).keys()}
        except Exception:
            pass
    for c in dedup:
        c["_already_exists"] = c["email"] in existing

    return {"contacts": dedup, "files": file_summaries}


@app.post("/api/crm/bulk-commit")
async def crm_bulk_commit(request: Request, session: dict = Depends(require_session)):
    """Persist the user-approved contacts. Exact email matches are skipped."""
    from src.modules.crm import load_crm, save_crm
    uid = session["user_id"]
    body = await request.json()
    contacts = body.get("contacts") or []
    if not isinstance(contacts, list):
        raise HTTPException(400, "contacts must be a list")

    db = load_crm(_udir(uid))
    existing = db.get("contacts") or {}
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    added = 0
    skipped_existing = 0
    skipped_invalid = 0
    for c in contacts:
        email = (c.get("email") or "").strip().lower()
        if not email or "@" not in email:
            skipped_invalid += 1
            continue
        if email in existing:
            skipped_existing += 1
            continue
        existing[email] = {
            "email":         email,
            "name":          c.get("name") or "",
            "company":       c.get("company") or "",
            "role":          c.get("role") or "",
            "phone":         c.get("phone") or "",
            "linkedin":      c.get("linkedin") or "",
            "status":        c.get("status") or "other",
            "priority":      c.get("priority") or "medium",
            "summary":       c.get("summary") or "",
            "notes":         c.get("notes") or "",
            "thread_count":  0,
            "last_contact":  "",
            "updated_at":    today,
            "manual":        True,
            "source":        "bulk_upload",
            "ignore":        False,
        }
        added += 1

    db["contacts"] = existing
    save_crm(_udir(uid), db)
    return {"added": added, "skipped_existing": skipped_existing, "skipped_invalid": skipped_invalid}


@app.post("/api/projects/bulk-upload")
async def projects_bulk_upload(files: list[UploadFile] = File(...),
                               session: dict = Depends(require_session)):
    """Parse uploaded files into a preview list of Projects (NOT yet saved)."""
    from src.modules.bulk_loader import parse_project_file
    uploads = await _read_uploads(files)
    if not uploads:
        return {"projects": [], "files": []}

    ai = AIClient()
    all_projects: list[dict] = []
    file_summaries: list[dict] = []
    for data, name in uploads:
        try:
            parsed = parse_project_file(data, name, ai)
        except Exception as e:
            file_summaries.append({"name": name, "count": 0, "error": str(e)[:200]})
            continue
        file_summaries.append({"name": name, "count": len(parsed), "error": None})
        all_projects.extend(parsed)

    # Dedup by lowercase name across files
    seen: set = set()
    dedup: list[dict] = []
    for p in all_projects:
        n = (p.get("name") or "").lower().strip()
        if not n or n in seen:
            continue
        seen.add(n)
        dedup.append(p)

    # Mark which ones already exist (by name match against existing projects)
    uid = session["user_id"]
    existing_names: set = set()
    proj_path = _udir(uid) / "projects.json"
    if proj_path.exists():
        try:
            existing_names = {
                (p.get("name") or "").lower().strip()
                for p in json.loads(proj_path.read_text()).get("projects", {}).values()
            }
        except Exception:
            pass
    for p in dedup:
        p["_already_exists"] = (p["name"] or "").lower().strip() in existing_names

    return {"projects": dedup, "files": file_summaries}


@app.post("/api/projects/bulk-commit")
async def projects_bulk_commit(request: Request, session: dict = Depends(require_session)):
    """Persist user-approved projects. Existing name matches are skipped."""
    from src.modules.projects import load_projects, save_projects
    uid = session["user_id"]
    body = await request.json()
    projects = body.get("projects") or []
    if not isinstance(projects, list):
        raise HTTPException(400, "projects must be a list")

    db = load_projects(_udir(uid))
    existing = db.get("projects") or {}
    existing_names_lower = {(p.get("name") or "").lower() for p in existing.values()}
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    def _make_id(name: str) -> str:
        slug = re.sub(r"[^\w\s-]", "", (name or "")).strip().lower()
        slug = re.sub(r"[\s-]+", "-", slug)[:60]
        base = slug or "project"
        candidate = base
        i = 1
        while candidate in existing:
            i += 1
            candidate = f"{base}-{i}"
        return candidate

    added = 0
    skipped_existing = 0
    skipped_invalid = 0
    for p in projects:
        name = (p.get("name") or "").strip()
        if not name:
            skipped_invalid += 1
            continue
        if name.lower() in existing_names_lower:
            skipped_existing += 1
            continue
        pid = _make_id(name)
        existing[pid] = {
            "id":               pid,
            "name":             name,
            "summary":          p.get("summary") or "",
            "status":           p.get("status") or "ongoing",
            "category":         p.get("category") or "other",
            "momentum":         "steady",
            "participants":     p.get("participants") or [],
            "key_topics":       p.get("key_topics") or [],
            "next_action":      p.get("next_action") or "",
            "deadline":         p.get("deadline") or "",
            "thread_count":     0,
            "conversation_ids": [],
            "last_activity":    today,
            "ignore":           False,
            "updated_at":       today,
            "source":           "bulk_upload",
        }
        existing_names_lower.add(name.lower())
        added += 1

    db["projects"] = existing
    save_projects(_udir(uid), db)
    return {"added": added, "skipped_existing": skipped_existing, "skipped_invalid": skipped_invalid}


# ── Excel exports (CRM / Projects / Companies) ────────────
# Full-dataset xlsx downloads. Filename includes today's date so multiple
# downloads in a row stack cleanly in the user's Downloads folder.

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _xlsx_response(
    sheet_title: str,
    headers: list[str],
    rows: list[list],
    filename_stem: str,
) -> Response:
    """Build a single-sheet xlsx in memory and return it as a FastAPI
    attachment Response. Shared by all three export endpoints below so
    column behaviour stays consistent."""
    import io
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(headers)
    for r in rows:
        # openpyxl rejects bare lists/dicts in cells — pre-stringify any
        # collection-shaped value the caller didn't flatten itself.
        ws.append([
            (", ".join(str(x) for x in v) if isinstance(v, (list, tuple, set))
             else json.dumps(v, ensure_ascii=False) if isinstance(v, dict)
             else v)
            for v in r
        ])
    buf = io.BytesIO()
    wb.save(buf)
    today = datetime.now().strftime("%Y-%m-%d")
    fname = f"{filename_stem}-{today}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type=_XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.get("/api/crm/export.xlsx")
def export_crm_xlsx(session: dict = Depends(require_session)):
    uid = session["user_id"]
    from src.modules.crm import load_crm
    contacts = (load_crm(_udir(uid)).get("contacts") or {}).values()
    headers = [
        "Email", "Name", "Company", "Role", "Phone", "LinkedIn", "Website",
        "Status", "Priority", "Last Contact", "Thread Count",
        "Summary", "Writing Style", "Notes", "Tags", "Source",
        "Ignore", "Archived", "Updated At", "Added At",
    ]
    rows = []
    for c in contacts:
        rows.append([
            c.get("email", ""), c.get("name", ""), c.get("company", ""),
            c.get("role", ""), c.get("phone", ""), c.get("linkedin", ""),
            c.get("website", ""), c.get("status", ""), c.get("priority", ""),
            c.get("last_contact", ""), c.get("thread_count", 0),
            c.get("summary", ""), c.get("writing_style", ""), c.get("notes", ""),
            c.get("tags") or [], c.get("source", ""),
            bool(c.get("ignore", False)), bool(c.get("archived", False)),
            c.get("updated_at", ""), c.get("added_at", ""),
        ])
    rows.sort(key=lambda r: (r[8] or "", r[1] or ""))  # status, name
    return _xlsx_response("Contacts", headers, rows, "crm-contacts")


@app.get("/api/projects/export.xlsx")
def export_projects_xlsx(session: dict = Depends(require_session)):
    uid = session["user_id"]
    from src.modules.projects import load_projects
    projects = (load_projects(_udir(uid)).get("projects") or {}).values()
    headers = [
        "ID", "Name", "Category", "Status", "Momentum",
        "Summary", "Next Action", "Deadline", "Last Activity", "Thread Count",
        "Participants", "Key Topics", "Ignore", "Archived",
        "Updated At",
    ]
    rows = []
    for p in projects:
        rows.append([
            p.get("id", ""), p.get("name", ""), p.get("category", ""),
            p.get("status", ""), p.get("momentum", ""),
            p.get("summary", ""), p.get("next_action", ""),
            p.get("deadline", ""), p.get("last_activity", ""),
            p.get("thread_count", 0),
            p.get("participants") or [], p.get("key_topics") or [],
            bool(p.get("ignore", False)), bool(p.get("archived", False)),
            p.get("updated_at", ""),
        ])
    rows.sort(key=lambda r: (r[8] or ""), reverse=True)  # last_activity desc
    return _xlsx_response("Projects", headers, rows, "projects")


@app.get("/api/expenses/export.xlsx")
def export_expenses_xlsx(session: dict = Depends(require_session)):
    """Stream the user's expenses_master.xlsx directly. Unlike the other
    export endpoints (which build a fresh sheet from JSON), the expenses
    ledger IS already an .xlsx file written incrementally by m05_expense
    and the edit-row endpoints — so this just hands the live file back."""
    uid  = session["user_id"]
    path = _udir(uid) / "expenses" / "expenses_master.xlsx"
    if not path.exists():
        raise HTTPException(404, "No expense ledger yet")
    from fastapi.responses import FileResponse
    return FileResponse(
        path=path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="expenses.xlsx",
    )


@app.get("/api/companies/export.xlsx")
def export_companies_xlsx(session: dict = Depends(require_session)):
    uid = session["user_id"]
    from src.modules.companies import load_companies
    companies = (load_companies(_udir(uid)).get("companies") or {}).values()
    headers = [
        "Name", "Key", "Aliases", "Derived Status", "Priority",
        "Monitor Intelligence", "Ignore", "Manual",
        "Contact Count", "Project Count", "Total Threads", "Last Activity",
        "Contact Emails", "Project Names", "Notes",
        "Added At", "Updated At",
    ]
    rows = []
    for c in companies:
        contact_emails = [ct.get("email", "") for ct in (c.get("contacts") or [])]
        project_names  = [p.get("name", "")  for p in (c.get("projects") or [])]
        rows.append([
            c.get("name", ""), c.get("key", ""),
            c.get("aliases") or [], c.get("derived_status", ""),
            c.get("priority", ""),
            bool(c.get("monitor_intelligence", True)),
            bool(c.get("ignore", False)), bool(c.get("manual", False)),
            c.get("contact_count", 0), c.get("project_count", 0),
            c.get("thread_count_total", 0), c.get("last_activity", ""),
            contact_emails, project_names, c.get("notes", ""),
            c.get("added_at", ""), c.get("updated_at", ""),
        ])
    rows.sort(key=lambda r: (r[11] or ""), reverse=True)  # last_activity desc
    return _xlsx_response("Companies", headers, rows, "companies")


# ── M03 Meeting scan ──────────────────────────────────────

@app.post("/api/m03/scan")
def trigger_m03_scan(background_tasks: BackgroundTasks,
                     session: dict = Depends(require_session),
                     force: bool = False,
                     use_mock: bool = False):
    uid = session["user_id"]

    def _run():
        try:
            from src.modules.m03_meeting import run as m03_run
            from src.ai import AIClient
            token    = auth.get_valid_access_token(uid)
            graph    = GraphClient(token)
            ai       = AIClient()
            settings = _read_json(_user_settings(uid))
            result   = m03_run(graph, ai, _udir(uid), settings=settings, force=force, use_mock=use_mock)
            # Refresh read sections from the newly populated Meeting DB
            recent_meetings.run(_udir(uid))
            meeting_action_items.run(_udir(uid))
            print(f"[M03] Scan complete for {uid}: {result.get('processed', 0)} processed")

            # Send Teams notification for each processed meeting
            processed_results = [r for r in result.get("results", []) if r.get("status") == "processed"]
            if processed_results:
                lines = [f"✅ Meeting scan complete — {len(processed_results)} meeting(s) processed"]
                for r in processed_results[:5]:
                    n_actions = len(r.get("action_items", []))
                    draft_saved = r.get("followup_draft_saved", False)
                    title = r.get("title", "(untitled)")
                    date  = (r.get("date") or "")[:10]
                    lines.append(
                        f"• **{title}**" + (f" ({date})" if date else "")
                        + f" — {n_actions} action item(s)"
                        + (" · follow-up draft saved to Drafts" if draft_saved else "")
                    )
                try:
                    bot_uid, chat_id = _find_bot_for_user(uid)
                    if bot_uid and chat_id:
                        bot_token  = auth.get_valid_access_token(bot_uid)
                        bot_graph  = GraphClient(bot_token)
                        bot_graph.send_chat_message(chat_id, "\n".join(lines))
                except Exception as _ne:
                    print(f"[M03] Teams notification failed: {_ne}")

        except Exception as e:
            print(f"[M03] Scan failed for {uid}: {e}")

    background_tasks.add_task(_run)
    return {"ok": True, "message": "Meeting scan started"}


# ── Projects: read-only list + manual edit ───────────────

_PROJECT_EDITABLE_FIELDS = {
    "status", "momentum", "category", "summary", "next_action",
    "name", "deadline", "ignore",
}
_PROJECT_VALID_STATUSES  = {"ongoing", "needs_attention", "paused", "early_stage", "completed"}
_PROJECT_VALID_MOMENTUMS = {"accelerating", "steady", "slowing", "stalled"}


@app.get("/api/projects")
def get_projects(session: dict = Depends(require_session)):
    """List all projects from projects.json."""
    uid       = session["user_id"]
    proj_path = _udir(uid) / "projects.json"
    if not proj_path.exists():
        return {"last_scan": None, "total": 0, "projects": []}
    try:
        data     = json.loads(proj_path.read_text())
        projects = list(data.get("projects", {}).values())
        projects.sort(key=lambda p: p.get("last_activity", ""), reverse=True)
        return {
            "last_scan": data.get("last_scan"),
            "total":     len(projects),
            "projects":  projects,
        }
    except Exception as e:
        raise HTTPException(500, str(e))


# ── Companies ────────────────────────────────────────────
# Derived view over CRM + Projects, with per-company settings (most
# importantly monitor_intelligence, consumed by the company_intelligence
# section). companies.json is the single source of truth for "which
# companies to monitor" — CRM contact priority/ignore no longer participates
# in that decision.

@app.get("/api/companies")
def get_companies(session: dict = Depends(require_session)):
    uid = session["user_id"]
    from src.modules.companies import load_companies
    db = load_companies(_udir(uid))
    companies = list((db.get("companies") or {}).values())
    companies.sort(
        key=lambda c: (c.get("last_activity") or "", (c.get("name") or "").lower()),
        reverse=True,
    )
    return {
        "last_scan": db.get("last_scan"),
        "total":     len(companies),
        "companies": companies,
    }


@app.patch("/api/companies/{key}")
def patch_company(key: str, body: dict, session: dict = Depends(require_session)):
    uid = session["user_id"]
    from src.modules.companies import update_company
    try:
        return update_company(_udir(uid), key, body or {})
    except KeyError:
        raise HTTPException(404, f"Company not found: {key}")


@app.post("/api/companies")
def create_company(body: dict, session: dict = Depends(require_session)):
    """Manually add a company that doesn't have to appear in CRM/Projects."""
    uid = session["user_id"]
    from src.modules.companies import add_manual_company
    try:
        return add_manual_company(
            _udir(uid),
            name=body.get("name") or "",
            notes=body.get("notes") or "",
            priority=body.get("priority") or "medium",
            monitor_intelligence=bool(body.get("monitor_intelligence", True)),
        )
    except ValueError as e:
        raise HTTPException(409 if "already exists" in str(e) else 400, str(e))


@app.delete("/api/companies/{key}")
def remove_company(key: str, session: dict = Depends(require_session)):
    uid = session["user_id"]
    from src.modules.companies import delete_company
    if not delete_company(_udir(uid), key):
        raise HTTPException(404, f"Company not found: {key}")
    return {"ok": True}


@app.post("/api/companies/scan")
def trigger_companies_scan(background_tasks: BackgroundTasks,
                           session: dict = Depends(require_session)):
    uid = session["user_id"]
    background_tasks.add_task(_build_companies_for_user, uid)
    return {"ok": True, "message": "Company rebuild started"}


@app.post("/api/commitments/{commitment_id}/done")
def commitment_mark_done(commitment_id: str, session: dict = Depends(require_session)):
    from src.modules.commitments_state import mark_done
    mark_done(_udir(session["user_id"]), commitment_id, method="user")
    return {"ok": True, "id": commitment_id, "state": "done"}


@app.post("/api/commitments/{commitment_id}/snooze")
def commitment_snooze(commitment_id: str, body: dict | None = None,
                      session: dict = Depends(require_session)):
    from src.modules.commitments_state import mark_snoozed
    days = int((body or {}).get("days") or 3)
    mark_snoozed(_udir(session["user_id"]), commitment_id, days=days)
    return {"ok": True, "id": commitment_id, "state": "snoozed", "days": days}


# ── Email draft composer ──────────────────────────────────

def _html_to_plain(html: str) -> str:
    """Cheap HTML→text used when AI gets the original email body. Same shape
    as src.modules.crm._strip_html; inlined here to avoid importing the
    whole crm module."""
    import re as _re
    import html as _html
    text = _re.sub(r"<[^>]+>", " ", html or "")
    text = _html.unescape(text)
    text = _re.sub(r"[ \t]+", " ", text)
    text = _re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _plain_to_html(text: str) -> str:
    """Convert plain-text draft body to the HTML payload Outlook expects.
    Preserves blank-line paragraph breaks and single-line line breaks."""
    import html as _html
    escaped = _html.escape(text or "")
    paras = escaped.split("\n\n")
    return "".join(f"<p>{p.replace(chr(10), '<br>')}</p>" for p in paras if p.strip() or True)


def _fetch_original_email(uid: str, email_id: str) -> dict:
    """Return {subject, from_email, from_name, to_email, to_name, body_text,
    sent_date} for any mailbox message (inbox or sent items).
    Raises HTTPException(404) on failure."""
    token = auth.get_valid_access_token(uid)
    graph = GraphClient(token)
    try:
        msg = graph.get_message(email_id)
    except Exception as e:
        raise HTTPException(404, f"Could not fetch email: {e}")
    from_obj = (msg.get("from") or {}).get("emailAddress") or {}
    to_list  = msg.get("toRecipients") or []
    to_obj   = (to_list[0].get("emailAddress") if to_list else {}) or {}
    body_obj = msg.get("body") or {}
    body_content = body_obj.get("content") or msg.get("bodyPreview") or ""
    if (body_obj.get("contentType") or "").lower() == "html":
        body_content = _html_to_plain(body_content)
    return {
        "subject":    msg.get("subject") or "(no subject)",
        "from_email": (from_obj.get("address") or "").strip(),
        "from_name":  (from_obj.get("name") or "").strip(),
        "to_email":   (to_obj.get("address") or "").strip(),
        "to_name":    (to_obj.get("name") or "").strip(),
        "sent_date":  (msg.get("sentDateTime") or msg.get("receivedDateTime") or "")[:10],
        "body_text":  body_content[:4000],
    }


@app.post("/api/drafts/generate")
def draft_generate(body: dict, session: dict = Depends(require_session)):
    """Generate a draft for one of two modes:
      - mode='reply' (default): inbound email I need to respond to.
        To = original sender.
      - mode='followup': outbound email I sent that hasn't been replied to.
        To = original recipient. Tone: polite nudge.
    Both pull the original message via Graph and fuse with profile + writing-
    style + per-item AI hints. Returns plain-text body."""
    from src.modules.profile import load_profile_context, get_user_signature, append_signature_to_body
    uid     = session["user_id"]
    payload = body or {}
    email_id = (payload.get("email_id") or "").strip()
    mode     = (payload.get("mode") or "reply").strip().lower()
    if mode not in ("reply", "followup"):
        raise HTTPException(400, f"unknown mode: {mode}")
    if not email_id:
        raise HTTPException(400, "email_id required")

    orig = _fetch_original_email(uid, email_id)

    profile_ctx   = load_profile_context(_udir(uid))
    settings      = _read_json(_user_settings(uid))
    writing_style = settings.get("writing_style_note") or "Professional, concise, friendly."
    display_name  = settings.get("display_name") or "the executive"

    reason       = (payload.get("reason") or "").strip()
    opening      = (payload.get("suggested_opening") or "").strip()
    tone         = (payload.get("reply_tone") or "professional").strip()
    urgency      = (payload.get("urgency") or "medium").strip()
    days_waiting = payload.get("days_waiting")

    if mode == "reply":
        instr_path = _udir(uid) / "instructions" / "reply_needed.md"
        user_instruction = instr_path.read_text() if instr_path.exists() else ""
        subj = orig["subject"]
        reply_subject = subj if subj.lower().startswith("re:") else f"Re: {subj}"
        to_addr = orig["from_email"]
        prompt = (
            f"You are drafting an email reply on behalf of {display_name}.\n\n"
            f"BUSINESS CONTEXT:\n{profile_ctx}\n\n"
            f"WRITING STYLE (match this voice exactly):\n{writing_style}\n\n"
            f"USER REPLY PREFERENCES:\n{user_instruction or '(none)'}\n\n"
            f"ORIGINAL EMAIL from {orig['from_name'] or orig['from_email']} <{orig['from_email']}>:\n"
            f"Subject: {orig['subject']}\n\n{orig['body_text']}\n\n"
            f"WHY THIS NEEDS A REPLY:\n{reason or '(unspecified)'}\n\n"
            f"Suggested tone: {tone}\n"
            f"Suggested opening: {opening or '(none)'}\n\n"
            "Write the COMPLETE reply body. Plain text only — no markdown, no "
            "subject line, no extra commentary. Address the specific question(s) "
            "in the original email. Do not invent facts not present above. Do "
            "NOT write a sign-off — the user's real signature is appended "
            "automatically after this body."
        )
    else:  # followup
        instr_path = _udir(uid) / "instructions" / "followup_needed.md"
        user_instruction = instr_path.read_text() if instr_path.exists() else ""
        subj = orig["subject"]
        # Distinct subject so recipient sees it's a nudge, not a re-send.
        reply_subject = subj if subj.lower().startswith(("re:", "following up")) else f"Following up: {subj}"
        to_addr = orig["to_email"]
        recipient_label = orig["to_name"] or orig["to_email"] or "the recipient"
        days_str = f"{days_waiting} days" if isinstance(days_waiting, int) and days_waiting > 0 else "a while"
        prompt = (
            f"You are drafting a polite follow-up nudge on behalf of {display_name}.\n\n"
            f"BUSINESS CONTEXT:\n{profile_ctx}\n\n"
            f"WRITING STYLE (match this voice exactly):\n{writing_style}\n\n"
            f"USER FOLLOWUP PREFERENCES:\n{user_instruction or '(none)'}\n\n"
            f"ORIGINAL EMAIL I sent on {orig['sent_date']} to {recipient_label} <{orig['to_email']}>:\n"
            f"Subject: {orig['subject']}\n\n{orig['body_text']}\n\n"
            f"They have not replied in {days_str}. Urgency: {urgency}.\n"
            f"Context for why a follow-up matters now: {reason or '(general nudge)'}\n\n"
            "Write a SHORT, polite follow-up message body. Rules:\n"
            "- Do NOT quote the full original email back to them.\n"
            "- Acknowledge the prior thread in one sentence ('Just circling back on...').\n"
            "- Make it easy for them to respond — ask a specific question or "
            "offer to help unblock them.\n"
            "- Tone: friendly but professional. Never aggressive or guilt-tripping.\n"
            "- Plain text only, no markdown, no subject line. Do NOT write a sign-off — "
            "the user's real signature is appended automatically."
        )

    if not to_addr or "@" not in to_addr:
        raise HTTPException(
            500,
            f"Could not determine recipient ({'sender' if mode=='reply' else 'recipient'} "
            f"missing on original email)",
        )

    try:
        body_text = AIClient().generate(prompt).strip()
    except Exception as e:
        raise HTTPException(500, f"AI generation failed: {e}")
    body_text = body_text.strip('"').strip("'").strip()

    # Append the user's real signature (extracted from sent emails during
    # onboarding). Empty string for legacy users → body unchanged. Helper
    # handles HTML-vs-plain detection so an HTML signature with links and
    # external logos doesn't break the plain-text body format.
    body_text = append_signature_to_body(body_text, get_user_signature(settings))

    return {
        "subject": reply_subject,
        "to":      to_addr,
        "body":    body_text,
    }


@app.post("/api/drafts/refine")
def draft_refine(body: dict, session: dict = Depends(require_session)):
    """Take the current draft body + a free-form user request ("make it
    shorter", "add thanks") and return a revised draft body."""
    uid     = session["user_id"]
    payload = body or {}
    email_id     = (payload.get("email_id") or "").strip()
    current_body = (payload.get("current_body") or "").strip()
    user_request = (payload.get("user_request") or "").strip()
    if not current_body or not user_request:
        raise HTTPException(400, "current_body and user_request required")

    settings      = _read_json(_user_settings(uid))
    writing_style = settings.get("writing_style_note") or "Professional, concise, friendly."
    display_name  = settings.get("display_name") or "the executive"

    orig_snippet = ""
    if email_id:
        try:
            orig = _fetch_original_email(uid, email_id)
            orig_snippet = orig["body_text"][:1500]
        except Exception:
            pass

    prompt = (
        f"You are refining {display_name}'s draft email reply.\n\n"
        f"WRITING STYLE:\n{writing_style}\n\n"
        f"ORIGINAL EMAIL (context only — do not quote):\n{orig_snippet or '(not available)'}\n\n"
        f"CURRENT DRAFT:\n{current_body}\n\n"
        f"USER WANTS YOU TO:\n{user_request}\n\n"
        "Return ONLY the revised draft body — no preamble, no quotes around "
        "it, no explanation. Preserve everything the user did not ask to "
        "change. Plain text only."
    )
    try:
        revised = AIClient().generate(prompt).strip()
    except Exception as e:
        raise HTTPException(500, f"AI refine failed: {e}")
    return {"body": revised.strip('"').strip("'").strip()}


@app.post("/api/drafts/save")
def draft_save(body: dict, session: dict = Depends(require_session)):
    """Create a real Outlook draft via Graph. Plain-text body in → HTML out."""
    uid     = session["user_id"]
    payload = body or {}
    to       = (payload.get("to") or "").strip()
    subject  = (payload.get("subject") or "").strip()
    body_txt = (payload.get("body") or "").strip()

    missing = [n for n, v in (("to", to), ("subject", subject), ("body", body_txt)) if not v]
    if missing:
        raise HTTPException(400, f"Missing required field(s): {', '.join(missing)}")
    if "@" not in to:
        raise HTTPException(400, f"'to' is not a valid email address: {to!r}")

    token = auth.get_valid_access_token(uid)
    graph = GraphClient(token)
    try:
        result = graph.create_draft(subject=subject, body=_plain_to_html(body_txt), to=to)
    except Exception as e:
        raise HTTPException(500, f"Draft save failed: {e}")
    return {
        "ok":       True,
        "draft_id": result.get("id"),
        "web_link": result.get("webLink") or "",
    }


@app.patch("/api/projects/{project_id}")
def patch_project(project_id: str, body: dict, session: dict = Depends(require_session)):
    """Manually update a project. Allowed fields: status, momentum, category,
    summary, next_action, name, deadline, ignore."""
    uid       = session["user_id"]
    proj_path = _udir(uid) / "projects.json"
    if not proj_path.exists():
        raise HTTPException(404, "Projects DB not built yet")
    data = json.loads(proj_path.read_text())
    projects = data.get("projects", {})
    if project_id not in projects:
        raise HTTPException(404, f"Project '{project_id}' not found")
    updates = {k: v for k, v in body.items() if k in _PROJECT_EDITABLE_FIELDS}
    if "status" in updates and updates["status"] not in _PROJECT_VALID_STATUSES:
        raise HTTPException(400, f"Invalid status. Must be one of {sorted(_PROJECT_VALID_STATUSES)}")
    if "momentum" in updates and updates["momentum"] not in _PROJECT_VALID_MOMENTUMS:
        raise HTTPException(400, f"Invalid momentum. Must be one of {sorted(_PROJECT_VALID_MOMENTUMS)}")
    projects[project_id].update(updates)
    projects[project_id]["updated_at"] = today_local_str(_udir(uid))
    data["projects"] = projects
    # Atomic write
    tmp = proj_path.with_suffix(".tmp")
    tmp.write_text(json.dumps(data, indent=2, ensure_ascii=False))
    tmp.replace(proj_path)
    return projects[project_id]


# ── Expenses: full history from Excel + manual edit ──────

_EXPENSE_EDITABLE_FIELDS = {"Vendor", "Date", "Amount", "Currency", "Category", "GST_HST", "Net_Amount"}


def _expense_row_id(row: dict) -> str:
    """Stable id computed from Msg_ID + Att_ID (or fallback to row content)."""
    import hashlib
    key = (str(row.get("Msg_ID") or "") + "::" + str(row.get("Att_ID") or ""))
    if not key.strip("::"):
        # Fallback: hash the whole row
        key = json.dumps({k: str(v) for k, v in row.items()}, sort_keys=True)
    return hashlib.sha1(key.encode()).hexdigest()[:16]


@app.get("/api/expenses/all")
def get_all_expenses(session: dict = Depends(require_session)):
    """Return all rows from expenses_master.xlsx as a JSON list with stable ids."""
    uid  = session["user_id"]
    path = _udir(uid) / "expenses" / "expenses_master.xlsx"
    if not path.exists():
        return []
    try:
        from openpyxl import load_workbook
        ws      = load_workbook(path).active
        headers = [c.value for c in ws[1]]
        rows = [
            dict(zip(headers, row))
            for row in ws.iter_rows(min_row=2, values_only=True)
        ]
        for r in rows:
            r["id"] = _expense_row_id(r)
        return rows
    except Exception:
        return []


@app.patch("/api/expenses/{row_id}")
def patch_expense(row_id: str, body: dict, session: dict = Depends(require_session)):
    """Manually update a row in expenses_master.xlsx. Allowed fields: Vendor,
    Date, Amount, Currency, Category, GST_HST, Net_Amount."""
    uid  = session["user_id"]
    path = _udir(uid) / "expenses" / "expenses_master.xlsx"
    if not path.exists():
        raise HTTPException(404, "No expense ledger yet")
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    wb = load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col_index = {h: i + 1 for i, h in enumerate(headers)}  # 1-based

    target_row = None
    for row_num in range(2, ws.max_row + 1):
        row_dict = {h: ws.cell(row=row_num, column=col_index[h]).value for h in headers}
        if _expense_row_id(row_dict) == row_id:
            target_row = row_num
            break
    if target_row is None:
        raise HTTPException(404, f"Expense row '{row_id}' not found")

    updates = {k: v for k, v in body.items() if k in _EXPENSE_EDITABLE_FIELDS}
    for field, value in updates.items():
        if field in col_index:
            ws.cell(row=target_row, column=col_index[field]).value = value

    wb.save(path)

    updated_row = {h: ws.cell(row=target_row, column=col_index[h]).value for h in headers}
    updated_row["id"] = _expense_row_id(updated_row)
    return updated_row


@app.post("/api/expenses/scan")
async def scan_expenses_historical(
    body: dict, background_tasks: BackgroundTasks,
    session: dict = Depends(require_session),
):
    """Manually trigger a historical receipt scan over email attachments.
    Body: {days: int} — typically 90 or 180. Runs in background; results
    stream into expenses_master.xlsx incrementally so the user can see new
    rows appear by polling /api/expenses/all."""
    uid  = session["user_id"]
    days = int((body or {}).get("days") or 90)
    days = max(7, min(days, 365))

    # Stream logs into results/expenses.json (status=running + logs[-20:]) so
    # the global Activity drawer can poll /api/sections/expenses and show
    # live progress, same way run_section() does.
    results_dir = _udir(uid) / "results"
    results_dir.mkdir(parents=True, exist_ok=True)
    expense_result_path = results_dir / "expenses.json"
    _write_json(expense_result_path, {
        "id": "expenses", "status": "running",
        "started_at": datetime.now(timezone.utc).isoformat(),
        "scan_days": days,
        "logs": [f"Starting scan: last {days} days"],
    })

    def _run():
        from src.sections.expenses import run as expenses_run
        logs: list[str] = [f"Starting scan: last {days} days"]
        def _progress(msg: str):
            logs.append(msg)
            print(f"[ExpenseScan/{uid[:8]}] {msg}")
            _write_json(expense_result_path, {
                "id": "expenses", "status": "running",
                "started_at": datetime.now(timezone.utc).isoformat(),
                "scan_days": days,
                "logs": logs[-20:],
            })
        try:
            token = auth.get_valid_access_token(uid)
            graph = GraphClient(token)
            ai    = AIClient()
            result = expenses_run(graph, ai, _udir(uid), days=days, progress=_progress)
            result["logs"] = logs[-20:]
            _write_json(expense_result_path, result)
            print(f"[ExpenseScan] Done for {uid} — captured {result.get('count', 0)} new items ({days}d)")
        except Exception as e:
            _write_json(expense_result_path, {
                "id": "expenses", "status": "error",
                "error": str(e), "logs": logs[-20:],
                "last_run": datetime.now(timezone.utc).isoformat(),
            })
            print(f"[ExpenseScan] Failed for {uid}: {e}")

    background_tasks.add_task(_run)
    # Rough heuristic — ~1 min per ~15 days of inbox (depends on volume + OCR speed)
    est_min = max(2, days // 15)
    return {"ok": True, "days": days, "estimated_minutes": est_min}


@app.post("/api/expenses")
def create_expense(body: dict, session: dict = Depends(require_session)):
    """Manually add a new receipt row to expenses_master.xlsx."""
    uid  = session["user_id"]
    expenses_dir = _udir(uid) / "expenses"
    expenses_dir.mkdir(parents=True, exist_ok=True)
    path = expenses_dir / "expenses_master.xlsx"
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    headers = [
        "Date", "Vendor", "Amount", "Currency", "GST_HST", "Net_Amount",
        "Category", "Attachment", "Email_Subject", "From", "Msg_ID", "Att_ID", "Processed_Date",
    ]
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses"
        ws.append(headers)

    today = today_local_str(_udir(uid))
    manual_id = f"manual_{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
    new_row = [
        body.get("Date") or today,
        body.get("Vendor", ""),
        body.get("Amount", ""),
        body.get("Currency", "CAD"),
        body.get("GST_HST", ""),
        body.get("Net_Amount", ""),
        body.get("Category", "Other"),
        body.get("Attachment", ""),
        body.get("Email_Subject", "(manual entry)"),
        body.get("From", ""),
        manual_id,                # Msg_ID
        manual_id,                # Att_ID — stable id source
        today,                    # Processed_Date
    ]
    ws.append(new_row)
    wb.save(path)

    row_dict = dict(zip(headers, new_row))
    row_dict["id"] = _expense_row_id(row_dict)
    return row_dict


@app.get("/api/expenses/{row_id}/photo")
def get_expense_photo(row_id: str, session: dict = Depends(require_session)):
    """Stream the receipt photo from OneDrive (CEO Platform/Receipts/{attachment})."""
    from fastapi.responses import Response
    uid  = session["user_id"]
    path = _udir(uid) / "expenses" / "expenses_master.xlsx"
    if not path.exists():
        raise HTTPException(404, "No expense ledger yet")
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    wb = load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col_index = {h: i + 1 for i, h in enumerate(headers)}

    attachment_name = None
    for row_num in range(2, ws.max_row + 1):
        row_dict = {h: ws.cell(row=row_num, column=col_index[h]).value for h in headers}
        if _expense_row_id(row_dict) == row_id:
            attachment_name = row_dict.get("Attachment")
            break
    if not attachment_name:
        raise HTTPException(404, "Receipt row not found or has no attachment")

    try:
        token = auth.get_valid_access_token(uid)
        graph = GraphClient(token)
        from urllib.parse import quote
        encoded = quote(f"CEO Platform/Receipts/{attachment_name}", safe="/")
        # Get item metadata + downloadUrl
        meta = graph.get(f"/me/drive/root:/{encoded}")
        download_url = meta.get("@microsoft.graph.downloadUrl")
        mime = (meta.get("file") or {}).get("mimeType", "application/octet-stream")
        if not download_url:
            content = graph.download(f"/me/drive/root:/{encoded}:/content")
        else:
            import requests as _req
            r = _req.get(download_url, timeout=20)
            r.raise_for_status()
            content = r.content
        return Response(content=content, media_type=mime)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(404, f"Photo not in OneDrive: {e}")


@app.delete("/api/expenses/{row_id}")
def delete_expense(row_id: str, session: dict = Depends(require_session)):
    """Delete a row from expenses_master.xlsx."""
    uid  = session["user_id"]
    path = _udir(uid) / "expenses" / "expenses_master.xlsx"
    if not path.exists():
        raise HTTPException(404, "No expense ledger yet")
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    wb = load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col_index = {h: i + 1 for i, h in enumerate(headers)}

    target_row = None
    for row_num in range(2, ws.max_row + 1):
        row_dict = {h: ws.cell(row=row_num, column=col_index[h]).value for h in headers}
        if _expense_row_id(row_dict) == row_id:
            target_row = row_num
            break
    if target_row is None:
        raise HTTPException(404, f"Expense row '{row_id}' not found")

    ws.delete_rows(target_row, 1)
    wb.save(path)
    return {"ok": True}


# ── Company Intelligence Watchlist ────────────────────────

# ── Outreach (batch draft emails from OneDrive folder) ────

@app.post("/api/outreach/run")
async def trigger_outreach(request: Request, session: dict = Depends(require_session)):
    """Batch-generate Outlook drafts for contacts found in a OneDrive folder.
    Body: {"folder": "Conferences/TechConf", "context_note": "met at TechConf 2026"}
    folder defaults to settings.outreach_folder if omitted.
    """
    uid  = session["user_id"]
    body = await request.json()
    folder       = (body.get("folder") or "").strip()
    context_note = (body.get("context_note") or "").strip()

    try:
        from src.modules.outreach import run as _run_outreach
        token    = auth.get_valid_access_token(uid)
        graph    = GraphClient(token)
        ai       = AIClient()
        settings = _read_json(_user_settings(uid))
        result   = _run_outreach(
            graph=graph, ai=ai, data_dir=_udir(uid),
            settings=settings, folder=folder, context_note=context_note,
        )
        return result
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/outreach/last")
def get_last_outreach(session: dict = Depends(require_session)):
    """Return the most recent outreach run results."""
    uid = session["user_id"]
    f   = _udir(uid) / "outreach" / "results.json"
    if not f.exists():
        return {"status": "not_run"}
    try:
        return json.loads(f.read_text())
    except Exception:
        return {"status": "not_run"}


@app.get("/api/contacts/list")
def list_contacts(session: dict = Depends(require_session),
                  source: str = "",
                  tag: str = "",
                  since_hours: int = 0):
    """Flat list of CRM contacts for the frontend table view.

    Query params:
      source       — filter by source field (e.g. 'teams_card', 'email_scan')
      tag          — filter by tag (case-insensitive substring match)
      since_hours  — only contacts added in the last N hours

    Returns items sorted by added_at desc (newest first).
    """
    uid = session["user_id"]
    from src.modules.crm import load_crm
    crm = load_crm(_udir(uid))
    contacts = crm.get("contacts", {}).values()

    items = []
    cutoff = None
    if since_hours and since_hours > 0:
        cutoff = datetime.now(timezone.utc) - timedelta(hours=since_hours)
    tag_lower = tag.lower() if tag else ""

    for c in contacts:
        if source and c.get("source") != source:
            continue
        if tag_lower:
            if not any(tag_lower in t.lower() for t in (c.get("tags") or [])):
                continue
        if cutoff:
            added_at = c.get("added_at", "")
            if not added_at:
                continue
            try:
                dt = datetime.fromisoformat(added_at.replace("Z", "+00:00"))
            except Exception:
                continue
            if dt < cutoff:
                continue
        items.append({
            "email":        c.get("email", ""),
            "name":         c.get("name", ""),
            "company":      c.get("company", ""),
            "role":         c.get("role", ""),
            "phone":        c.get("phone", ""),
            "linkedin":     c.get("linkedin", ""),
            "source":       c.get("source", ""),
            "tags":         c.get("tags") or [],
            "added_at":     c.get("added_at", ""),
            "last_contact": c.get("last_contact", ""),
            "draft_link":   c.get("draft_link", ""),
            "priority":     c.get("priority", ""),
            "status":       c.get("status", ""),
            "notes":        c.get("notes", ""),
        })

    items.sort(key=lambda x: x.get("added_at", ""), reverse=True)
    return {"items": items, "count": len(items)}


# ── User Profile (business_profile + market_segments) ────

@app.get("/api/profile")
def get_profile(session: dict = Depends(require_session)):
    from src.modules.profile import load_business_profile, load_market_segments, load_personal_profile
    uid = session["user_id"]
    return {
        "business_profile": load_business_profile(_udir(uid)),
        "market_segments":  load_market_segments(_udir(uid)),
        "personal_profile": load_personal_profile(_udir(uid)),
    }


@app.post("/api/profile/business")
async def save_business_profile_endpoint(request: Request, session: dict = Depends(require_session)):
    from src.modules.profile import save_business_profile
    uid  = session["user_id"]
    body = await request.json()
    text = str(body.get("content") or body.get("text") or "")
    save_business_profile(_udir(uid), text)
    return {"ok": True}


@app.post("/api/profile/segments")
async def save_market_segments_endpoint(request: Request, session: dict = Depends(require_session)):
    from src.modules.profile import save_market_segments
    uid  = session["user_id"]
    body = await request.json()
    text = str(body.get("content") or body.get("text") or "")
    save_market_segments(_udir(uid), text)
    return {"ok": True}


@app.post("/api/profile/personal")
async def save_personal_profile_endpoint(request: Request, session: dict = Depends(require_session)):
    from src.modules.profile import save_personal_profile
    uid  = session["user_id"]
    body = await request.json()
    text = str(body.get("content") or body.get("text") or "")
    save_personal_profile(_udir(uid), text)
    return {"ok": True}


@app.get("/api/profile/status")
def get_profile_status(session: dict = Depends(require_session)):
    from src.modules.profile import load_init_status, init_has_started, save_init_status
    uid  = session["user_id"]
    udir = _udir(uid)
    # If init has never been triggered for this user (e.g. existing session predating
    # this feature, or fresh OAuth that hasn't been confirmed yet), surface the
    # confirmation gate so the frontend can show "Logged in as X — start setup?"
    if not init_has_started(udir):
        token_data = auth.load_user_tokens(uid) or {}
        username = token_data.get("username", "")
        save_init_status(udir, "awaiting_confirmation",
                         current_message=f"Logged in as {username}. Confirm to start setup." if username
                                         else "Confirm to start setup.")
    return load_init_status(udir)


@app.post("/api/profile/confirm")
def confirm_profile(session: dict = Depends(require_session)):
    from src.modules.profile import save_init_status
    save_init_status(_udir(session["user_id"]), "user_confirmed")
    return {"ok": True}


@app.post("/api/profile/refresh-signature")
def refresh_signature(session: dict = Depends(require_session)):
    """Run ONLY the signature-extraction step from profile_init — pull the
    user's recent sent emails, AI-extract the recurring HTML signature
    block, persist to settings.email_signature.

    This is the same logic run_profile_init does inline, lifted out so the
    frontend can trigger it lazily (on first app load when the field is
    missing) without forcing the user through a full Restart Onboarding
    (which re-scans 12 months of inbox, rebuilds CRM, etc.). Fast: one
    Graph fetch + one AI call, typically 3-5 seconds total.

    Returns the resolved signature (helpful for the caller to confirm it
    wrote something usable). Synthesises a fallback from existing
    settings (display_name + title + company + email) when AI can't find
    a consistent block — so a brand-new account with no sent history still
    gets a usable signature."""
    from src.modules.profile_init import (
        _collect_signature_samples, _extract_signature,
        _synthesize_signature, _user_title_from_signatures,
    )

    uid      = session["user_id"]
    token    = auth.get_valid_access_token(uid)
    graph    = GraphClient(token)
    ai       = AIClient()
    settings = _read_json(_user_settings(uid))

    samples = _collect_signature_samples(graph, log=lambda m: print(f"[RefreshSig] {m}"))
    html_samples  = samples.get("html", [])
    plain_samples = samples.get("plain", [])
    user_title    = _user_title_from_signatures(plain_samples)
    signature     = _extract_signature(html_samples, ai, log=lambda m: print(f"[RefreshSig] {m}"))
    fallback      = False
    if not signature:
        signature = _synthesize_signature(settings, user_title)
        fallback  = True

    if signature:
        settings["email_signature"] = signature
        _write_json(_user_settings(uid), settings)
        print(f"[RefreshSig] {uid}: stored {'synthesised' if fallback else 'extracted'} "
              f"signature ({len(signature)} chars)")
    return {
        "signature": signature,
        "source":    "fallback" if fallback else "extracted",
        "html":      "<" in signature,
    }


@app.post("/api/init/start")
def start_init(session: dict = Depends(require_session)):
    """User has confirmed the right account is logged in — start the first-login init chain.
    Idempotent: if init is already running or completed, returns the current stage without re-triggering.
    Uses the history_months saved in settings (default 12) — for explicit control, use
    POST /api/onboarding/preferences instead."""
    from src.modules.profile import load_init_status, save_init_status
    uid = session["user_id"]
    stage = load_init_status(_udir(uid)).get("stage", "pending")
    if stage in ("generating", "draft_ready", "user_confirmed"):
        return {"ok": True, "stage": stage, "message": "Init already started or complete"}
    # awaiting_confirmation or pending → kick off
    save_init_status(_udir(uid), "generating")
    threading.Thread(target=_run_first_login_init, args=(uid,), daemon=True).start()
    print(f"[init] User confirmed — starting first-login chain for {uid}")
    return {"ok": True, "stage": "generating"}


@app.post("/api/onboarding/enrich-website")
async def enrich_company_website(request: Request, session: dict = Depends(require_session)):
    """Onboarding Step 1 helper. User pastes their company URL; we call
    Gemini's Google Search grounding to infer company_name + description +
    market segments. Frontend uses the result to pre-fill Step 3's editable
    company name field and bundles the rest into the final preferences POST.

    Always returns 200 with whatever we managed to extract. On AI failure,
    falls back to a URL-based heuristic for company_name and empty values
    for the rest.

    Body: {url: str}
    Returns: {company_name: str, description: str, segments: list[str], source: 'ai'|'fallback'}
    """
    from src.modules.website_intel import (
        extract_company_info, extract_company_name_fallback,
    )
    body = await request.json()
    url = str(body.get("url") or "").strip()
    if not url:
        raise HTTPException(400, "url is required")

    try:
        info = extract_company_info(url, AIClient())
        # AI sometimes returns empty company_name even on a 200 response.
        if not info.get("company_name"):
            info["company_name"] = extract_company_name_fallback(url)
            info["source"] = "ai_partial"
        else:
            info["source"] = "ai"
        return info
    except Exception as e:
        print(f"[Onboarding/enrich] AI extraction failed for {url}: {e}")
        return {
            "company_name":   extract_company_name_fallback(url),
            "headquarters":   "",
            "what_they_do":   "",
            "business_lines": [],
            "products":       [],
            "segments":       [],
            "role_emails":    [],
            "description":    "",
            "source":         "fallback",
            "error":          str(e)[:120],
        }


@app.post("/api/onboarding/preferences")
async def submit_onboarding_preferences(request: Request, session: dict = Depends(require_session)):
    """Wizard-driven init: persists company + bio context provided by the new
    5-step wizard, applies default briefing/monitor presets, then kicks off
    the init chain in the background.

    Body fields (all optional except where noted):
      company_website_url:     str — user's pasted URL (Step 1)
      company_name:            str — AI-extracted or user-edited (Step 3)
      company_description:     str — legacy alias for company_what_they_do
      company_what_they_do:    str — AI-extracted, 2-3 sentence company self-description
      company_headquarters:    str — "City, Region, Country"
      company_business_lines:  list[str] — distinct service areas
      company_products:        list[str] — named products/services
      company_market_segments: list[str] — industries
      company_role_emails:     list[str] — functional emails (info@, sales@, etc.)
      linkedin_bio:            str — user pasted (Step 2, optional)
      history_months:          int — defaults to 12 (UI no longer exposes this)
      briefings_preset:        str — defaults to 'recommended'
      monitor_preset:          str — defaults to 'recommended'
      personal_profile:        str — legacy field, kept for backwards compat
    """
    from src.modules.profile import (
        load_init_status, save_init_status, save_personal_profile,
    )
    from src.modules.schedules import (
        apply_briefing_preset, apply_monitor_preset, load_schedules,
    )

    uid  = session["user_id"]
    body = await request.json()

    history_months   = max(1, min(int(body.get("history_months") or 12), 24))
    briefings_preset = (body.get("briefings_preset") or "recommended").lower()
    monitor_preset   = (body.get("monitor_preset")   or "recommended").lower()
    personal_text    = body.get("personal_profile")

    # New: company + user context from the 5-step wizard
    company_website_url     = str(body.get("company_website_url") or "").strip()
    company_name            = str(body.get("company_name") or "").strip()
    company_description     = str(body.get("company_description") or "").strip()
    company_market_segments = body.get("company_market_segments") or []
    if not isinstance(company_market_segments, list):
        company_market_segments = []
    company_market_segments = [str(s).strip() for s in company_market_segments if str(s).strip()][:10]
    linkedin_bio            = str(body.get("linkedin_bio") or "").strip()

    # Step 1 (website_intel) richer fields — fed into business_profile v1.
    company_headquarters    = str(body.get("company_headquarters") or "").strip()
    company_what_they_do    = str(body.get("company_what_they_do") or "").strip()
    def _list_field(key: str, cap: int) -> list[str]:
        raw_list = body.get(key) or []
        if not isinstance(raw_list, list):
            return []
        return [str(s).strip() for s in raw_list if str(s).strip()][:cap]
    company_business_lines  = _list_field("company_business_lines", 8)
    company_products        = _list_field("company_products",       8)
    company_role_emails     = _list_field("company_role_emails",    8)

    # 1. Persist everything to settings so _run_first_login_init / profile_init
    #    pick it up. Branding (company_name) is also read by the frontend
    #    sidebar to rebrand the dashboard after onboarding.
    settings = _read_json(_user_settings(uid))
    settings["onboarding_history_months"] = history_months
    if company_website_url:     settings["company_website_url"]     = company_website_url
    if company_name:            settings["company_name"]            = company_name
    if company_description:     settings["company_description"]     = company_description
    if company_market_segments: settings["company_market_segments"] = company_market_segments
    if linkedin_bio:            settings["linkedin_bio"]            = linkedin_bio
    if company_headquarters:    settings["company_headquarters"]    = company_headquarters
    if company_what_they_do:    settings["company_what_they_do"]    = company_what_they_do
    if company_business_lines:  settings["company_business_lines"]  = company_business_lines
    if company_products:        settings["company_products"]        = company_products
    if company_role_emails:     settings["company_role_emails"]     = company_role_emails
    _write_json(_user_settings(uid), settings)
    tz = settings.get("timezone") or "UTC"

    # 2. Legacy personal_profile field (the old 4-step wizard's Step 4 textarea).
    #    New wizard doesn't send this — profile_init will draft it from linkedin_bio
    #    + signatures during the init chain.
    if personal_text:
        save_personal_profile(_udir(uid), str(personal_text))

    # 3. Schedule presets (skip 'custom' so we don't clobber prior config)
    if briefings_preset != "custom":
        apply_briefing_preset(_udir(uid), briefings_preset, tz)
        # Hot-register the new briefings with the live scheduler so cron starts ticking
        # without waiting for a server restart.
        for b in load_schedules(_udir(uid)).get("briefings", []):
            try:
                _register_briefing(uid, b)
            except Exception as e:
                print(f"[Onboarding] briefing register failed for {b.get('id')}: {e}")
    if monitor_preset != "custom":
        apply_monitor_preset(_udir(uid), monitor_preset)

    # 4. Trigger init (idempotent)
    stage = load_init_status(_udir(uid)).get("stage", "pending")
    if stage in ("generating", "draft_ready", "user_confirmed"):
        return {"ok": True, "stage": stage, "message": "Init already running"}
    save_init_status(_udir(uid), "generating")
    threading.Thread(
        target=_run_first_login_init,
        args=(uid,),
        kwargs={"history_months": history_months},
        daemon=True,
    ).start()
    print(f"[Onboarding] wizard submitted for {uid} — history={history_months}mo, "
          f"briefings={briefings_preset}, monitor={monitor_preset}, "
          f"company={company_name or '(none)'}, has_bio={bool(linkedin_bio)}")
    return {"ok": True, "stage": "generating", "history_months": history_months}


@app.get("/api/init/status")
def get_init_status(session: dict = Depends(require_session)):
    """Frontend polls this to know what to show: awaiting_confirmation / generating / draft_ready / user_confirmed."""
    from src.modules.profile import load_init_status
    return load_init_status(_udir(session["user_id"]))


@app.post("/api/onboarding/restart")
def restart_onboarding(session: dict = Depends(require_session)):
    """Wipe init artifacts and send the user back through the wizard. Unlike
    /api/init/reset (which immediately re-runs init), this leaves stage at
    'awaiting_confirmation' so the OnboardingWizard renders again. The user
    re-picks history depth / briefing preset / monitor preset before any
    scan starts."""
    from src.modules.profile import save_init_status
    uid  = session["user_id"]
    udir = _udir(uid)
    for fname in ("crm.json", "projects.json"):
        p = udir / fname
        if p.exists():
            try: p.unlink()
            except Exception: pass
    status_path = udir / "profile" / "init_status.json"
    if status_path.exists():
        try: status_path.unlink()
        except Exception: pass
    save_init_status(udir, "awaiting_confirmation",
                     current_message="Awaiting new preferences from wizard.")
    print(f"[Onboarding] Restart requested for {uid} — wizard will run again")
    return {"ok": True, "stage": "awaiting_confirmation"}


@app.post("/api/init/reset")
def reset_init(session: dict = Depends(require_session)):
    """Reset init state and immediately start a fresh init chain.
    Use this when init is stuck mid-way (e.g. process restart left stage='generating' with no live worker).
    Deletes init_status.json and partial crm.json/projects.json so the chain starts from scratch."""
    uid = session["user_id"]
    udir = _udir(uid)
    # Clear init artifacts so the chain rebuilds cleanly
    for fname in ("crm.json", "projects.json"):
        p = udir / fname
        if p.exists():
            try:
                p.unlink()
            except Exception:
                pass
    status_path = udir / "profile" / "init_status.json"
    if status_path.exists():
        try:
            status_path.unlink()
        except Exception:
            pass
    # Kick a fresh run
    from src.modules.profile import save_init_status
    save_init_status(udir, "generating", current_message="Restarting setup...")
    threading.Thread(target=_run_first_login_init, args=(uid,), daemon=True).start()
    print(f"[init] Reset and restart triggered for {uid}")
    return {"ok": True, "stage": "generating", "message": "Init reset and restarted"}


@app.post("/api/profile/regenerate")
def regenerate_profile(session: dict = Depends(require_session)):
    from src.modules.profile import save_init_status, load_init_status
    uid = session["user_id"]
    # Only set status to "generating" if user hasn't confirmed yet.
    # Confirmed users see a Settings spinner instead — they shouldn't be re-routed to Onboarding.
    if load_init_status(_udir(uid)).get("stage") != "user_confirmed":
        save_init_status(_udir(uid), "generating")
    threading.Thread(target=_regenerate_profile_for_user, args=(uid,), daemon=True).start()
    return {"ok": True}


# ── DB Cleanup (duplicate detection + stale archival) ────

def _run_cleanup_scan_for_user(uid: str, auto_apply: bool = False, send_digest: bool = False) -> None:
    """Run a full scan for one user. If auto_apply, immediately approve high-confidence
    duplicates and/or stale records based on user's settings. If send_digest, push a
    Teams Adaptive Card summary to the user's bot chat afterwards."""
    from src.modules.db_cleaner import (
        run_full_scan, approve_candidates, load_pending, load_last_scan, build_digest_card,
    )
    # Skip if no data to scan (e.g. bot accounts have no crm/projects)
    if not (_udir(uid) / "crm.json").exists() and not (_udir(uid) / "projects.json").exists():
        return

    try:
        ai = AIClient()
        settings = _read_json(_user_settings(uid))
        run_full_scan(_udir(uid), ai)

        auto_applied_count = 0
        if auto_apply:
            pending = load_pending(_udir(uid))
            auto_ids: list[str] = []
            if settings.get("auto_merge_high_confidence") in (True, "true"):
                auto_ids += [c["id"] for c in pending
                             if c.get("confidence") == "high"
                             and c.get("type") in ("project_duplicate", "contact_duplicate")]
            if settings.get("auto_archive_stale") in (True, "true"):
                auto_ids += [c["id"] for c in pending
                             if c.get("type") in ("stale_project", "stale_contact")]
            if auto_ids:
                result = approve_candidates(_udir(uid), auto_ids)
                auto_applied_count = len(auto_ids)
                print(f"[DBCleanup] {uid} auto-applied {auto_applied_count} candidates: {result}")

        if send_digest and settings.get("cleanup_digest_enabled") not in (False, "false"):
            summary = load_last_scan(_udir(uid)).get("summary", {})
            # Subtract auto-applied from the high/stale totals shown to the user
            if auto_applied_count:
                summary = dict(summary)
                summary["total"] = max(0, summary.get("total", 0) - auto_applied_count)

            origin = os.getenv("FRONTEND_URL") or FRONTEND_URL
            cleanup_url = f"{origin.rstrip('/')}/?page=records&tab=cleanup"
            card = build_digest_card(summary, cleanup_url, auto_applied=auto_applied_count)
            _send_card_to_owner_bot_chat(uid, card)
    except Exception as e:
        print(f"[DBCleanup] Scan failed for {uid}: {e}")


def _send_card_to_owner_bot_chat(owner_uid: str, card: dict) -> None:
    """Find the bot account linked to this owner and push a card to their 1:1 chat."""
    from src.modules.email_monitor import _send_adaptive_card
    sessions_dir = auth.DATA_DIR / "_sessions"
    if not sessions_dir.exists():
        return
    for f in sessions_dir.glob("*.json"):
        bot_uid = f.stem
        bp = _bot_state_path(bot_uid)
        if not bp.exists():
            continue
        try:
            bs = _read_json(bp)
            if (bs.get("enabled")
                    and bs.get("is_registered_bot")
                    and bs.get("owner_uid") == owner_uid
                    and bs.get("chat_id")):
                bot_graph = GraphClient(auth.get_valid_access_token(bot_uid))
                _send_adaptive_card(bot_graph, bs["chat_id"], card)
                return
        except Exception as e:
            print(f"[DBCleanup] Card push failed for owner={owner_uid} bot={bot_uid}: {e}")


def _weekly_cleanup_scan_all_users() -> None:
    sessions_dir = auth.DATA_DIR / "_sessions"
    if not sessions_dir.exists():
        return
    for f in sessions_dir.glob("*.json"):
        threading.Thread(
            target=_run_cleanup_scan_for_user,
            kwargs={"uid": f.stem, "auto_apply": True, "send_digest": True},
            daemon=True,
        ).start()


@app.get("/api/db-cleanup/pending")
def db_cleanup_pending(session: dict = Depends(require_session)):
    from src.modules.db_cleaner import load_pending, load_last_scan
    uid = session["user_id"]
    return {
        "candidates": load_pending(_udir(uid)),
        "last_scan":  load_last_scan(_udir(uid)),
    }


@app.post("/api/db-cleanup/scan")
def db_cleanup_scan(session: dict = Depends(require_session)):
    uid = session["user_id"]
    threading.Thread(
        target=_run_cleanup_scan_for_user,
        args=(uid, False),  # manual scan: never auto-apply (user wants to review)
        daemon=True,
    ).start()
    return {"ok": True, "message": "Scan started in background"}


@app.post("/api/db-cleanup/approve")
async def db_cleanup_approve(request: Request, session: dict = Depends(require_session)):
    from src.modules.db_cleaner import approve_candidates
    uid  = session["user_id"]
    body = await request.json()
    ids  = body.get("ids") or []
    if not isinstance(ids, list) or not ids:
        raise HTTPException(400, "ids must be a non-empty list")
    return approve_candidates(_udir(uid), [str(i) for i in ids])


@app.post("/api/db-cleanup/reject")
async def db_cleanup_reject(request: Request, session: dict = Depends(require_session)):
    from src.modules.db_cleaner import reject_candidates
    uid  = session["user_id"]
    body = await request.json()
    ids  = body.get("ids") or []
    if not isinstance(ids, list) or not ids:
        raise HTTPException(400, "ids must be a non-empty list")
    return reject_candidates(_udir(uid), [str(i) for i in ids])


@app.post("/api/db-cleanup/merge-projects-direct")
async def db_cleanup_merge_projects_direct(request: Request, session: dict = Depends(require_session)):
    """Manual merge: user picks both projects, no AI candidate required."""
    from src.modules.db_cleaner import merge_projects
    uid  = session["user_id"]
    body = await request.json()
    keep_id  = str(body.get("keep_id") or "").strip()
    merge_id = str(body.get("merge_id") or "").strip()
    if not keep_id or not merge_id:
        raise HTTPException(400, "keep_id and merge_id are required")
    if keep_id == merge_id:
        raise HTTPException(400, "keep_id and merge_id must be different")
    result = merge_projects(_udir(uid), keep_id, merge_id)
    if not result.get("ok"):
        raise HTTPException(400, result.get("error", "merge failed"))
    return result


@app.post("/api/db-cleanup/merge-contacts-direct")
async def db_cleanup_merge_contacts_direct(request: Request, session: dict = Depends(require_session)):
    """Manual merge: user picks both contacts, no AI candidate required."""
    from src.modules.db_cleaner import merge_contacts
    uid  = session["user_id"]
    body = await request.json()
    keep_email  = str(body.get("keep_email") or "").strip()
    merge_email = str(body.get("merge_email") or "").strip()
    if not keep_email or not merge_email:
        raise HTTPException(400, "keep_email and merge_email are required")
    if keep_email.lower() == merge_email.lower():
        raise HTTPException(400, "keep_email and merge_email must be different")
    result = merge_contacts(_udir(uid), keep_email, merge_email)
    if not result.get("ok"):
        raise HTTPException(400, result.get("error", "merge failed"))
    return result


@app.post("/api/db-cleanup/archive")
async def db_cleanup_archive_direct(request: Request, session: dict = Depends(require_session)):
    """Manual archive: set archived=true on a single project or contact."""
    from src.modules.db_cleaner import archive_record
    uid  = session["user_id"]
    body = await request.json()
    kind = str(body.get("kind") or "").strip()
    ident = str(body.get("id") or "").strip()
    if kind not in ("project", "contact"):
        raise HTTPException(400, "kind must be 'project' or 'contact'")
    if not ident:
        raise HTTPException(400, "id is required")
    result = archive_record(_udir(uid), kind, ident)
    if not result.get("ok"):
        raise HTTPException(400, result.get("error", "archive failed"))
    return result


# ── Health check ──────────────────────────────────────────

@app.get("/health")
def health():
    return {"status": "ok"}


# ── Per-user briefing schedules (schedules.json + APScheduler) ────────────

def _briefing_job_id(uid: str, briefing_id: str) -> str:
    return f"user_{uid}_briefing_{briefing_id}"


def _run_briefing_for_user(uid: str, briefing_id: str, force: bool = False) -> None:
    """Run all skills in a briefing sequentially, each pushed to Teams.
    force=True bypasses the enabled check (used by Test Run).
    Sends a 'complete' footer to Teams so the user knows the briefing ran even
    if individual skills had no results."""
    from src.modules.schedules import load_schedules
    schedules = load_schedules(_udir(uid))
    briefing = next((b for b in schedules.get("briefings", []) if b.get("id") == briefing_id), None)
    if not briefing:
        return
    if not force and not briefing.get("enabled"):
        return
    skill_ids = briefing.get("skills", [])
    print(f"[Briefing] Firing '{briefing.get('name')}' for {uid}{' (test run)' if force else ''}: {skill_ids}")

    pushed = []
    empty = []
    failed = []
    for sid in skill_ids:
        if sid not in _SECTION_RUNNERS:
            failed.append((sid, "unknown skill"))
            continue
        try:
            # Mirror _run_section_for_user but track pushed vs empty
            results_dir = _udir(uid) / "results"
            results_dir.mkdir(parents=True, exist_ok=True)
            # Freshness check: if ai_summary already refreshed this dep within
            # the last 30 minutes (or any other recent run wrote it), reuse the
            # file instead of double-running the same skill in one briefing.
            existing = _read_json(results_dir / f"{sid}.json")
            if existing and _within(existing.get("last_run"), 30 * 60):
                result = existing
            else:
                token    = auth.get_valid_access_token(uid)
                graph    = GraphClient(token)
                ai       = AIClient()
                settings = _read_json(_user_settings(uid))
                result   = _SECTION_RUNNERS[sid](graph, ai, _udir(uid), settings, None)
                _write_json(results_dir / f"{sid}.json", result)
            msg = _format_section_for_teams(result, get_user_tz(_udir(uid)))
            if msg:
                bot_uid, chat_id = _find_bot_for_user(uid)
                if bot_uid and chat_id:
                    bot_token = auth.get_valid_access_token(bot_uid)
                    bot_graph = GraphClient(bot_token)
                    bot_graph.send_chat_message(chat_id, msg)
                    _log_outbound_to_history(uid, msg)
                pushed.append(sid)
            else:
                empty.append(sid)
        except Exception as e:
            failed.append((sid, str(e)[:120]))
            print(f"[Briefing]   Skill {sid} failed: {e}")

    # Brief footer so user knows the whole briefing finished
    bot_uid, chat_id = _find_bot_for_user(uid)
    if bot_uid and chat_id:
        suffix = " (test run)" if force else ""
        total = len(pushed) + len(empty)
        if failed:
            footer = (
                f"⚠️ **Briefing '{briefing.get('name')}' finished with errors{suffix}** — "
                f"{total} ok, {len(failed)} failed: {', '.join(s for s, _ in failed)}"
            )
        else:
            footer = f"✅ **Briefing '{briefing.get('name')}' complete{suffix}** — {total} skill(s) reported."
        try:
            bot_token = auth.get_valid_access_token(bot_uid)
            bot_graph = GraphClient(bot_token)
            bot_graph.send_chat_message(chat_id, footer)
            _log_outbound_to_history(uid, footer)
        except Exception as e:
            print(f"[Briefing] Footer send failed for {uid}: {e}")


def _register_briefing(uid: str, briefing: dict) -> None:
    from src.modules.schedules import cron_to_apscheduler_kwargs, is_valid_cron
    job_id = _briefing_job_id(uid, briefing["id"])
    try:
        _scheduler.remove_job(job_id)
    except Exception:
        pass
    if not briefing.get("enabled"):
        return
    cron = briefing.get("cron", "")
    if not is_valid_cron(cron):
        print(f"[Scheduler] Skipping {job_id}: invalid cron {cron!r}")
        return
    kwargs = cron_to_apscheduler_kwargs(cron)
    tz = briefing.get("tz") or "UTC"
    try:
        _scheduler.add_job(
            _run_briefing_for_user,
            trigger="cron",
            kwargs={"uid": uid, "briefing_id": briefing["id"]},
            id=job_id,
            replace_existing=True,
            max_instances=1,
            coalesce=True,
            timezone=tz,
            **kwargs,
        )
    except Exception as e:
        # Likely invalid timezone — fall back to UTC
        print(f"[Scheduler] tz={tz!r} rejected for {job_id} ({e}); using UTC")
        _scheduler.add_job(
            _run_briefing_for_user,
            trigger="cron",
            kwargs={"uid": uid, "briefing_id": briefing["id"]},
            id=job_id,
            replace_existing=True,
            max_instances=1,
            coalesce=True,
            **kwargs,
        )


def _unregister_briefing(uid: str, briefing_id: str) -> None:
    try:
        _scheduler.remove_job(_briefing_job_id(uid, briefing_id))
    except Exception:
        pass


def _load_all_user_schedules_at_startup() -> None:
    """Scan .data/_sessions/*.json for every uid, load their schedules.json, register briefing jobs."""
    from src.modules.schedules import load_schedules
    sessions_dir = auth.DATA_DIR / "_sessions"
    if not sessions_dir.exists():
        return
    n = 0
    for f in sessions_dir.glob("*.json"):
        uid = f.stem
        schedules = load_schedules(_udir(uid))
        for briefing in schedules.get("briefings", []):
            try:
                _register_briefing(uid, briefing)
                if briefing.get("enabled"):
                    n += 1
            except Exception as e:
                print(f"[Scheduler] Failed to register briefing {uid}/{briefing.get('id')}: {e}")
    print(f"[Scheduler] Registered {n} briefing job(s) from schedules.json")


# ── Schedule API endpoints ────────────────────────────────

@app.get("/api/schedules")
def get_schedules(session: dict = Depends(require_session)):
    """Return full schedule config: briefings + email_monitor + meeting."""
    from src.modules.schedules import load_schedules
    return load_schedules(_udir(session["user_id"]))


@app.post("/api/schedules/briefings")
def create_briefing(body: dict, session: dict = Depends(require_session)):
    """Add a new briefing bundle."""
    from src.modules.schedules import add_briefing, NON_SCHEDULABLE_SECTIONS
    uid = session["user_id"]
    skills = body.get("skills", [])
    bad = [s for s in skills if s in NON_SCHEDULABLE_SECTIONS]
    if bad:
        raise HTTPException(400, f"Non-schedulable skills: {bad}")
    unknown = [s for s in skills if s not in _SECTION_RUNNERS]
    if unknown:
        raise HTTPException(400, f"Unknown skills: {unknown}")
    try:
        entry = add_briefing(_udir(uid), body)
    except ValueError as e:
        raise HTTPException(400, str(e))
    _register_briefing(uid, entry)
    return entry


@app.patch("/api/schedules/briefings/{briefing_id}")
def patch_briefing(briefing_id: str, body: dict, session: dict = Depends(require_session)):
    from src.modules.schedules import update_briefing, NON_SCHEDULABLE_SECTIONS
    uid = session["user_id"]
    if "skills" in body:
        bad = [s for s in body["skills"] if s in NON_SCHEDULABLE_SECTIONS]
        if bad:
            raise HTTPException(400, f"Non-schedulable skills: {bad}")
        unknown = [s for s in body["skills"] if s not in _SECTION_RUNNERS]
        if unknown:
            raise HTTPException(400, f"Unknown skills: {unknown}")
    try:
        entry = update_briefing(_udir(uid), briefing_id, body)
    except KeyError as e:
        raise HTTPException(404, str(e))
    except ValueError as e:
        raise HTTPException(400, str(e))
    _register_briefing(uid, entry)
    return entry


@app.delete("/api/schedules/briefings/{briefing_id}")
def remove_briefing(briefing_id: str, session: dict = Depends(require_session)):
    from src.modules.schedules import delete_briefing
    uid = session["user_id"]
    try:
        delete_briefing(_udir(uid), briefing_id)
    except KeyError as e:
        raise HTTPException(404, str(e))
    _unregister_briefing(uid, briefing_id)
    return {"ok": True}


@app.post("/api/bot/chat")
def bot_chat(body: dict, session: dict = Depends(require_session)):
    """Send a message to the same Audrey bot agent that runs in Teams.
    Re-uses src.bot.reply — full tool access, shared SQLite history."""
    uid = session["user_id"]
    text = (body.get("message") or "").strip()
    if not text:
        raise HTTPException(400, "message required")
    data_dir = _udir(uid)
    settings = _read_json(_user_settings(uid))
    bot_state = _read_json(data_dir / "teams_bot.json") or {}

    try:
        owner_graph = GraphClient(auth.get_valid_access_token(uid))
    except Exception as e:
        raise HTTPException(500, f"Auth failed: {e}")

    try:
        from src import bot as _bot
        reply_text, new_state = _bot.reply(
            bot_state, text,
            graph=owner_graph,        # fallback — web has no separate bot account
            owner_graph=owner_graph,
            settings=settings,
            wiki_dir=data_dir / "wiki",
            data_dir=data_dir,
            user_model_path=data_dir / "user_model.json",
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Bot error: {e}")

    try:
        _write_json(data_dir / "teams_bot.json", new_state)
    except Exception as e:
        print(f"[BotChat] State save failed for {uid}: {e}")

    return {"response": reply_text}


@app.get("/api/bot/history")
def bot_history(limit: int = 20, session: dict = Depends(require_session)):
    """Return the most recent N turns of the bot conversation (shared with Teams)."""
    from src.modules.db_helpers import open_sqlite
    uid = session["user_id"]
    db_path = _udir(uid) / "bot_history.db"
    if not db_path.exists():
        return {"turns": []}
    limit = max(1, min(int(limit or 20), 100))
    try:
        con = open_sqlite(db_path)
        rows = con.execute(
            "SELECT role, content, ts FROM history ORDER BY ts DESC LIMIT ?",
            (limit,),
        ).fetchall()
        con.close()
        rows.reverse()
        return {"turns": [{"role": r, "content": c, "ts": t} for r, c, t in rows]}
    except Exception as e:
        raise HTTPException(500, f"History read failed: {e}")


@app.post("/api/schedules/briefings/{briefing_id}/run")
def test_run_briefing(briefing_id: str, background_tasks: BackgroundTasks,
                     session: dict = Depends(require_session)):
    """Trigger a briefing immediately (Test Run). Runs in background."""
    from src.modules.schedules import load_schedules
    uid = session["user_id"]
    schedules = load_schedules(_udir(uid))
    briefing = next((b for b in schedules.get("briefings", []) if b.get("id") == briefing_id), None)
    if not briefing:
        raise HTTPException(404, f"Briefing '{briefing_id}' not found")
    background_tasks.add_task(_run_briefing_for_user, uid, briefing_id, True)
    return {"ok": True, "skills": briefing.get("skills", [])}


@app.put("/api/schedules/email_monitor")
def put_email_monitor(body: dict, session: dict = Depends(require_session)):
    from src.modules.schedules import update_email_monitor
    return update_email_monitor(_udir(session["user_id"]), body)


@app.put("/api/schedules/meeting")
def put_meeting_config(body: dict, session: dict = Depends(require_session)):
    from src.modules.schedules import update_meeting
    return update_meeting(_udir(session["user_id"]), body)


# ── Outlook draft redirector (mobile fallback) ────────────
# Teams chat is plain HTML — there's no way to ask the client "try the app,
# else open the web". So we route the click through this endpoint, which
# returns a tiny HTML page that tries the ms-outlook:// scheme first and
# falls back to the original webLink after a short delay. On mobile with
# Outlook installed → app opens; everywhere else → browser opens the draft.
_OUTLOOK_HOST_ALLOWLIST = (
    "outlook.office.com",
    "outlook.office365.com",
    "outlook.live.com",
)


@app.get("/r/draft")
def redirect_to_draft(url: str = "", request: Request = None):
    """Public mobile-friendly redirector for Outlook draft webLinks.
    Whitelists Outlook hosts so the endpoint can't be abused as an open
    redirect (phishing) — anything else returns 400.

    Mobile gets an HTTP 302 straight to ms-outlook://drafts so there's no
    HTML page to render — the browser hands the scheme to the OS instantly,
    Outlook app opens, no "Opening Outlook…" flash. Desktop still gets the
    iframe-based HTML page since it needs the multi-scheme fallback chain
    and a graceful web fallback when Outlook desktop isn't installed.

    JS strategy on the returned page:
      1. Feed ms-outlook://emails/<itemid> to a hidden iframe (undocumented
         scheme — Outlook may or may not recognize it; if it does, the app
         opens directly to that draft).
      2. After 600ms, feed ms-outlook://drafts to the iframe — if the app
         is installed but didn't accept #1, it now opens to the drafts
         folder (the new draft will be at the top by modified time).
      3. After another 800ms with the page still in foreground, fall back
         to the original webLink (browser-based).

    Using an iframe instead of window.location lets us silently try custom
    schemes without browser error pages when no handler is registered."""
    from urllib.parse import urlparse, parse_qs
    import html as _html

    if not url:
        raise HTTPException(400, "Missing url parameter")
    try:
        parsed = urlparse(url)
    except Exception:
        raise HTTPException(400, "Malformed url")
    if parsed.scheme not in ("http", "https"):
        raise HTTPException(400, "Only http(s) URLs allowed")
    host = (parsed.hostname or "").lower()
    if not any(host == h or host.endswith("." + h) for h in _OUTLOOK_HOST_ALLOWLIST):
        raise HTTPException(400, f"Host '{host}' not in Outlook allowlist")

    # Mobile: 302 straight to the OS-handled scheme. No HTML, no "Opening
    # Outlook…" flash — the browser hands the scheme off to Outlook the
    # instant it sees the response header.
    ua = (request.headers.get("user-agent", "") if request else "").lower()
    is_mobile = ("iphone" in ua) or ("ipad" in ua) or ("ipod" in ua) or \
                ("android" in ua and "mobile" in ua)
    if is_mobile:
        return RedirectResponse(url="ms-outlook://drafts", status_code=302)

    # Try to extract the draft itemid from the webLink for the experimental
    # per-item deep link. webLink format is typically
    #   .../mail/deeplink/compose?itemid=AAMk...&popoutv2=1
    qs = parse_qs(parsed.query)
    itemid = (qs.get("itemid") or qs.get("ItemID") or [""])[0]

    safe_url = _html.escape(url, quote=True)
    body = f"""<!doctype html>
<html lang="en"><head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Opening Outlook…</title>
<style>
  body{{font-family:-apple-system,Segoe UI,Roboto,sans-serif;text-align:center;padding:48px 20px;color:#333}}
  a{{color:#0078d4;text-decoration:none}}
  iframe{{display:none}}
</style></head>
<body>
<p>Opening Outlook…</p>
<p><a id="webfallback" href="{safe_url}">If nothing happens, tap here to open in Outlook web.</a></p>
<iframe id="launcher" src="about:blank"></iframe>
<script>
  var iframe = document.getElementById("launcher");
  var itemid = {json.dumps(itemid)};
  var webUrl = {json.dumps(url)};
  // iOS Safari blocks iframe-triggered custom schemes for anti-phishing;
  // we have to use top-level navigation (window.location) instead, which
  // limits us to one scheme attempt before falling back to web.
  var isIOS = /iPad|iPhone|iPod/.test(navigator.userAgent);

  function launch(scheme) {{ try {{ iframe.src = scheme; }} catch (e) {{}} }}

  // iOS path: iOS Safari only reliably triggers custom schemes via
  // top-level navigation, but a top-level nav to an unregistered scheme
  // shows a "Cannot Open Page" error and destroys our JS context. So we
  // *probe* with iframe first (silent failure if unsupported — Outlook iOS
  // might happen to accept it via the iframe path even though Mac doesn't),
  // then top-level-navigate to the known-good drafts scheme as fallback,
  // then web.
  if (isIOS) {{
    if (itemid) {{
      launch("ms-outlook://emails/" + encodeURIComponent(itemid));
    }}
    setTimeout(function () {{
      if (document.hidden) return;
      // Known-good fallback — Outlook iOS always accepts this.
      window.location.href = "ms-outlook://drafts";
      setTimeout(function () {{
        if (document.hidden) return;
        window.location.replace(webUrl);
      }}, 1500);
    }}, 400);
    /* don't run the desktop iframe sequence below */
  }} else {{

  // Walk through a list of candidate Outlook URL schemes. The first one the
  // OS recognises hands off to the Outlook app and backgrounds this page;
  // the rest never fire (document.hidden short-circuit). If none are
  // recognised we fall through to ms-outlook://drafts (known-good — lands
  // on the drafts folder), then finally the browser webLink.
  //
  // Microsoft only documents `ms-outlook://compose`, `drafts`, `inbox`,
  // `search`, etc — there's no public per-item scheme. The first six below
  // are undocumented guesses based on community reports; harmless if
  // unsupported.
  var candidates = [];
  if (itemid) {{
    var enc = encodeURIComponent(itemid);
    // Classic Microsoft Outlook URL handler — pre-Acompli, designed
    // specifically for "open this item in its own popup window".
    candidates.push("outlook:"        + enc);
    candidates.push("outlook://"      + enc);
    candidates.push("outlook://open/" + enc);
    // ms-outlook scheme variants — undocumented per-item endpoints.
    candidates.push("ms-outlook://emails/"  + enc);
    candidates.push("ms-outlook://email/"   + enc);
    candidates.push("ms-outlook://message/" + enc);
    candidates.push("ms-outlook://item/"    + enc);
    candidates.push("ms-outlook://mail/"    + enc);
    candidates.push("ms-outlook://drafts/"  + enc);
    candidates.push("ms-outlook://show?id="    + enc);
    candidates.push("ms-outlook://open?id="    + enc);
    candidates.push("ms-outlook://view?id="    + enc);
    candidates.push("ms-outlook://compose?id=" + enc);
  }}
  candidates.push("ms-outlook://drafts");           // safe fallback — opens folder

  var step = 0;
  function next() {{
    if (document.hidden) return;            // app took over — stop here
    if (step >= candidates.length) {{
      // Nothing accepted; open in browser as last resort.
      window.location.replace(webUrl);
      return;
    }}
    launch(candidates[step++]);
    setTimeout(next, 300);
  }}
  next();
  }} /* end !isIOS desktop branch */
</script>
</body></html>"""
    from fastapi.responses import HTMLResponse
    return HTMLResponse(content=body, status_code=200)


# ── Startup ───────────────────────────────────────────────

_frontend_dist = Path(__file__).parent.parent / "frontend" / "dist"
if _frontend_dist.exists():
    app.mount("/", StaticFiles(directory=str(_frontend_dist), html=True), name="static")


def _detect_stuck_init_chains():
    """Watchdog: if any user's init chain stops emitting progress for too long,
    mark the in_progress step as failed so the UI can move on. Defense in depth
    against any Gemini call or other operation that escapes our timeout coverage.

    Triggers every 60 seconds. Considers an init chain stuck when:
      - stage == 'generating'
      - last_update on init_status.json is older than 5 minutes
      - at least one step is still in_progress

    On detection: marks each in_progress step as failed (with a diagnostic
    message), bumps stage to 'draft_ready' so the frontend's Enter button
    enables, and lets the user choose to confirm or restart. Does NOT touch
    any other state — partial CRM data / signatures / etc. stay intact for
    a subsequent restart to use.
    """
    from datetime import datetime, timezone
    from src.modules.profile import load_init_status, update_init_step, save_init_status

    if not auth.DATA_DIR.exists():
        return
    cutoff_ts = datetime.now(timezone.utc).timestamp() - 300  # 5 minutes

    for udir in auth.DATA_DIR.iterdir():
        if not udir.is_dir() or udir.name.startswith("_") or udir.name.startswith("."):
            continue
        try:
            status = load_init_status(udir)
            if status.get("stage") != "generating":
                continue
            last_update_str = status.get("last_update") or ""
            try:
                last_ts = datetime.fromisoformat(last_update_str.replace("Z", "+00:00")).timestamp()
            except Exception:
                continue
            if last_ts >= cutoff_ts:
                continue  # recent enough — actively progressing

            in_progress = [s for s in status.get("steps", []) if s.get("status") == "in_progress"]
            if not in_progress:
                continue  # already past whatever was running

            print(f"[StuckDetector] {udir.name}: init chain silent for "
                  f"{int((datetime.now(timezone.utc).timestamp() - last_ts) // 60)}min — "
                  f"failing {len(in_progress)} step(s): "
                  f"{[s['key'] for s in in_progress]}")
            msg = ("No progress in 5+ minutes — likely a Gemini hang or "
                   "network issue. Restart onboarding to retry.")
            for step in in_progress:
                update_init_step(udir, step["key"], "failed", msg)
            save_init_status(udir, "draft_ready",
                             current_message="Watchdog: init aborted after silent timeout")
        except Exception as e:
            print(f"[StuckDetector] error checking {udir.name}: {e}")


@app.on_event("startup")
def startup_event():
    # Clean up any sections left "running" by the previous (killed) process
    # — must run before scheduler starts so a fresh scheduled fire sees a
    # clean state.
    _reset_stuck_running_results()
    _scheduler.start()
    # Ops: record per-job success/failure timing for the fleet-health endpoint.
    _scheduler.add_listener(_record_job_event, EVENT_JOB_EXECUTED | EVENT_JOB_ERROR)
    _scheduler.add_job(
        _detect_stuck_init_chains,
        trigger="interval",
        seconds=60,
        id="init_stuck_detector",
        replace_existing=True,
        max_instances=1,
        coalesce=True,
    )
    _scheduler.add_job(
        _poll_teams_bot_all_users,
        trigger="interval",
        seconds=10,
        id="teams_bot_poll",
        replace_existing=True,
        max_instances=1,
    )
    _scheduler.add_job(
        _refresh_crm_all_users,
        trigger="cron",
        hour=6, minute=0,
        id="crm_daily_refresh",
        replace_existing=True,
        max_instances=1,
    )
    _scheduler.add_job(
        _refresh_projects_all_users,
        trigger="cron",
        hour=6, minute=30,
        id="projects_daily_refresh",
        replace_existing=True,
        max_instances=1,
    )
    _scheduler.add_job(
        _refresh_companies_all_users,
        trigger="cron",
        hour=6, minute=45,
        id="companies_daily_refresh",
        replace_existing=True,
        max_instances=1,
    )
    _scheduler.add_job(
        _poll_email_monitor_all_users,
        trigger="interval",
        minutes=1,
        id="email_monitor_poll",
        replace_existing=True,
        max_instances=1,
        coalesce=True,
    )
    _scheduler.add_job(
        _poll_expense_scan_all_users,
        trigger="interval",
        minutes=1,
        id="expense_scan_poll",
        replace_existing=True,
        max_instances=1,
        coalesce=True,
    )
    _scheduler.add_job(
        _poll_meeting_prep_all_users,
        trigger="interval",
        minutes=5,
        id="meeting_prep_poll",
        replace_existing=True,
        max_instances=1,
        coalesce=True,
    )
    _scheduler.add_job(
        _weekly_cleanup_scan_all_users,
        trigger="cron",
        day_of_week="mon", hour=7, minute=0,
        id="db_cleanup_weekly",
        replace_existing=True,
        max_instances=1,
    )
    _scheduler.add_job(
        _poll_meeting_recordings_all_users,
        trigger="interval",
        minutes=20,
        id="meeting_recordings_poll",
        replace_existing=True,
        max_instances=1,
        coalesce=True,
    )
    # Ops: synthetic canary probe (no-op unless CANARY_UID is set) — exercises
    # token→Graph for one account every 15 min so breakage surfaces end-to-end.
    _scheduler.add_job(
        _run_canary_probe,
        trigger="interval",
        minutes=15,
        id="canary_probe",
        replace_existing=True,
        max_instances=1,
        coalesce=True,
    )
    # Ops: push-quality spot-check (gated OFF by default — set PUSH_QA_ENABLED=true
    # to turn on; it makes Gemini judge/grounding calls and costs quota).
    _scheduler.add_job(
        _poll_push_qa_all_users,
        trigger="interval",
        minutes=60,
        id="push_qa_poll",
        replace_existing=True,
        max_instances=1,
        coalesce=True,
    )
    _heal_zombie_bot_flags_at_startup()
    _load_all_user_schedules_at_startup()
    print("[Scheduler] Teams bot 10s | Email monitor 1m | Expense scan 1m | Meeting prep 5m | Meeting recordings 20m | CRM refresh daily 06:00 UTC | Projects refresh daily 06:30 UTC | Companies refresh daily 06:45 UTC | DB cleanup weekly Mon 07:00 UTC")
